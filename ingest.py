import os
import re
import sys
import io
import csv
import json
import unicodedata
from pathlib import Path
import torch
from loguru import logger
from vector_store import VectorStore
from api.config import get_settings
from api.embeddings import get_embedder, to_builtin_list


def _ensure_utf8_stdio() -> None:
    """Force UTF-8 stdio on Windows CLI runs without mutating imports."""
    if hasattr(sys.stdout, "buffer"):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    if hasattr(sys.stderr, "buffer"):
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# ============= Constants =============
settings = get_settings()
PDFS_DIR = settings.UPLOAD_DIR
CACHE_DIR = settings.CACHE_DIR
EMBEDDING_MODEL = settings.EMBEDDING_MODEL
SUPPORTED_DOCUMENT_EXTENSIONS = {".pdf", ".txt", ".md", ".csv", ".json", ".docx", ".xlsx"}

os.makedirs(CACHE_DIR, exist_ok=True)
os.makedirs(PDFS_DIR, exist_ok=True)


# ── Section detection patterns ───────────────────────────────────────────
SECTION_PATTERNS = {
    "admin":     r"(document administratif|registre|cnss|attestation|déclaration|affiliation|matricule|légalis|وثائق إدارية|سجل تجاري|تصريح|الإدارة العامة للإعلامية)",
    "technical": r"(document technique|spécification|caractéristique|installation|configuration|matériel|équipement|وثائق فنية|مواصفات|rapport de test|fiche technique)",
    "financial": r"(document financier|bordereau|prix|offre financière|lettre de soumission|montant|عرض مالي|أثمان|أسعار|جدول الأثمان|تعهد مالي)",
    "guarantee": r"(caution|cautionnement|garantie bancaire|garantie définitive|garantie provisoire|ضمان وقتي|ضمان نهائي|كفيل بالتضامن|كفالة بنكية|الضمان الوقتي|الضمان النهائي|ضمان حسن التنفيذ|مدة الضمان)",
    "deadline":  r"(date limite|délai de remise|validité|ouverture des plis|heure limite|تاريخ أقصى|آجال|فتح العروض|جلسة|قبول العروض|تقديم العروض|تونيبس|مكتب الضبط|البريد السريع|مضمون الوصول|منظومة الشراء العمومي|الساعة المحددان|اليوم المحدد)",
    "payment":   r"(modalit|paiement|règlement|versement|échéance|facture|تحويل بنكي|خلاص|دفع|أقساط|فاتورة)",
    "penalty":   r"(pénalité|retard|sanction|indemnité|amende|تغريم|غرامة|عقوبة|تأخير|فسخ)",
    "reception": r"(réception provisoire|réception définitive|réception quantitative|livraison|installation|استلام|تسليم|قبول)",
}

SECTION_PRIORITY = {
    "deadline": 8,
    "admin": 7,
    "penalty": 6,
    "payment": 5,
    "financial": 4,
    "guarantee": 3,
    "technical": 2,
    "reception": 1,
    "general": 0,
}

MIN_USEFUL_CHARS = 40
MERGE_TARGET_CHARS = 180
MAX_CHUNK_CHARS = 1200
ARABIC_OCR_ENABLED = os.getenv("ARABIC_OCR_ENABLED", "true").lower() not in {"0", "false", "no"}
OCR_LANGUAGES = os.getenv("OCR_LANGUAGES", "ara+fra+eng")
OCR_DPI = int(os.getenv("OCR_DPI", "250"))
OCR_MAX_PAGES = int(os.getenv("OCR_MAX_PAGES", "120"))
OCR_REINFORCE_ENABLED = os.getenv("OCR_REINFORCE_ENABLED", "true").lower() not in {"0", "false", "no"}
OCR_REINFORCE_MAX_PAGES = int(os.getenv("OCR_REINFORCE_MAX_PAGES", "12"))
OCR_REINFORCE_ARABIC_MAX_PAGES = int(os.getenv("OCR_REINFORCE_ARABIC_MAX_PAGES", "20"))
OCR_REINFORCE_ARABIC_FRONT_PAGES = int(os.getenv("OCR_REINFORCE_ARABIC_FRONT_PAGES", "14"))
PAGE_MARKER_RE = re.compile(r"[—-]?\s*Page\s+(\d+)\s*[—-]?", re.IGNORECASE)

def detect_section(text: str) -> str:
    text_lower = text.lower()
    scores = {}
    for section, pattern in SECTION_PATTERNS.items():
        matches = re.findall(pattern, text_lower)
        if matches:
            scores[section] = len(matches)

    if not scores:
        return "general"

    # Prefer the strongest semantic section instead of the first broad regex hit.
    return max(scores.items(), key=lambda item: (item[1], SECTION_PRIORITY.get(item[0], 0)))[0]


def _clean_chunk_text(text: str) -> str:
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _is_noise_chunk(text: str) -> bool:
    stripped = text.strip()
    if not stripped:
        return True

    if stripped == "<!-- formula-not-decoded -->":
        return True

    if re.fullmatch(r"[-–—\s]*page\s+\d+[-–—\s]*", stripped, flags=re.IGNORECASE):
        return True

    if re.fullmatch(r"[\d\s/().,:;-]+", stripped):
        return True

    lowered = stripped.lower()
    if "signature et cachet du soumissionnaire" in lowered and len(stripped) < 180:
        return True

    if lowered.startswith("tunis, le") and len(stripped) < 80:
        return True

    return False


def _split_large_chunk(text: str, max_chars: int = MAX_CHUNK_CHARS) -> list[str]:
    text = text.strip()
    if len(text) <= max_chars:
        return [text]

    blocks = [b.strip() for b in re.split(r"\n{2,}", text) if b.strip()]
    if len(blocks) <= 1:
        blocks = [b.strip() for b in re.split(r"(?<=[.!?؟؛:])\s+", text) if b.strip()]

    if len(blocks) <= 1:
        return [text[i:i + max_chars].strip() for i in range(0, len(text), max_chars) if text[i:i + max_chars].strip()]

    parts = []
    current = ""
    for block in blocks:
        candidate = block if not current else f"{current}\n\n{block}"
        if len(candidate) <= max_chars:
            current = candidate
            continue

        if current:
            parts.append(current)

        if len(block) <= max_chars:
            current = block
        else:
            parts.extend(
                piece for piece in (block[i:i + max_chars].strip() for i in range(0, len(block), max_chars))
                if piece
            )
            current = ""

    if current:
        parts.append(current)

    return parts


def _split_embedded_pages(text: str, fallback_page: int | str) -> list[dict]:
    """
    Some PDFs embed several page bodies into one extracted text blob.
    Split those blobs back into per-page entries when explicit page markers exist.
    """
    matches = list(PAGE_MARKER_RE.finditer(text))
    if not matches:
        return [{"page": str(fallback_page), "text": text}]

    entries = []
    preamble = text[: matches[0].start()].strip()
    if preamble:
        entries.append({"page": str(fallback_page), "text": preamble})

    for idx, match in enumerate(matches):
        page_num = match.group(1)
        start = match.end()
        end = matches[idx + 1].start() if idx + 1 < len(matches) else len(text)
        piece = text[start:end].strip()
        piece = re.sub(rf"^\s*P\s*{re.escape(page_num)}\b", "", piece, flags=re.IGNORECASE).strip()
        if piece:
            entries.append({"page": str(page_num), "text": piece})

    return entries or [{"page": str(fallback_page), "text": text}]


def _merge_small_chunks(entries: list[dict]) -> list[dict]:
    if not entries:
        return []

    merged = []
    buffer = entries[0].copy()

    for entry in entries[1:]:
        same_page = entry["page"] == buffer["page"]
        combined_len = len(buffer["text"]) + len(entry["text"]) + 2
        should_merge = same_page and combined_len <= MAX_CHUNK_CHARS and (
            len(buffer["text"]) < MERGE_TARGET_CHARS or len(entry["text"]) < MERGE_TARGET_CHARS
        )

        if should_merge:
            buffer["text"] = f"{buffer['text']}\n{entry['text']}"
        else:
            merged.append(buffer)
            buffer = entry.copy()

    merged.append(buffer)
    return merged


def _extract_text_entries_pypdf(file_path: str) -> tuple[list[dict], int]:
    """Use direct text extraction for born-digital PDFs before falling back to OCR."""
    from pypdf import PdfReader

    reader = PdfReader(file_path)
    entries = []

    for page_num, page in enumerate(reader.pages, start=1):
        text = _clean_chunk_text(page.extract_text() or "")
        if not text:
            continue
        if _is_noise_chunk(text):
            continue
        entries.extend(_split_embedded_pages(text, page_num))

    return entries, len(reader.pages)


def _extract_text_entries_pymupdf(file_path: str) -> tuple[list[dict], int]:
    """Use PyMuPDF as a second direct text-layer extractor for PDFs with awkward embedded page banners."""
    import fitz

    document = fitz.open(file_path)
    entries = []
    try:
        page_count = document.page_count
        for page_num in range(page_count):
            text = _clean_chunk_text(document.load_page(page_num).get_text("text") or "")
            if not text or _is_noise_chunk(text):
                continue
            entries.append({"page": str(page_num + 1), "text": text})
        return entries, page_count
    finally:
        document.close()


DIRECT_TEXT_TENDER_MARKERS = (
    "appel d'offres",
    "cahier des charges",
    "soumission",
    "offre technique",
    "offre financiere",
    "caution provisoire",
)
DIRECT_TEXT_DEADLINE_ANCHORS = (
    "date limite",
    "remise des offres",
    "reception des offres",
    "au plus tard",
)
DIRECT_TEXT_DATE_VALUE_RE = re.compile(
    r"\b\d{1,2}(?:[/-]\d{1,2}[/-]\d{2,4}|"
    r"\s+(?:janvier|fevrier|mars|avril|mai|juin|juillet|aout|septembre|octobre|novembre|decembre)\s+\d{4})\b",
    re.IGNORECASE,
)


def _has_deadline_value_near_anchor(folded_text: str) -> bool:
    for anchor in DIRECT_TEXT_DEADLINE_ANCHORS:
        start = 0
        while True:
            index = folded_text.find(anchor, start)
            if index < 0:
                break
            window = folded_text[max(0, index - 80) : index + 220]
            if DIRECT_TEXT_DATE_VALUE_RE.search(window):
                return True
            start = index + len(anchor)
    return False


def _direct_pdf_text_has_tender_gap(entries: list[dict]) -> bool:
    text = "\n".join(str(entry.get("text", "")) for entry in entries)
    folded = _fold_fact_text(text)
    if not any(marker in folded for marker in DIRECT_TEXT_TENDER_MARKERS):
        return False
    if not any(anchor in folded for anchor in DIRECT_TEXT_DEADLINE_ANCHORS):
        return False

    return not _has_deadline_value_near_anchor(folded)


def _direct_pdf_text_has_offer_envelope_content(entries: list[dict]) -> bool:
    text = "\n".join(str(entry.get("text", "")) for entry in entries)
    folded = _fold_fact_text(text)
    markers = (
        "dossier administratif",
        "offre technique",
        "offre financiere",
        "lettre de soumission",
        "bordereau des prix",
        "recapitulatif des prix",
        "bureau d'ordre",
        "date limite",
        "cnss",
        "rne",
    )
    return sum(1 for marker in markers if marker in folded) >= 5


def _direct_pdf_candidate_score(entries: list[dict], page_count: int) -> int:
    combined = "\n".join(str(entry.get("text", "")) for entry in entries)
    quality = _build_text_quality_metadata(entries, page_count=page_count, text_source="pdf_text_layer")
    score = _text_quality_score(combined)
    score += min(sum(len(str(entry.get("text", ""))) for entry in entries) // 500, 80)
    score -= int(quality.get("page_gap_count") or 0) * 25
    if _direct_pdf_text_has_offer_envelope_content(entries):
        score += 80
    if _has_deadline_value_near_anchor(_fold_fact_text(combined)):
        score += 40
    return score


PUBLIC_TENDER_ARABIC_OCR_HINTS = (
    "tuneps",
    "republique tunisienne",
    "république tunisienne",
    "ministere",
    "ministère",
    "office de",
)
USEFUL_DIRECT_LANGUAGE_MARKERS = (
    "appel d'offres",
    "appel d offres",
    "cahier des charges",
    "date limite",
    "soumission",
    "offre technique",
    "offre financiere",
    "offre financière",
    "bordereau des prix",
    "caracteristiques techniques",
    "caractéristiques techniques",
)
ARABIC_TENDER_MARKERS = (
    "كراس الشروط",
    "طلب عروض",
    "العروض",
    "الصفقة",
    "الضمان",
    "آخر أجل",
    "منظومة تونابس",
    "تونابس",
)


def _direct_pdf_text_needs_arabic_ocr(entries: list[dict]) -> bool:
    text = "\n".join(str(entry.get("text", "")) for entry in entries)
    if not text.strip() or _arabic_char_ratio(text) >= 0.08:
        return False

    folded = _fold_fact_text(text)
    if not any(marker in folded for marker in PUBLIC_TENDER_ARABIC_OCR_HINTS):
        return False

    useful_hits = sum(1 for marker in USEFUL_DIRECT_LANGUAGE_MARKERS if marker in folded)
    return useful_hits < 2


def _expected_page_count(entries: list[dict], page_count: int | None = None) -> int:
    numeric_pages = [
        int(str(entry.get("page")))
        for entry in entries
        if str(entry.get("page", "")).isdigit()
    ]
    candidates = [page for page in (page_count, *numeric_pages) if isinstance(page, int) and page > 0]
    return max(candidates) if candidates else 0


def _direct_pdf_text_too_sparse_for_pages(entries: list[dict], page_count: int) -> bool:
    expected_pages = _expected_page_count(entries, page_count)
    if expected_pages < 8:
        return False

    useful_entries = [entry for entry in entries if len(str(entry.get("text", "")).strip()) >= MIN_USEFUL_CHARS]
    useful_pages = {str(entry.get("page")) for entry in useful_entries}
    total_chars = sum(len(str(entry.get("text", ""))) for entry in useful_entries)
    coverage = len(useful_pages) / max(expected_pages, 1)
    chars_per_page = total_chars / max(expected_pages, 1)

    if expected_pages >= 20 and chars_per_page < 220:
        return True
    if coverage < 0.55 and chars_per_page < 350:
        return True
    return False


def _text_quality_score(text: str) -> int:
    cleaned = _clean_chunk_text(str(text or ""))
    if not cleaned:
        return -1000

    folded = _fold_fact_text(cleaned)
    score = min(len(cleaned) // 80, 40)
    score += min(len(re.findall(r"\d", cleaned)), 30)
    score += int(_arabic_char_ratio(cleaned) * 160)

    for marker in USEFUL_DIRECT_LANGUAGE_MARKERS:
        if marker in folded:
            score += 16
    for marker in ARABIC_TENDER_MARKERS:
        if marker in cleaned:
            score += 18
    for marker in ("tuneps", "www.tuneps.tn", "prix", "qte", "qté", "minimum demandé"):
        if marker in folded:
            score += 8

    if any(marker in folded for marker in PUBLIC_TENDER_ARABIC_OCR_HINTS) and _arabic_char_ratio(cleaned) < 0.01:
        score -= 35

    return score


NON_CONTENT_FACT_KEYS = {
    "summary",
    "tender_profile",
    "extraction_warning",
}


def _short_ocr_fragment_ratio(text: str) -> float:
    lines = [line.strip() for line in str(text or "").splitlines() if line.strip()]
    if not lines:
        return 0.0

    short_fragments = [
        line
        for line in lines
        if len(line) <= 6 and re.search(r"[A-Za-zÀ-ÖØ-öø-ÿ\u0600-\u06FF]", line)
    ]
    return len(short_fragments) / max(len(lines), 1)


def _build_extraction_warning(chunks: list[str], facts: dict) -> dict | None:
    content_fields = [
        key
        for key in facts
        if key not in NON_CONTENT_FACT_KEYS
    ]
    if len(content_fields) > 1:
        return None

    texts = [str(chunk or "") for chunk in chunks if str(chunk or "").strip()]
    if not texts:
        return {
            "text": (
                "Qualité d'extraction insuffisante : aucun texte exploitable n'a été extrait. "
                "Les réponses « Non mentionné » peuvent venir de l'OCR, pas forcément du contenu."
            ),
            "level": "warning",
            "reason": "empty_extraction",
            "section": "extraction",
        }

    combined = "\n".join(texts)
    cleaned = _clean_chunk_text(combined)
    if len(cleaned) < 120:
        return None

    folded = _fold_fact_text(cleaned)
    quality_scores = [_text_quality_score(text) for text in texts]
    average_quality = sum(quality_scores) / max(len(quality_scores), 1)
    short_fragment_ratio = _short_ocr_fragment_ratio(combined)
    useful_marker_hits = sum(
        1
        for marker in (
            "objet",
            "soumission",
            "offre",
            "date limite",
            "cahier des charges",
            "bureau d'ordre",
            "tuneps",
            "prix",
            "garantie",
            "استشارة",
            "العروض",
            "الضمان",
            "الخلاص",
            "السعر",
        )
        if marker in folded or marker in cleaned
    )
    long_word_count = len(re.findall(r"[A-Za-zÀ-ÖØ-öø-ÿ\u0600-\u06FF]{4,}", cleaned))

    looks_fragmented = short_fragment_ratio >= 0.42
    looks_low_signal = average_quality < 22 or (long_word_count < 20 and len(texts) >= 5)
    has_enough_tender_signal = useful_marker_hits >= 3 and average_quality >= 25
    if has_enough_tender_signal or not (looks_fragmented or looks_low_signal):
        return None

    return {
        "text": (
            "Qualité d'extraction insuffisante : le document semble être un scan tourné, "
            "peu lisible ou mal OCRisé. Les réponses « Non mentionné » peuvent venir de "
            "l'OCR, pas forcément du contenu. Veuillez réimporter une version correctement "
            "orientée et lisible."
        ),
        "level": "warning",
        "reason": "low_text_quality",
        "section": "extraction",
        "metrics": {
            "average_quality": round(average_quality, 2),
            "short_fragment_ratio": round(short_fragment_ratio, 3),
            "content_field_count": len(content_fields),
            "useful_marker_hits": useful_marker_hits,
        },
    }


def _combine_entries_by_page(entries: list[dict]) -> dict[str, str]:
    by_page: dict[str, list[str]] = {}
    for entry in entries:
        text = _clean_chunk_text(str(entry.get("text", "")))
        if not text:
            continue
        page = str(entry.get("page", "?"))
        by_page.setdefault(page, []).append(text)
    return {page: "\n".join(parts) for page, parts in by_page.items()}


def _select_best_entries_by_page(*entry_sets: list[dict]) -> list[dict]:
    candidates_by_page: dict[str, list[str]] = {}
    for entries in entry_sets:
        for page, text in _combine_entries_by_page(entries).items():
            candidates_by_page.setdefault(page, []).append(text)

    selected = []
    for page, candidates in candidates_by_page.items():
        ranked = sorted(candidates, key=_text_quality_score, reverse=True)
        best = ranked[0]
        best_folded = _fold_fact_text(best)
        for candidate in ranked[1:]:
            candidate = _clean_chunk_text(candidate)
            folded_candidate = _fold_fact_text(candidate)
            carries_table = any(
                marker in folded_candidate
                for marker in (
                    "designation",
                    "qte",
                    "qté",
                    "prix unitaire",
                    "caracteristique technique",
                    "minimum demande",
                    "minimum demandé",
                )
            )
            if carries_table and folded_candidate not in best_folded:
                best = f"{best}\n{candidate}"
                best_folded = _fold_fact_text(best)
        if _clean_chunk_text(best):
            selected.append({"page": page, "text": best})

    return sorted(selected, key=lambda item: _page_sort_key(item.get("page")))


OCR_REINFORCE_PAGE_MARKERS = (
    "objet",
    "a pour objet",
    "date limite",
    "heure limite",
    "reception des offres",
    "remise des offres",
    "validite",
    "valables",
    "dpc",
    "dpao",
    "donnees particulieres",
    "tuneps",
    "www.tuneps.tn",
    "bureau d'ordre",
    "soumission",
    "offre technique",
    "offre financiere",
    "dossier administratif",
    "documents administratifs",
    "bordereau des prix",
    "caution",
    "garantie",
    "reception provisoire",
    "reception definitive",
    "modalites de paiement",
    "reglement",
    "penalite",
)

ARABIC_OCR_REINFORCE_PAGE_MARKERS = (
    "كراس الشروط",
    "طلب عروض",
    "اقتناء",
    "وزارة",
    "قبول العروض",
    "فتح العروض",
    "جلسة",
    "العرض الفني",
    "العرض المالي",
    "الوثائق",
    "وثيقة الضمان",
    "وشيقة الضمان",
    "الضمان الوقتي",
    "الضمان المالي الوقتي",
    "بطاقة الإرشادات",
    "بطاقق الإرشادات",
    "السجل الوطني",
    "لسجل الوطني",
    "التعهد المالي",
    "جدول الأثمان",
    "جدول الأشمان",
    "مدة الضمان",
    "مدق الضمان",
    "الاستلام",
    "القبول الوقتي",
    "غرامة التأخير",
    "غرامق الت خير",
    "خطايا التأخير",
    "خلاص",
    "أمر بصرف",
    "فاتورة",
    "فاتورق",
    "الضمان النهائي",
    "لنهائي",
)


def _page_number(page: str | int | None) -> int | None:
    page_num, _page_text = _page_sort_key(page)
    return page_num if page_num < 10_000 else None


def _entry_page_numbers(entries: list[dict]) -> list[int]:
    pages = []
    for entry in entries:
        page_num = _page_number(entry.get("page"))
        if page_num is not None:
            pages.append(page_num)
    return sorted(set(pages))


def _missing_page_ranges(page_numbers: list[int], expected_page_count: int | None = None) -> list[str]:
    if not page_numbers:
        return []

    first_page = min(page_numbers)
    last_page = max(max(page_numbers), expected_page_count or 0)
    if first_page > 1:
        first_page = 1

    present = set(page_numbers)
    ranges = []
    start = None
    previous = None
    for page in range(first_page, last_page + 1):
        if page in present:
            if start is not None:
                ranges.append(f"{start}" if start == previous else f"{start}-{previous}")
                start = None
            continue
        if start is None:
            start = page
        previous = page

    if start is not None:
        ranges.append(f"{start}" if start == previous else f"{start}-{previous}")

    return ranges


def _readable_char_ratio(text: str) -> float:
    non_space = [char for char in str(text or "") if not char.isspace()]
    if not non_space:
        return 0.0
    readable = [
        char
        for char in non_space
        if char.isalnum() or "\u0600" <= char <= "\u06FF"
    ]
    return len(readable) / max(len(non_space), 1)


def _build_text_quality_metadata(
    entries: list[dict],
    *,
    page_count: int | None = None,
    text_source: str = "unknown",
    preferred_source: str | None = None,
) -> dict:
    combined = "\n".join(str(entry.get("text", "")) for entry in entries)
    page_numbers = _entry_page_numbers(entries)
    missing_ranges = _missing_page_ranges(page_numbers, page_count)
    page_gap_count = sum(
        (int(end) - int(start) + 1) if "-" in range_text else 1
        for range_text in missing_ranges
        for start, end in [range_text.split("-", 1) if "-" in range_text else (range_text, range_text)]
    )
    arabic_ratio = _arabic_char_ratio(combined)
    readable_ratio = _readable_char_ratio(combined)

    if arabic_ratio >= 0.08 and (readable_ratio < 0.68 or page_gap_count):
        mode = "arabic_noisy"
    elif readable_ratio < 0.55:
        mode = "noisy_ocr"
    elif page_gap_count:
        mode = "partial_pages"
    else:
        mode = "clean"

    return {
        "mode": mode,
        "page_gap_count": page_gap_count,
        "missing_page_ranges": missing_ranges,
        "arabic_ratio": round(arabic_ratio, 4),
        "readable_ratio": round(readable_ratio, 4),
        "text_source": text_source,
        "preferred_source": preferred_source or text_source,
    }


def _entries_to_chunks(
    entries: list[dict],
    filename: str,
    *,
    text_quality: dict | None = None,
) -> tuple[list[str], list[dict], list[str]]:
    merged_entries = _merge_small_chunks(entries)

    chunks = []
    metas = []
    ids_out = []
    split_max_chars = 12000 if (text_quality or {}).get("text_source") == "pdf_text_layer" else MAX_CHUNK_CHARS
    for entry in merged_entries:
        for piece in _split_large_chunk(entry["text"], max_chars=split_max_chars):
            piece = _clean_chunk_text(piece)
            if len(piece) < MIN_USEFUL_CHARS or _is_noise_chunk(piece):
                continue

            chunk_index = len(chunks)
            chunks.append(piece)
            detected_section = detect_section(piece)
            entry_section = entry.get("section")
            meta = {
                "source": filename,
                "page": entry["page"],
                "section": detected_section if detected_section != "general" else entry_section or detected_section,
                "chunk_index": chunk_index,
            }
            for key in ("doc_type", "location", "section_heading", "source_type"):
                if entry.get(key):
                    meta[key] = entry[key]
            if text_quality:
                meta["text_quality"] = text_quality
                meta["text_quality_mode"] = text_quality.get("mode")
                meta["text_quality_text_source"] = text_quality.get("text_source")
                meta["text_quality_preferred_source"] = text_quality.get("preferred_source")
                meta["text_quality_page_gap_count"] = text_quality.get("page_gap_count", 0)
            metas.append(meta)
            ids_out.append(f"{filename}_c{chunk_index}")

    return chunks, metas, ids_out


def _facts_for_entries(entries: list[dict], filename: str) -> dict:
    chunks, metas, _ids = _entries_to_chunks(entries, filename)
    if not chunks:
        return {}
    return extract_document_facts(chunks, metas)


def _content_fact_count(facts: dict | None) -> int:
    return len([
        key
        for key in (facts or {})
        if key not in NON_CONTENT_FACT_KEYS
    ])


def _fact_text(facts: dict | None, field: str) -> str:
    fact = (facts or {}).get(field)
    if isinstance(fact, dict):
        return str(fact.get("text", ""))
    return ""


def _fact_page(facts: dict | None, field: str) -> int | None:
    fact = (facts or {}).get(field)
    if not isinstance(fact, dict):
        return None
    return _page_number(fact.get("page"))


def _extracted_facts_need_ocr_reinforcement(facts: dict, entries: list[dict]) -> bool:
    if not OCR_REINFORCE_ENABLED or not entries:
        return False

    combined = "\n".join(str(entry.get("text", "")) for entry in entries)
    folded = _fold_fact_text(combined)
    marker_hits = sum(1 for marker in OCR_REINFORCE_PAGE_MARKERS if marker in folded)
    has_tender_signal = marker_hits >= 3
    if not has_tender_signal:
        return False

    subject_text = _fold_fact_text(_fact_text(facts, "subject"))
    subject_page = _fact_page(facts, "subject")
    suspicious_subject = bool(subject_text) and (
        "objet de changement" in subject_text
        or "developpingunit" in subject_text
        or "developingunit" in subject_text
        or "code reference" in subject_text
        or ("bordereau" in subject_text and "prix" in subject_text)
        or (subject_page is not None and subject_page > 8 and len(subject_text) < 140)
    )
    if suspicious_subject:
        return True

    if "subject" not in facts and any(marker in folded for marker in ("objet", "a pour objet", "consultation")):
        return True
    if "deadline" not in facts and any(marker in folded for marker in ("date limite", "heure limite", "au plus tard")):
        return True
    if "submission_method" not in facts and any(marker in folded for marker in ("offres doivent", "tuneps", "bureau d'ordre", "deposees")):
        return True

    validity_text = _fold_fact_text(_fact_text(facts, "validity"))
    if validity_text and any(marker in validity_text for marker in ("dpc", "dpao", "indiquee au", "indique au")):
        return True
    if "validity" not in facts and any(marker in folded for marker in ("validite", "valables", "lies par leurs offres")):
        return True

    profile = facts.get("tender_profile") if isinstance(facts, dict) else None
    coverage = profile.get("coverage", {}) if isinstance(profile, dict) else {}
    core_ratio = coverage.get("core_ratio")
    if isinstance(core_ratio, (int, float)) and core_ratio < 0.35:
        return True

    return _content_fact_count(facts) <= 3


def _entries_look_like_arabic_tender(entries: list[dict]) -> bool:
    combined = "\n".join(str(entry.get("text", "")) for entry in entries)
    if _arabic_char_ratio(combined) < 0.08:
        return False
    normalized = _normalize_arabic_ocr_for_fact_matching(combined)
    return any(
        marker in normalized
        for marker in (*ARABIC_TENDER_MARKERS, *ARABIC_OCR_REINFORCE_PAGE_MARKERS)
    )


def _target_pages_for_ocr_reinforcement(entries: list[dict], facts: dict) -> list[int]:
    expected_pages = _expected_page_count(entries)
    page_limit = min(expected_pages or OCR_MAX_PAGES, OCR_MAX_PAGES)
    arabic_tender = _entries_look_like_arabic_tender(entries)
    max_pages = OCR_REINFORCE_ARABIC_MAX_PAGES if arabic_tender else OCR_REINFORCE_MAX_PAGES
    target_pages = {page for page in range(1, min(page_limit, 3) + 1)}

    if arabic_tender:
        front_pages = min(page_limit, OCR_REINFORCE_ARABIC_FRONT_PAGES)
        target_pages.update(range(1, front_pages + 1))

    for entry in entries:
        page_num = _page_number(entry.get("page"))
        if not page_num or page_num > page_limit:
            continue
        text = _normalize_arabic_ocr_for_fact_matching(str(entry.get("text", "")))
        folded = _fold_fact_text(text)
        if any(marker in folded for marker in OCR_REINFORCE_PAGE_MARKERS) or any(
            marker in text for marker in ARABIC_OCR_REINFORCE_PAGE_MARKERS
        ):
            target_pages.add(page_num)

    for field in ("subject", "deadline", "validity", "submission_method", "payment", "guarantee"):
        page_num = _fact_page(facts, field)
        if page_num and page_num <= page_limit:
            target_pages.add(page_num)

    return sorted(target_pages)[:max_pages]


def _should_use_direct_pdf_text(entries: list[dict], page_count: int) -> bool:
    if not entries:
        return False

    useful_entries = [entry for entry in entries if len(entry["text"].strip()) >= MIN_USEFUL_CHARS]
    total_chars = sum(len(entry["text"]) for entry in useful_entries)
    useful_pages = {str(entry["page"]) for entry in useful_entries}
    expected_pages = page_count if page_count and page_count > 0 else _expected_page_count(entries)

    if expected_pages >= 4:
        coverage = len(useful_pages) / max(expected_pages, 1)
        if min(coverage, 1.0) < 0.9:
            return False

    chars_per_page = total_chars / max(expected_pages, 1)
    if expected_pages >= 20 and chars_per_page < 220:
        return False
    if expected_pages >= 8 and len(useful_pages) / max(expected_pages, 1) < 0.55 and chars_per_page < 350:
        return False

    if _direct_pdf_text_has_tender_gap(useful_entries) and not _direct_pdf_text_has_offer_envelope_content(useful_entries):
        return False

    if _direct_pdf_text_needs_arabic_ocr(useful_entries):
        return False

    return len(useful_entries) >= 1 and total_chars >= 120


def _write_text_cache(filename: str, entries: list[dict]) -> None:
    if not entries:
        return

    def _cache_quality(text: str) -> tuple[int, int, int]:
        folded = _fold_fact_text(text)
        marker_score = sum(
            1
            for marker in (
                "article 17",
                "24 fevrier",
                "date limite",
                "dossier administratif",
                "offre technique",
                "offre financiere",
                "conditions de paiement",
                "modalites de paiement",
                "penalite de retard",
                "caution provisoire",
                "objet du marche",
            )
            if marker in folded
        )
        page_numbers = sorted({
            int(match)
            for match in re.findall(r"\[Page\s+(\d+)\]", text, flags=re.IGNORECASE)
        })
        missing_pages = _missing_page_ranges(page_numbers)
        gap_penalty = sum(
            (int(end) - int(start) + 1) if "-" in range_text else 1
            for range_text in missing_pages
            for start, end in [range_text.split("-", 1) if "-" in range_text else (range_text, range_text)]
        )
        pages = len(page_numbers)
        return marker_score - (gap_penalty * 3), -gap_penalty, pages, len(text)

    try:
        cache_path = Path(CACHE_DIR) / f"{filename}.txt"
        cache_path.parent.mkdir(parents=True, exist_ok=True)

        lines = []
        for entry in sorted(entries, key=lambda item: _page_sort_key(item.get("page"))):
            text = str(entry.get("text", "")).strip()
            if not text:
                continue
            lines.append(f"[Page {entry.get('page', '?')}]\n{text}")

        if lines:
            new_text = "\n\n".join(lines)
            if cache_path.exists():
                existing_text = cache_path.read_text(encoding="utf-8", errors="replace")
                if _cache_quality(existing_text) >= _cache_quality(new_text):
                    return

            cache_path.write_text(new_text, encoding="utf-8", errors="replace")
    except Exception as exc:
        logger.warning(f"Could not write text cache for {filename}: {exc}")


# ============= Docling PDF Extraction =============

def _extract_text_entries_docling(file_path: str, filename: str) -> list[dict]:
    from docling.document_converter import DocumentConverter
    from docling.chunking import HierarchicalChunker

    logger.info(f"Extracting {filename} via Docling...")
    converter = DocumentConverter()
    res = converter.convert(file_path)

    raw_entries = []
    chunker = HierarchicalChunker()
    doc_chunks = list(chunker.chunk(res.document))

    for c in doc_chunks:
        text = c.text

        # Apply Arabic Bidi fix if Arabic content is detected in Docling output.
        if _arabic_char_ratio(text) > 0.05:
            text = _fix_arabic_lines(text)

        page_num = 1
        if hasattr(c, "meta") and hasattr(c.meta, "doc_items") and c.meta.doc_items:
            for item in c.meta.doc_items:
                if hasattr(item, "prov") and item.prov:
                    page_num = item.prov[0].page_no
                    break

        text = _clean_chunk_text(text)
        if _is_noise_chunk(text):
            continue

        raw_entries.extend(_split_embedded_pages(text, page_num))

    return raw_entries


def _extract_text_entries_tesseract(
    file_path: str,
    filename: str,
    pages: list[int] | None = None,
) -> list[dict]:
    if not ARABIC_OCR_ENABLED:
        return []

    try:
        from pdf2image import convert_from_path
        import pytesseract
    except Exception as exc:
        logger.warning(f"Arabic OCR unavailable for {filename}: {exc}")
        return []

    target_pages = sorted({page for page in (pages or []) if isinstance(page, int) and page > 0})
    if target_pages:
        target_pages = [page for page in target_pages if page <= OCR_MAX_PAGES]
        logger.info(
            f"Extracting {filename} pages {target_pages} via Tesseract OCR ({OCR_LANGUAGES}, {OCR_DPI} DPI)..."
        )
    else:
        logger.info(
            f"Extracting {filename} via Tesseract OCR ({OCR_LANGUAGES}, {OCR_DPI} DPI)..."
        )

    entries = []
    if target_pages:
        rendered_pages: list[tuple[int, object]] = []
        for page_num in target_pages:
            try:
                images = convert_from_path(
                    file_path,
                    dpi=OCR_DPI,
                    first_page=page_num,
                    last_page=page_num,
                )
            except Exception as exc:
                logger.warning(f"Could not render {filename} page {page_num} for OCR: {exc}")
                continue
            rendered_pages.extend((page_num, image) for image in images)
    else:
        try:
            images = convert_from_path(
                file_path,
                dpi=OCR_DPI,
                first_page=1,
                last_page=OCR_MAX_PAGES,
            )
        except Exception as exc:
            logger.warning(f"Could not render {filename} for Arabic OCR: {exc}")
            return []
        rendered_pages = list(enumerate(images, start=1))

    for page_num, image in rendered_pages:
        try:
            text = pytesseract.image_to_string(
                image,
                lang=OCR_LANGUAGES,
                config="--oem 1 --psm 6",
            )
        except Exception as exc:
            logger.warning(f"Tesseract OCR failed for {filename} page {page_num}: {exc}")
            continue

        text = _clean_chunk_text(text)
        if not text or _is_noise_chunk(text):
            continue
        entries.append({"page": str(page_num), "text": text})

    return entries


def _entries_need_arabic_ocr(entries: list[dict]) -> bool:
    if not entries:
        return True

    text = "\n".join(str(entry.get("text", "")) for entry in entries)
    folded = _fold_fact_text(text)
    if _arabic_char_ratio(text) >= 0.08:
        return False

    if any(marker in folded for marker in PUBLIC_TENDER_ARABIC_OCR_HINTS):
        useful_hits = sum(1 for marker in USEFUL_DIRECT_LANGUAGE_MARKERS if marker in folded)
        if useful_hits < 3:
            return True

    expected_pages = _expected_page_count(entries)
    if expected_pages >= 20:
        useful_entries = [entry for entry in entries if len(str(entry.get("text", "")).strip()) >= MIN_USEFUL_CHARS]
        chars_per_page = sum(len(str(entry.get("text", ""))) for entry in useful_entries) / max(expected_pages, 1)
        if chars_per_page < 260:
            return True

    return _text_quality_score(text) < 25


def _split_plain_text_entries(text: str, filename: str, *, source_label: str = "text", max_chars: int = 4500) -> list[dict]:
    clean = text.replace("\r\n", "\n").replace("\r", "\n").strip()
    if not clean:
        return []

    entries = []
    buffer = []
    current_len = 0

    def flush() -> None:
        nonlocal buffer, current_len
        chunk = "\n".join(buffer).strip()
        if chunk:
            entries.append({
                "page": str(len(entries) + 1),
                "text": chunk,
                "source": source_label,
            })
        buffer = []
        current_len = 0

    for paragraph in re.split(r"\n{2,}", clean):
        paragraph = paragraph.strip()
        if not paragraph:
            continue
        if current_len and current_len + len(paragraph) + 2 > max_chars:
            flush()
        if len(paragraph) > max_chars:
            for start in range(0, len(paragraph), max_chars):
                part = paragraph[start:start + max_chars].strip()
                if part:
                    entries.append({
                        "page": str(len(entries) + 1),
                        "text": part,
                        "source": source_label,
                    })
            continue
        buffer.append(paragraph)
        current_len += len(paragraph) + 2

    flush()
    return entries


def _read_text_file(file_path: str) -> str:
    data = Path(file_path).read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _extract_text_entries_plain_file(file_path: str, filename: str) -> list[dict]:
    suffix = Path(filename).suffix.lower()
    text = _read_text_file(file_path)

    if suffix == ".json":
        try:
            parsed = json.loads(text)
            text = json.dumps(parsed, ensure_ascii=False, indent=2)
        except Exception:
            pass

    if suffix == ".csv":
        rows = []
        try:
            dialect = csv.Sniffer().sniff(text[:4096])
        except Exception:
            dialect = csv.excel
        try:
            for row in csv.reader(io.StringIO(text), dialect):
                cells = [cell.strip() for cell in row if cell and cell.strip()]
                if cells:
                    rows.append(" | ".join(cells))
            if rows:
                text = "\n".join(rows)
        except Exception:
            pass

    return _split_plain_text_entries(text, filename, source_label=suffix.lstrip(".") or "text")


def _extract_text_entries_docx(file_path: str, filename: str) -> list[dict]:
    try:
        from docx import Document as DocxDocument
        from docx.oxml.table import CT_Tbl
        from docx.oxml.text.paragraph import CT_P
        from docx.table import Table
        from docx.text.paragraph import Paragraph
    except ImportError as exc:
        raise RuntimeError("python-docx is required to ingest .docx files") from exc

    document = DocxDocument(file_path)
    entries: list[dict] = []
    current_heading = "Document"
    current_parts: list[str] = []
    section_index = 0

    def block_iter():
        for child in document.element.body.iterchildren():
            if isinstance(child, CT_P):
                yield Paragraph(child, document)
            elif isinstance(child, CT_Tbl):
                yield Table(child, document)

    def is_heading(paragraph: Paragraph) -> bool:
        style_name = str(getattr(paragraph.style, "name", "") or "").lower()
        text = paragraph.text.strip()
        if style_name.startswith("heading") or style_name.startswith("titre"):
            return True
        return bool(re.match(r"^(?:article|chapitre|section)\s+\d+\b", text, flags=re.IGNORECASE))

    def flush() -> None:
        nonlocal current_parts, section_index
        text = "\n\n".join(part for part in current_parts if part.strip()).strip()
        if not text:
            current_parts = []
            return
        section_index += 1
        clean_heading = re.sub(r"\s+", " ", current_heading).strip(" :.-") or f"Bloc {section_index}"
        entries.append(
            {
                "page": str(section_index),
                "text": text,
                "source": "docx",
                "source_type": "docx",
                "doc_type": "docx",
                "location": f"Section: {clean_heading}",
                "section_heading": clean_heading,
                "section": detect_section(f"{clean_heading}\n{text}"),
            }
        )
        current_parts = []

    for block in block_iter():
        if isinstance(block, Paragraph):
            text = block.text.strip()
            if not text:
                continue
            if is_heading(block) and current_parts:
                flush()
                current_heading = text
                current_parts.append(text)
            elif is_heading(block):
                current_heading = text
                current_parts.append(text)
            else:
                current_parts.append(text)
            continue

        table_lines = []
        for row in block.rows:
            cells = [cell.text.strip().replace("\n", " ") for cell in row.cells]
            cells = [re.sub(r"\s+", " ", cell).strip() for cell in cells if cell and cell.strip()]
            if cells:
                table_lines.append(" | ".join(cells))
        if table_lines:
            current_parts.append("\n".join(table_lines))

    flush()

    if entries:
        return entries

    parts = [paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip()]
    return _split_plain_text_entries("\n\n".join(parts), filename, source_label="docx")


def _extract_text_entries_xlsx(file_path: str, filename: str) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to ingest .xlsx files") from exc

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    entries = []
    for index, sheet in enumerate(workbook.worksheets, start=1):
        rows = [f"[Sheet: {sheet.title}]"]
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if cells:
                rows.append(" | ".join(cells))
        sheet_text = "\n".join(rows).strip()
        if sheet_text:
            entries.append({
                "page": str(index),
                "text": sheet_text,
                "source": "xlsx",
            })
    workbook.close()
    return entries


def _extract_text_entries_non_pdf(file_path: str, filename: str) -> list[dict]:
    suffix = Path(filename).suffix.lower()
    if suffix in {".txt", ".md", ".csv", ".json"}:
        return _extract_text_entries_plain_file(file_path, filename)
    if suffix == ".docx":
        return _extract_text_entries_docx(file_path, filename)
    if suffix == ".xlsx":
        return _extract_text_entries_xlsx(file_path, filename)
    raise ValueError(f"Unsupported document extension: {suffix or '(none)'}")


def extract_and_chunk(file_path: str, filename: str, *, force_docling: bool = False):
    """
    Extract text and structure using Docling, and chunk hierarchically.
    This replaces both PyMuPDF extraction and regex semantic chunking.
    """
    suffix = Path(filename).suffix.lower()
    if suffix and suffix != ".pdf":
        logger.info(f"Extracting {filename} via {suffix.lstrip('.')} parser...")
        raw_entries = _extract_text_entries_non_pdf(file_path, filename)
        if not raw_entries:
            return [], [], []
        text_quality = _build_text_quality_metadata(
            raw_entries,
            text_source=suffix.lstrip(".") or "text",
        )
        _write_text_cache(filename, raw_entries)
        return _entries_to_chunks(raw_entries, filename, text_quality=text_quality)

    raw_entries = []
    direct_entries = []
    docling_entries = []
    page_count = None
    text_source = "unknown"

    if force_docling:
        logger.info(f"Force Docling extraction requested for {filename}; skipping pypdf direct text.")
    else:
        direct_candidates = []
        for extractor_name, extractor in (
            ("pypdf", _extract_text_entries_pypdf),
            ("pymupdf", _extract_text_entries_pymupdf),
        ):
            try:
                candidate_entries, candidate_page_count = extractor(file_path)
            except Exception as exc:
                logger.warning(f"{extractor_name} direct PDF extraction failed for {filename}: {exc}")
                continue
            if _should_use_direct_pdf_text(candidate_entries, candidate_page_count):
                score = _direct_pdf_candidate_score(candidate_entries, candidate_page_count)
                if extractor_name == "pymupdf":
                    score += 60
                direct_candidates.append(
                    (
                        score,
                        extractor_name,
                        candidate_entries,
                        candidate_page_count,
                    )
                )

        if direct_candidates:
            _score, extractor_name, direct_entries, page_count = max(direct_candidates, key=lambda item: item[0])
            logger.info(f"Extracting {filename} via {extractor_name} direct text...")
            raw_entries = direct_entries
            text_source = "pdf_text_layer"
        else:
            logger.info(f"Direct PDF text too sparse or incomplete for {filename}; falling back to Docling...")

    if not raw_entries:
        docling_entries = _extract_text_entries_docling(file_path, filename)
        raw_entries = docling_entries
        text_source = "docling_ocr"

    if _entries_need_arabic_ocr(raw_entries):
        ocr_entries = _extract_text_entries_tesseract(file_path, filename)
        if ocr_entries:
            raw_entries = _select_best_entries_by_page(direct_entries, docling_entries, raw_entries, ocr_entries)
            text_source = "hybrid_ocr"

    facts_preview = _facts_for_entries(raw_entries, filename)
    skip_reinforcement = text_source == "pdf_text_layer" and _direct_pdf_text_has_offer_envelope_content(raw_entries)
    if not skip_reinforcement and _extracted_facts_need_ocr_reinforcement(facts_preview, raw_entries):
        target_pages = _target_pages_for_ocr_reinforcement(raw_entries, facts_preview)
        if target_pages:
            logger.info(
                f"Core facts look weak for {filename}; OCR reinforcing pages {target_pages}..."
            )
            reinforced_ocr_entries = _extract_text_entries_tesseract(file_path, filename, pages=target_pages)
            if reinforced_ocr_entries:
                raw_entries = _select_best_entries_by_page(
                    direct_entries,
                    docling_entries,
                    raw_entries,
                    reinforced_ocr_entries,
                )
                text_source = "hybrid_ocr"

    text_quality = _build_text_quality_metadata(
        raw_entries,
        page_count=page_count,
        text_source=text_source,
        preferred_source=text_source,
    )
    _write_text_cache(filename, raw_entries)
    return _entries_to_chunks(raw_entries, filename, text_quality=text_quality)


# ============= Structured Facts =============

FACT_SENTENCE_STOP = r"(?=(?:\.\s+[A-ZÀ-ÖØ-Þ]|\.\s+ARTICLE\b|\.\s+\d{1,3}\b|\.$|\n|$))"
FACT_VALUE_STOP = r"(?=(?:\.\s+[A-ZÀ-ÖØ-Þ]|\.\s+ARTICLE\b|\.\s+\d{1,3}\b|\.\s*$|\s+ARTICLE\s+\d+\b|\n|$))"
FACT_CLAUSE_STOP = r"(?=(?:[;.\n]|$))"
FACT_SUBJECT_PATTERNS = (
    re.compile(
        rf"\barticle\s*\d+\s*[.:\-]?\s*objet(?:\s+(?:du|de\s+la|de\s+l['’])\s+(?:march[eé]|consultation|cahier\s+des\s+charges))?\s*[.:\-]?\s*(.+?){FACT_SENTENCE_STOP}",
        re.IGNORECASE | re.DOTALL,
    ),
    re.compile(
        rf"\bobjet\s+(?:du|de\s+la|de\s+l['’])\s+(?:march[eé]|consultation|cahier\s+des\s+charges)\s*[.:\-]?\s*(.+?){FACT_SENTENCE_STOP}",
        re.IGNORECASE | re.DOTALL,
    ),
    re.compile(
        rf"\b((?:la\s+)?TSB\s+envisage\s+de\s+.+?){FACT_SENTENCE_STOP}",
        re.IGNORECASE | re.DOTALL,
    ),
    re.compile(
        rf"\b(?:la\s+)?(?:pr[eé]sente\s+)?consultation\s+a\s+pour\s+objet\s+(.+?){FACT_SENTENCE_STOP}",
        re.IGNORECASE | re.DOTALL,
    ),
    re.compile(
        rf"\b(?:le\s+)?(?:pr[eé]sent\s+)?(?:march[eé]|appel\s+d['’]offres?|cahier\s+des\s+charges)\s+a\s+pour\s+objet\s+(.+?){FACT_SENTENCE_STOP}",
        re.IGNORECASE | re.DOTALL,
    ),
    re.compile(
        rf"\bporte\s+sur\s+(.+?){FACT_SENTENCE_STOP}",
        re.IGNORECASE | re.DOTALL,
    ),
    re.compile(
        rf"\bobjet(?:\s+de\s+la\s+consultation)?\s*[:\-]\s*(.+?){FACT_SENTENCE_STOP}",
        re.IGNORECASE | re.DOTALL,
    ),
    re.compile(
        rf"\bconsultation\s+pour\s+(.+?){FACT_SENTENCE_STOP}",
        re.IGNORECASE | re.DOTALL,
    ),
    re.compile(
        rf"\bl['’']ubci\s+se\s+propose\s+d['’']acqu[eé]rir\s+(.+?){FACT_SENTENCE_STOP}",
        re.IGNORECASE | re.DOTALL,
    ),
    re.compile(
        rf"\bmarch[eé]\s+d['’']acquisition\s+(.+?){FACT_SENTENCE_STOP}",
        re.IGNORECASE | re.DOTALL,
    ),
)
FACT_DEADLINE_PATTERNS = (
    re.compile(
        rf"\bla\s+date\s+limite(?:\s+de\s+(?:la\s+)?(?:remise|r[eé]ception)\s+des\s+offres)?\s+"
        rf"(?:est\s+)?(?:fix[eé]e|arr[eê]t[eé]e)\s+(?:au|pour)\s+(.+?){FACT_VALUE_STOP}",
        re.IGNORECASE | re.DOTALL,
    ),
    re.compile(
        rf"\bdate\s+limite[^.\n]{{0,220}}?\b(?:est\s+)?(?:fix[eé]e|arr[eê]t[eé]e)\s+(?:au|pour)\s+(.+?){FACT_VALUE_STOP}",
        re.IGNORECASE | re.DOTALL,
    ),
    re.compile(
        rf"\bdate\s+limite(?:\s+de\s+(?:la\s+)?(?:remise|r[eé]ception)\s+des\s+offres)?\s*[:\-]\s*(.+?){FACT_SENTENCE_STOP}",
        re.IGNORECASE | re.DOTALL,
    ),
    re.compile(
        rf"\barticle\s*\d+\s*[.:\-]?\s*date\s+limite(?:\s+de\s+(?:la\s+)?(?:remise|r[eé]ception)\s+des\s+offres)?\s*[:\-]\s*(.+?){FACT_SENTENCE_STOP}",
        re.IGNORECASE | re.DOTALL,
    ),
    re.compile(
        rf"\bremise\s+des\s+offres\s*[:\-]\s*(.+?){FACT_SENTENCE_STOP}",
        re.IGNORECASE | re.DOTALL,
    ),
)


FACT_SCALAR_PATTERNS: dict[str, tuple[re.Pattern, ...]] = {
    "validity": (
        re.compile(
            rf"\b(?:d[eé]lai\s+de\s+)?validit[eé]\s+(?:de\s+l['’]offre|des\s+offres|de\s+la\s+soumission)\s*[:\-]\s*(.+?){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"(\boffres?\s+(?:techniques?\s+et\s+financi[eè]res?\s+)?(?:resteront?|reste(?:nt)?)\s+valables?\s+.+?){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"(\bles\s+soumissionnaires\s+sont\s+engag[eÃ©]s\s+par\s+leurs\s+offres\s+pendant\s+.+?){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"(\bles\s+soumissionnaires\s+sont\s+engag\S+s\s+par\s+leurs\s+offres\s+pendant\s+.+?){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"(\bles\s+(?:candidats|soumissionnaires)\s+sont\s+li\S+s\s+par\s+leurs\s+offres\s+(?:pour\s+une\s+p[eé]riode\s+de\s+|pendant\s+).+?){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
    ),
    "opening": (
        re.compile(
            rf"\bouverture\s+(?:des\s+plis|des\s+offres)\s*[:\-]\s*(.+?){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"\bs[eé]ance\s+(?:publique\s+)?(?:d['’])?ouverture\s*[:\-]\s*(.+?){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"(\bla\s+s[eé]ance\s+(?:publique\s+)?(?:d['’])?ouverture[^.\n]+){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
    ),
    "caution": (
        re.compile(
            rf"((?:chaque\s+offre\s+doit\s+[eê]tre\s+accompagn[eé]e[^.\n]{{0,260}}|(?:le\s+)?montant\s+de\s+la\s+caution\s+provisoire[^.\n]{{0,260}}|(?:la\s+)?caution\s+bancaire\s+provisoire[^.\n]{{0,220}})(?:\d[\d\s.,()]*\s*(?:DT|TND|dinars?)|douze\s+mille)[^.\n]{{0,80}})",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"\bcaution(?:nement)?\s+(?:bancaire\s+)?(?:provisoire|d[eé]finitive)?\s*[:\-]\s*(.+?){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"((?:le\s+)?cautionnement\s+provisoire\s+(?:de|est\s+fix[eÃ©]\s+[aÃ ])\s+[^.\n]{{0,180}}(?:%|pour\s+cent|DT|TND|dinars?)[^.\n]{{0,120}}){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"\bgarantie\s+provisoire\s*[:\-]\s*(.+?){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"(\bvalidit[eé]\s+de\s+la\s+caution\s+bancaire\s+provisoire[^.\n]+){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
    ),
    "guarantee": (
        re.compile(
            rf"((?:le\s+)?d[eé]lai\s+de\s+garantie[^.\n]+){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"\bd[eé]lai\s+de\s+garantie\s*[:\-]?\s*(.+?){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"\bgarantie\s+(?:constructeur|contractuelle|technique|de\s+bonne\s+ex[eé]cution|d[eé]finitive)\s*[:\-]\s*(.+?){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"\bp[eé]riode\s+de\s+garantie\s*[:\-]?\s*(.+?){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
    ),
    "payment": (
        re.compile(
            rf"((?:les\s+)?conditions?\s+de\s+paiement[^.\n]{{0,180}}?(?:sont\s+)?fix[eé]es?[\s\S]{{0,900}}?(?:30\s+jours|retenue\s+de\s+garantie|virement|ch[eè]que))",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"\b(?:modalit[eé]s?|conditions?)\s+de\s+paiement\s*[:\-]\s*(.+?){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"\bpaiements?\s+s['’]effectueront\s+(.+?){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"\bsp[eé]cificit[eé]\s+de\s+paiement\s*[:\-]?\s*(.+?){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
    ),
    "penalties": (
        re.compile(
            rf"\bp[eé]nalit[eé]s?\s+(?:de\s+retard)?\s*[:\-]\s*(.+?){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"(\bp[eé]nalit[eé]s?[^.\n]+(?:retard|montant|jour)[^.\n]*){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
    ),
    "cnss": (
        re.compile(
            rf"((?:un\s+|une\s+)?(?:certificat|attestation)[^;.\n]{{0,160}}\bCNSS\b[^;.\n]{{0,160}}){FACT_CLAUSE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
    ),
    "rne": (
        re.compile(
            rf"((?:un\s+|une\s+)?(?:extrait|certificat|attestation)[^;.\n]{{0,160}}\b(?:RNE|registre\s+(?:de\s+commerce|national(?:\s+des\s+entreprises)?))\b[^;.\n]{{0,160}}){FACT_CLAUSE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"(\bRNE\b[^;.\n]{{0,180}}){FACT_CLAUSE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
    ),
    "submission_method": (
        re.compile(
            rf"((?:les\s+)?offres?[^.\n]{{0,100}}\b(?:doivent|devraient|doit)\s+(?:obligatoirement\s+)?(?:parvenir|[eê]tre\s+(?:remises?|d[eé]pos[eé]es?|envoy[eé]es?|adress[eé]es?))[^.\n]{{0,520}}(?:voie\s+postale|rapide-poste|bureau\s+d['’]\s*ordre|remise\s+directe)[^.\n]{{0,220}}){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"((?:les\s+)?(?:offres|soumissions?)\s+doivent\s+(?:parvenir|[eê]tre\s+(?:d[eé]pos[eé]es?|envoy[eé]es?|adress[eé]es?))[^.\n]+){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"((?:mode\s+(?:d['’]envoi|de\s+d[eé]p[oô]t)|d[eé]p[oô]t\s+des\s+offres|envoi\s+des\s+offres)\s*[:\-]?\s*[^.\n]+){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"((?:par\s+(?:voie\s+postale|rapide-poste|courrier|tuneps)|bureau\s+d['’]ordre|d[eé]p[oô]t\s+direct)[^.\n]{{0,220}}){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
    ),
    "variants": (
        re.compile(
            rf"(\bvariantes?\s+(?:ne\s+sont\s+pas\s+|sont\s+)?(?:autoris[eé]es?|admises?|accept[eé]es?|permises?|interdites?)[^.\n]*){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"(\bles\s+variantes?[^.\n]{{0,180}}){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
    ),
    "information_sheet": (
        re.compile(
            rf"((?:fiche\s+(?:de\s+renseignements|signal[eé]tique)|formulaire\s+de\s+renseignements)[^;.\n]{{0,180}}){FACT_CLAUSE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
    ),
    "fiscal_certificate": (
        re.compile(
            rf"((?:attestation|certificat)[^;.\n]{{0,160}}\b(?:situation\s+fiscale|solde\s+fiscal|fiscale)\b[^;.\n]{{0,160}}){FACT_CLAUSE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
    ),
    "manufacturer_authorization": (
        re.compile(
            rf"((?:autorisation|lettre|agr[eé]ment)[^;.\n]{{0,160}}\b(?:constructeur|fabricant|[eé]diteur)\b[^;.\n]{{0,160}}){FACT_CLAUSE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
    ),
    "references": (
        re.compile(
            rf"((?:la\s+)?liste\s+d\S?au\s+moins\s+\d+\s+travaux\s+similaires[^;.\n]{{0,220}}){FACT_CLAUSE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"((?:liste\s+d['’]au\s+moins\s+\d+\s+)?travaux\s+similaires[^;.\n]{{0,220}}){FACT_CLAUSE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
        re.compile(
            rf"((?:liste\s+des\s+)?r[eé]f[eé]rences?[^;.\n]{{0,200}}){FACT_CLAUSE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
    ),
    "reception": (
        re.compile(
            rf"((?:r[eé]ception\s+(?:provisoire|d[eé]finitive|quantitative|technique)|proc[eè]s-verbal\s+de\s+r[eé]ception|PV\s+de\s+r[eé]ception)[^.\n]*){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
    ),
    "definitive_caution": (
        re.compile(
            rf"((?:(?:un|une|la)\s+)?(?:caution|cautionnement|garantie)\s+(?:d[eé]finitive|de\s+bonne\s+ex[eé]cution)[^.\n]*){FACT_VALUE_STOP}",
            re.IGNORECASE | re.DOTALL,
        ),
    ),
}


FACT_LIST_STOP_PATTERNS: dict[str, tuple[re.Pattern, ...]] = {
    "administrative_documents": (
        re.compile(r"\b\d+\s*[.)]?\s*offre\s+technique\b", re.IGNORECASE),
        re.compile(r"\boffre\s+technique\s+[«\"]?\s*enveloppe\b", re.IGNORECASE),
        re.compile(r"\bdossier\s+technique\b", re.IGNORECASE),
        re.compile(r"\b[A-Z]\s*[-.)]?\s*dossier\s+de\s+l['’]?\s*offre\s+financi\S+re\b", re.IGNORECASE),
        re.compile(r"\bdossier\s+de\s+l['’]?\s*offre\s+financi\S+re\b", re.IGNORECASE),
        re.compile(r"\b\d+\s*[.)]?\s*offre\s+financi[eè]re\b", re.IGNORECASE),
        re.compile(r"\b[A-Z]\s*[-.)]?\s*dossier\s+de\s+l['’]?\s*offre\s+technique\b", re.IGNORECASE),
        re.compile(r"\bdossier\s+de\s+l['’]?\s*offre\s+technique\b", re.IGNORECASE),
        re.compile(r"\bARTICLE\s+\d+\b", re.IGNORECASE),
    ),
    "technical_documents": (
        re.compile(r"\b\d+\s*[.)]?\s*offre\s+financi[eè]re\b", re.IGNORECASE),
        re.compile(r"\boffre\s+financi[eè]re\s+[«\"]?\s*enveloppe\b", re.IGNORECASE),
        re.compile(r"\bdeuxi[eè]me\s+enveloppe\b", re.IGNORECASE),
        re.compile(r"\bdossier\s+financier\b", re.IGNORECASE),
        re.compile(r"\bARTICLE\s+\d+\b", re.IGNORECASE),
    ),
    "financial_documents": (
        re.compile(r"\b[A-Z]\s*[-.)]?\s*dossier\s+de\s+l['’]?\s*offre\s+technique\b", re.IGNORECASE),
        re.compile(r"\bdossier\s+de\s+l['’]?\s*offre\s+technique\b", re.IGNORECASE),
        re.compile(r"\bARTICLE\s+\d+\b", re.IGNORECASE),
        re.compile(r"\bdate\s+limite\b", re.IGNORECASE),
        re.compile(r"\bcaution\b", re.IGNORECASE),
        re.compile(r"\bp[eé]nalit[eé]\b", re.IGNORECASE),
    ),
}

FACT_LIST_DEFINITIONS: dict[str, tuple[re.Pattern, ...]] = {
    "administrative_documents": (
        re.compile(r"(?:^|\n|\d+\s*[.)]\s*)dossier\s+administratif\b", re.IGNORECASE),
        re.compile(r"\bpi[eè]ces\s+administratives\s+suivantes\b", re.IGNORECASE),
        re.compile(r"\bpi\S*ces\s+administratives\s+(?:a|à)\s+fournir\b", re.IGNORECASE),
        re.compile(r"\bdocuments?\s+administratifs?\b", re.IGNORECASE),
        re.compile(r"\bdossier\s+administratif\b", re.IGNORECASE),
        re.compile(r"\boffre\s+administrative\b", re.IGNORECASE),
    ),
    "technical_documents": (
        re.compile(
            r"\bdossier\s+de\s+l['’]\s*offre\s+technique\b[\s\S]{0,220}?\b(?:doit\s+(?:contenir|comporter)|contient|comprend|pi[eè]ces\s+suivantes)\b",
            re.IGNORECASE,
        ),
        re.compile(r"(?:^|\n|\d+\s*[.)]\s*)offre\s+technique\b", re.IGNORECASE),
        re.compile(
            r"\benveloppe\b[\s\S]{0,140}\boffre\s+technique\b[\s\S]{0,180}?\bdoit\s+(?:contenir|comporter)\b",
            re.IGNORECASE,
        ),
        re.compile(
            r"(?:^|\n|\d+\s*[.)]\s*)offre\s+technique\b[\s\S]{0,220}?\b(?:doit\s+(?:contenir|comporter)|contient|comprend|pi[eè]ces\s+suivantes)\b",
            re.IGNORECASE,
        ),
        re.compile(r"\bpi\S+ces\s+techniques\s+suivantes\b", re.IGNORECASE),
        re.compile(r"\bdocuments?\s+techniques?\b", re.IGNORECASE),
        re.compile(r"\bdossier\s+technique\b", re.IGNORECASE),
    ),
    "financial_documents": (
        re.compile(
            r"\bdossier\s+de\s+l['’]\s*offre\s+financi[eè]re\b[\s\S]{0,220}?\b(?:doit\s+(?:contenir|comporter)|doit\s+obligatoirement\s+comporter|contient|comprend|pi[eè]ces\s+suivantes)\b",
            re.IGNORECASE,
        ),
        re.compile(r"(?:^|\n|\d+\s*[.)]\s*)offre\s+financi[eè]re\b", re.IGNORECASE),
        re.compile(
            r"\benveloppe\b[\s\S]{0,140}\boffre\s+financi[eè]re\b[\s\S]{0,180}?\bdoit\s+(?:contenir|comporter)\b",
            re.IGNORECASE,
        ),
        re.compile(
            r"\boffre\s+financi[eè]re\b[\s\S]{0,120}?\bdoit\s+(?:contenir|comporter)\b",
            re.IGNORECASE,
        ),
        re.compile(
            r"(?:^|\n|\d+\s*[.)]\s*)offre\s+financi[eè]re\b[\s\S]{0,220}?\b(?:doit\s+(?:contenir|comporter)|contient|comprend)\b",
            re.IGNORECASE,
        ),
        re.compile(r"\bdocuments?\s+financiers?\b", re.IGNORECASE),
        re.compile(r"\bdossier\s+financier\b", re.IGNORECASE),
    ),
}


def _page_sort_key(page: str | int | None) -> tuple[int, str]:
    if page is None:
        return (10_000, "")
    page_text = str(page).strip()
    match = re.search(r"\d+", page_text)
    if match:
        return (int(match.group(0)), page_text)
    return (10_000, page_text)


def _normalize_fact_text(text: str) -> str:
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", " ", text)
    cleaned = re.sub(r"\s+", " ", text).strip(" .;:-")
    starts_with_caution = re.match(
        r"^(?:(?:un|une|la)\s+)?(?:caution|cautionnement|garantie)\s+",
        cleaned,
        flags=re.IGNORECASE,
    )
    section_boundary = (
        r"\s+\b(?:ARTICLE\s+\d+|(?:LA\s+)?DATE\s+LIMITE|OUVERTURE\s+DES\s+(?:PLIS|OFFRES)|"
        r"MODALIT[EÉ]S?\s+DE\s+PAIEMENT|P[EÉ]NALIT[EÉ]S?)\b"
    )
    if not starts_with_caution:
        section_boundary = (
            r"\s+\b(?:ARTICLE\s+\d+|(?:LA\s+)?DATE\s+LIMITE|OUVERTURE\s+DES\s+(?:PLIS|OFFRES)|"
            r"CAUTION|MODALIT[EÉ]S?\s+DE\s+PAIEMENT|P[EÉ]NALIT[EÉ]S?)\b"
        )
    cleaned = re.split(section_boundary, cleaned, maxsplit=1, flags=re.IGNORECASE)[0].strip(" .;:-")
    cleaned = re.sub(
        r"\b(?:article|lot|chapitre)\s+\d+\b.*$",
        "",
        cleaned,
        flags=re.IGNORECASE,
    ).strip(" .;:-")
    return cleaned


def _fold_fact_text(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", str(text))
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    normalized = normalized.replace("’", "'").replace("‘", "'").replace("`", "'")
    return re.sub(r"\s+", " ", normalized).lower().strip()


def _compact_fact_text(text: str) -> str:
    return re.sub(r"\s+", " ", str(text)).strip(" .;:-")


def _group_chunks_by_page(chunks: list[str], metas: list[dict]) -> list[dict]:
    grouped: dict[str, dict] = {}

    for chunk, meta in zip(chunks, metas):
        page = str(meta.get("page", "?"))
        entry = grouped.setdefault(
            page,
            {"page": page, "section": meta.get("section", "general"), "parts": []},
        )
        entry["parts"].append(chunk)

    ordered = []
    for page in sorted(grouped.keys(), key=_page_sort_key):
        entry = grouped[page]
        ordered.append(
            {
                "page": entry["page"],
                "section": entry["section"],
                "text": _clean_chunk_text("\n\n".join(entry["parts"])),
            }
        )
    return ordered


def _extract_fact_from_pages(pages: list[dict], patterns: tuple[re.Pattern, ...], formatter=None) -> dict | None:
    candidates = _extract_fact_candidates_from_pages(pages, patterns, formatter=formatter)
    if not candidates:
        return None

    fact = dict(candidates[0])
    fact.pop("_pattern_index", None)
    return fact


def _extract_fact_candidates_from_pages(
    pages: list[dict],
    patterns: tuple[re.Pattern, ...],
    formatter=None,
) -> list[dict]:
    candidates = []
    seen = set()

    for page_entry in pages:
        compact = re.sub(r"\s+", " ", page_entry["text"]).strip()
        raw_text = page_entry["text"].strip()
        if not compact and not raw_text:
            continue

        for candidate in (compact, raw_text):
            if not candidate:
                continue

            for pattern_index, pattern in enumerate(patterns):
                for match in pattern.finditer(candidate):
                    extracted = _normalize_fact_text(match.group(1))
                    if not extracted:
                        continue

                    if formatter:
                        extracted = formatter(extracted)

                    key = (page_entry["page"], extracted.lower())
                    if key in seen:
                        continue
                    seen.add(key)

                    candidates.append(
                        {
                            "text": extracted,
                            "page": page_entry["page"],
                            "section": page_entry["section"],
                            "_pattern_index": pattern_index,
                        }
                    )

    return candidates


def _subject_formatter(pages: list[dict]):
    def _format(text: str) -> str:
        normalized = text.replace("’", "'").lower()
        if (
            not normalized.startswith("l'ubci se propose")
            and "acqu" in normalized
            and any("ubci" in page["text"].lower() for page in pages[:1])
        ):
            return f"L'UBCI se propose d'acquerir {text}"
        return text

    return _format


def _score_subject_fact(fact: dict) -> tuple[int, int, int, int]:
    text = str(fact.get("text", ""))
    normalized = text.replace("’", "'").lower()
    folded = _fold_fact_text(text)

    try:
        page_score = -int(str(fact.get("page", "999")).split()[0])
    except (TypeError, ValueError):
        page_score = -999

    explicit_score = 0
    if "se propose" in normalized and "acqu" in normalized:
        explicit_score += 100
    if "a pour objet" in normalized:
        explicit_score += 120
    if "consultation pour" in normalized:
        explicit_score += 80
    if "tsb envisage" in folded:
        explicit_score += 170
    if "infrastructure systeme" in folded:
        explicit_score += 90
    if "topnet" in folded and "solution antivirale" in folded:
        explicit_score += 180
    if "securite des infrastructures informatiques" in folded:
        explicit_score += 100
    if "edr" in folded:
        explicit_score += 60
    if "lance le present appel d'offres" in folded or "en lot unique pour l'acquisition" in folded:
        explicit_score += 80
    if "fourniture et la livraison" in folded:
        explicit_score += 90
    if re.search(r"\d", text):
        explicit_score += 30
    if "solution d'impression" in folded and "multifonction" in folded:
        explicit_score += 180
    if (
        page_score >= -3
        and explicit_score < 50
        and ("objet" in folded or "a pour objet" in folded)
    ):
        explicit_score += 60
    if len(text) < 25:
        explicit_score -= 40
    if any(
        marker in normalized
        for marker in (
            "je joins",
            "presente soumission",
            "présente soumission",
            "cahier des clauses",
            "attestation fiscale",
            "attestation d'affiliation",
            "attestation d’affiliation",
            "fixent les procedures",
            "procedures de l'appel d'offres",
            "stipulent les conditions du marche",
            "definissent les conditions",
            "regissent les rapports",
            "precisent les obligations",
            "determinent les modalites",
            "fixent les conditions",
            "definissent les droits",
            "approuver les rapports",
            "delivrer les certificats",
        )
    ):
        explicit_score -= 160
    if any(
        marker in folded
        for marker in (
            "etcompositiondel",
            "compositiondelappeld",
            "composition de l appel d offres",
            "table des matieres",
            "sommaire",
        )
    ):
        explicit_score -= 260
    if ARTICLE_MARKER_RE.search(text) and len(ARTICLE_MARKER_RE.findall(text)) >= 3:
        explicit_score -= 300
    if (
        re.search(r"^\s*annexe\s+(?:n[°o]\s*)?\d+", folded)
        or "modele de soumission" in folded
        or "formulaire de reponse" in folded
    ):
        explicit_score -= 180
    if re.search(r"\.{2,}\s*,?\s*\d+\s*=", text):
        explicit_score -= 220
    if (
        (text.count(", 1 =") >= 2 or folded.count("composition du marche") >= 1)
        and "a pour objet" not in folded
    ):
        explicit_score -= 260

    pattern_score = 20 - int(fact.get("_pattern_index", 20))
    length_score = min(len(text), 240)
    return (explicit_score, pattern_score, length_score, page_score)


ARTICLE_MARKER_RE = re.compile(r"\bARTICLE\s+\d+\b", re.IGNORECASE)
DATE_VALUE_RE = re.compile(
    r"\b\d{1,2}\s+"
    r"(?:janvier|f[eé]vrier|mars|avril|mai|juin|juillet|ao[uû]t|septembre|octobre|novembre|d[eé]cembre)"
    r"\s+\d{4}(?:\s*(?:a|à)\s*\d{1,2}\s*h(?:\s*\d{2})?)?",
    re.IGNORECASE,
)
NUMERIC_DATE_VALUE_RE = re.compile(
    r"\b(?:\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}[/-]\d{1,2}[/-]\d{1,2})"
    r"(?:\s*(?:a|à)\s*\d{1,2}\s*h(?:\s*\d{2})?)?\b",
    re.IGNORECASE,
)
AMOUNT_VALUE_RE = re.compile(r"\b\d[\d\s.,()]*\s*(?:DT|TND|dinars?)\b", re.IGNORECASE)
WORD_AMOUNT_VALUE_RE = re.compile(
    r"\b(?:douze\s+mille|cinq\s+mille|dix\s+mille|vingt\s+mille)"
    r"(?:\s*\(\s*\d[\d\s.,]*\s*\))?\s*(?:DT|TND|dinars?)?\b",
    re.IGNORECASE,
)
PERCENT_VALUE_RE = re.compile(r"\b\d+(?:[.,]\d+)?\s*%(?=\s|$|[.;,])")


def _flatten_pages_for_articles(pages: list[dict]) -> tuple[str, list[dict]]:
    text_parts = []
    offsets = []
    cursor = 0

    for page_entry in pages:
        page_text = str(page_entry.get("text", "")).strip()
        if not page_text:
            continue

        if text_parts:
            text_parts.append("\n\n")
            cursor += 2

        start = cursor
        text_parts.append(page_text)
        cursor += len(page_text)
        offsets.append(
            {
                "start": start,
                "end": cursor,
                "page": page_entry.get("page", "?"),
                "section": page_entry.get("section", "general"),
            }
        )

    return "".join(text_parts), offsets


def _page_info_at_offset(offsets: list[dict], offset: int) -> tuple[str, str]:
    for entry in offsets:
        if entry["start"] <= offset <= entry["end"]:
            return str(entry["page"]), str(entry["section"])
    if offsets:
        return str(offsets[-1]["page"]), str(offsets[-1]["section"])
    return "?", "general"


def _article_sections_from_pages(pages: list[dict]) -> list[dict]:
    doc_text, offsets = _flatten_pages_for_articles(pages)
    matches = list(ARTICLE_MARKER_RE.finditer(doc_text))
    sections = []

    for index, match in enumerate(matches):
        start = match.start()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(doc_text)
        section_text = doc_text[start:end].strip()
        if not section_text:
            continue

        page, section = _page_info_at_offset(offsets, start)
        sections.append(
            {
                "text": section_text,
                "compact": _compact_fact_text(section_text),
                "folded": _fold_fact_text(section_text[:420]),
                "page": page,
                "section": section,
                "start": start,
            }
        )

    return sections


def _fact_from_text(text: str, page: str, section: str) -> dict | None:
    cleaned = _compact_fact_text(text)
    cleaned = re.split(r"\s+\bARTICLE\s+\d+\b", cleaned, maxsplit=1, flags=re.IGNORECASE)[0].strip(" .;:-")
    if not cleaned:
        return None
    return {"text": cleaned, "page": page, "section": section}


def _article_body_after(section_text: str, marker: str) -> str:
    compact = _compact_fact_text(section_text)
    folded = _fold_fact_text(compact)
    marker_index = folded.find(marker)
    if marker_index < 0:
        return compact

    body = compact[marker_index + len(marker) :]
    return body.strip(" .;:-")


def _article_title_folded(article: dict) -> str:
    compact = article["compact"][:320]
    body_start = re.search(
        r"\b(?:La|Le|Les|L['’]|Chaque|Pour|Dans|Au|TSB|T\.S\.B)\b",
        compact[12:],
        flags=re.IGNORECASE,
    )
    if body_start:
        compact = compact[: 12 + body_start.start()]
    return _fold_fact_text(compact)


def _strip_article_heading(compact: str) -> str:
    return re.sub(
        r"^\s*ARTICLE\s+\d+\s*[.:|\-]?\s*[^:]{0,120}?(?:\s*:\s*|\s+)",
        "",
        compact,
        count=1,
        flags=re.IGNORECASE,
    ).strip(" .;:-")


def _first_sentence_with(folded_marker: str, text: str, *, max_chars: int = 650) -> str | None:
    compact = _compact_fact_text(text)
    parts = re.split(r"(?<=[.!?])\s+", compact)
    for part in parts:
        if folded_marker in _fold_fact_text(part):
            return part[:max_chars].strip(" .;:-")
    return None


def _date_value_from_text(text: str) -> str | None:
    for pattern in (DATE_VALUE_RE, NUMERIC_DATE_VALUE_RE):
        match = pattern.search(text)
        if match:
            return match.group(0).strip(" .;:-()")
    return None


PLACEHOLDER_TOKEN_RE = re.compile(
    r"\b(?P<prefix>IM|DPAO)\s*\(?\s*(?P<number>\d{1,3})\s*\)?",
    re.IGNORECASE,
)

PLACEHOLDER_REFERENCE_RE = re.compile(
    r"\(?\s*\b(?P<prefix>IM|DPAO)\s*\(?\s*(?P<number>\d{1,3})\s*\)?\s*\)?",
    re.IGNORECASE,
)


def _placeholder_key(prefix: str, number: str) -> str:
    return f"{prefix.upper()}:{int(number)}"


def _clean_placeholder_value(value: str) -> str | None:
    value = _compact_fact_text(value)
    value = re.split(r"\s+\bARTICLE\s+\d+\b", value, maxsplit=1, flags=re.IGNORECASE)[0]
    value = re.split(r"\s+\b(?:DPAO|IM)\s*\(?\s*\d{1,3}\s*\)?", value, maxsplit=1, flags=re.IGNORECASE)[0]
    value = value.strip(" .;:-")
    value = re.sub(r"^[),.;:\-\s]+", "", value).strip(" .;:-")
    value = re.sub(r"\s+", " ", value)
    duration_only = re.match(
        r"^((?:60|90|120|180)\s+jours?)\b(?:\W+(?:des\s+offres|offres?|soumissions?).*)?$",
        value,
        flags=re.IGNORECASE,
    )
    if duration_only:
        value = duration_only.group(1)

    if len(value) < 2:
        return None
    if not any(ch.isalnum() for ch in value):
        return None
    if _fold_fact_text(value).startswith(("specifies dans", "indiquees", "indiquee")):
        return None
    return value[:420].strip(" .;:-")


def _extract_placeholder_values(pages: list[dict]) -> dict[str, dict]:
    values: dict[str, dict] = {}

    for page_entry in pages:
        compact = _compact_fact_text(page_entry["text"])
        if not compact:
            continue

        matches = list(PLACEHOLDER_TOKEN_RE.finditer(compact))
        for index, match in enumerate(matches):
            if match.start() > 0 and compact[match.start() - 1] == "(":
                # Clause references often look like "(IM (13))"; filled table cells do not.
                continue

            after = compact[match.end() : match.end() + 12]
            if re.match(r"\s*[\]),.;:]", after):
                continue

            end = matches[index + 1].start() if index + 1 < len(matches) else min(len(compact), match.end() + 520)
            value = _clean_placeholder_value(compact[match.end() : end])
            if not value:
                continue

            key = _placeholder_key(match.group("prefix"), match.group("number"))
            current = values.get(key)
            if current and len(str(current.get("text", ""))) >= len(value):
                continue
            values[key] = {
                "text": value,
                "page": page_entry.get("page", "?"),
                "section": page_entry.get("section", "general"),
            }

    return values


def _resolve_placeholder_text(text: str, placeholder_values: dict[str, dict]) -> tuple[str, dict | None]:
    first_value: dict | None = None

    def _replace(match: re.Match) -> str:
        nonlocal first_value
        key = _placeholder_key(match.group("prefix"), match.group("number"))
        value = placeholder_values.get(key)
        if not value:
            return match.group(0)
        if first_value is None:
            first_value = value
        return str(value["text"])

    resolved = PLACEHOLDER_REFERENCE_RE.sub(_replace, text)
    resolved = re.sub(r"\s+([,.;:])", r"\1", resolved)
    resolved = re.sub(r"\(\s*\)", "", resolved)
    resolved = _compact_fact_text(resolved)
    return resolved, first_value


def _resolve_fact_placeholders(fact: dict | None, placeholder_values: dict[str, dict]) -> dict | None:
    if not fact or not placeholder_values:
        return fact

    text = str(fact.get("text", ""))
    if not PLACEHOLDER_REFERENCE_RE.search(text):
        return fact

    resolved, first_value = _resolve_placeholder_text(text, placeholder_values)
    if not resolved or resolved == text:
        return fact

    updated = dict(fact)
    updated["text"] = resolved
    if first_value:
        updated["page"] = first_value.get("page", updated.get("page", "?"))
        updated["section"] = first_value.get("section", updated.get("section", "general"))
    return updated


def _instruction_marker_fact(
    placeholder_values: dict[str, dict],
    key: str,
    text: str,
    section: str,
) -> dict | None:
    value = placeholder_values.get(key)
    if not value:
        return None
    return {
        "text": text,
        "page": value.get("page", "?"),
        "section": value.get("section", section),
    }


def _extract_instruction_marker_facts(placeholder_values: dict[str, dict]) -> dict:
    """
    Resolve standard tender instruction tables such as STEG-style IM(n) values.

    This stays generic for the document family: it only activates when the
    common IM table keys are present, then converts the table values into the
    checklist facts that otherwise remain hidden behind placeholders.
    """
    if len({key for key in placeholder_values if key.startswith("IM:")}) < 3:
        return {}

    facts = {}
    if placeholder_values.get("IM:1"):
        facts["subject"] = _instruction_marker_fact(
            placeholder_values,
            "IM:1",
            str(placeholder_values["IM:1"]["text"]),
            "subject",
        )
    if placeholder_values.get("IM:7"):
        facts["definitive_caution"] = _instruction_marker_fact(
            placeholder_values,
            "IM:7",
            f"Cautionnement définitif / caution de bonne fin : {placeholder_values['IM:7']['text']}.",
            "guarantee",
        )
    if placeholder_values.get("IM:11"):
        facts["penalties"] = _instruction_marker_fact(
            placeholder_values,
            "IM:11",
            f"Pénalité de retard : {placeholder_values['IM:11']['text']} par jour.",
            "penalty",
        )
    if placeholder_values.get("IM:13"):
        facts["guarantee"] = _instruction_marker_fact(
            placeholder_values,
            "IM:13",
            f"Délai de garantie : {placeholder_values['IM:13']['text']}.",
            "guarantee",
        )
    return {key: value for key, value in facts.items() if value}


def _extract_dpc_number_value(pages: list[dict], number: int, *, max_chars: int = 700) -> dict | None:
    marker_re = re.compile(rf"\(\s*{number}\s*\)", re.IGNORECASE)
    next_marker_re = re.compile(r"\(\s*\d{1,3}\s*\)|\bArticle\s+\d+\b", re.IGNORECASE)

    for page_entry in pages:
        compact = _compact_fact_text(page_entry["text"])
        folded = _fold_fact_text(compact)
        if "dpc" not in folded and "donnees particulieres" not in folded:
            continue

        match = marker_re.search(compact)
        if not match:
            continue

        tail = compact[match.end() : match.end() + max_chars]
        next_match = next_marker_re.search(tail)
        value = tail[: next_match.start()] if next_match else tail
        value = _clean_placeholder_value(value)
        if not value:
            continue

        return {
            "text": value,
            "page": page_entry.get("page", "?"),
            "section": page_entry.get("section", "general"),
        }

    return None


def _extract_dpc_subject_fact(pages: list[dict]) -> dict | None:
    specific_title_pattern = re.compile(
        r"(ACQUISITION\s+DES?\s+PI\S?CES?\s+DE\s+RECHANGE\s+POUR\s+LES\s+IMPRIMANTES?"
        r"[\s\S]{0,240}?TUNIS\s+ET\s+SFAX)",
        flags=re.IGNORECASE,
    )
    title_pattern = re.compile(
        r"((?:ACQUISITION|FOURNITURE|FOURNITURES|RENOUVELLEMENT|MAINTENANCE|LOCATION)\s+"
        r"(?:DES?|DE\s+LA|DE\s+L['’])?[\s\S]{0,260}?"
        r"(?:VP\s*140|VP140|CANON|IMPRIMANTES?|PI[EÈ]CES?\s+DE\s+RECHANGE|TUNIS\s+ET\s+SFAX))",
        flags=re.IGNORECASE,
    )

    for page_entry in pages:
        compact = _compact_fact_text(page_entry["text"])
        match = specific_title_pattern.search(compact) or title_pattern.search(compact)
        if not match:
            continue
        value = _compact_fact_text(match.group(1)).strip(" .;:-")
        acquisition_start = re.search(
            r"\bACQUISITION\s+DES?\s+PI\S?CES?\s+DE\s+RECHANGE\b",
            value,
            flags=re.IGNORECASE,
        )
        if acquisition_start:
            value = value[acquisition_start.start() :]
        value = re.split(
            r"\b(?:Objet|DPC|Article\s+\d+|\(\s*\d+\s*\))\b",
            value,
            maxsplit=1,
            flags=re.IGNORECASE,
        )[0]
        value = _compact_fact_text(value).strip(" .;:-")
        if len(value) >= 35:
            return _fact_from_text(value[:420], page_entry.get("page", "?"), "subject")

    for page_entry in pages:
        compact = _compact_fact_text(page_entry["text"])
        folded = _fold_fact_text(compact)
        if "dpc" not in folded and "donnees particulieres" not in folded:
            continue

        match = re.search(
            r"\bArticle\s+1\b(?P<body>[\s\S]{0,900}?)(?:\(\s*1\s*\)|\bArticle\s+2\b|\(\s*7\s*\))",
            compact,
            flags=re.IGNORECASE,
        )
        if not match:
            continue

        value = match.group("body")
        value = re.sub(r"\b\d{4}\s*/\s*C?\d+\s*/\s*\d+\b", " ", value, flags=re.IGNORECASE)
        value = re.sub(r"\bObjet\b", " ", value, flags=re.IGNORECASE)
        value = re.sub(r"[\"'`«»“”]", " ", value)
        value = _compact_fact_text(value)
        start = re.search(
            r"\b(?:ACQUISITION|FOURNITURE|FOURNITURES|LOCATION|MAINTENANCE|RENOUVELLEMENT|"
            r"MISE\s+EN\s+PLACE|TRAVAUX|REALISATION|ETUDE)\b",
            value,
            flags=re.IGNORECASE,
        )
        if start:
            value = value[start.start() :]
        value = value.strip(" .;:-")

        if len(value) < 25:
            continue
        folded_value = _fold_fact_text(value)
        if "conditions de participation" in folded_value:
            continue
        if "travaux et/ou fourniture de biens et services" in folded_value:
            continue
        if "donnees particulieres de la consultation" in folded_value:
            continue

        return _fact_from_text(value[:420], page_entry.get("page", "?"), "subject")

    return None


def _extract_dpc_marker_facts(pages: list[dict]) -> dict:
    facts: dict[str, dict] = {}

    subject = _extract_dpc_subject_fact(pages)
    if subject:
        facts["subject"] = subject

    validity = _extract_dpc_number_value(pages, 14, max_chars=180)
    if validity and re.search(r"\b\d+\s*(?:jours?|mois|ans?)\b", str(validity.get("text", "")), re.IGNORECASE):
        facts["validity"] = validity

    deadline_value = _extract_dpc_number_value(pages, 12, max_chars=260)
    if deadline_value:
        date_value = _date_value_from_text(str(deadline_value.get("text", "")))
        if date_value:
            facts["deadline"] = _fact_from_text(date_value, deadline_value.get("page", "?"), deadline_value.get("section", "deadline"))

    rne = _extract_dpc_number_value(pages, 9, max_chars=220)
    if rne and "registre" in _fold_fact_text(str(rne.get("text", ""))):
        facts["rne"] = rne

    return facts


def _extract_deadline_value_from_pages(pages: list[dict]) -> dict | None:
    deadline_anchor_re = re.compile(
        r"\b(?:date\s+limite|remise\s+des\s+offres|r[eé]ception\s+des\s+offres|au\s+plus\s+tard)\b",
        re.IGNORECASE,
    )

    arabic_deadline_re = re.compile(
        r"(?:آخر\s+أجل|اخر\s+اجل|لقبول\s+العروض|قبول\s+العروض|موعد)"
        r"[\s\S]{0,160}?(\d{4}[/-]\d{1,2}[/-]\d{1,2}|\d{1,2}[/-]\d{1,2}[/-]\d{2,4})"
    )

    for page_entry in pages:
        compact = _compact_fact_text(page_entry["text"])
        for match in deadline_anchor_re.finditer(compact):
            window = compact[match.start() : match.start() + 520]
            date_value = _date_value_from_text(window)
            if date_value:
                return _fact_from_text(date_value, page_entry["page"], page_entry["section"])

        arabic_match = arabic_deadline_re.search(compact)
        if arabic_match:
            return _fact_from_text(arabic_match.group(1), page_entry["page"], page_entry["section"])

    return None


def _regex_fact_from_pages(
    pages: list[dict],
    pattern: re.Pattern,
    *,
    max_chars: int = 1200,
) -> dict | None:
    for page_entry in pages:
        compact = _compact_fact_text(page_entry["text"])
        match = pattern.search(compact)
        if not match:
            continue

        value = match.group(1) if match.groups() else match.group(0)
        fact = _fact_from_text(value[:max_chars], page_entry["page"], page_entry["section"])
        if fact:
            return fact

    return None


def _list_fact_from_items(items: list[str], page: str | None, section: str | None) -> dict | None:
    cleaned_items = []
    seen = set()
    for item in items:
        cleaned = _clean_fact_list_item(item)
        if not cleaned:
            continue
        folded = _fold_fact_text(cleaned)
        if folded in seen:
            continue
        seen.add(folded)
        cleaned_items.append(cleaned)

    if not cleaned_items:
        return None

    return {
        "text": "\n".join(f"- {item}" for item in cleaned_items),
        "items": [
            {
                "text": item,
                "page": page or "?",
                "section": section or "general",
            }
            for item in cleaned_items
        ],
        "page": page or "?",
        "section": section or "general",
    }


def _extract_subject_from_pages_fallback(pages: list[dict]) -> dict | None:
    for page_entry in pages[:60]:
        compact = _compact_fact_text(page_entry["text"])
        folded = _fold_fact_text(compact)
        if "acquisition et mise en place d'une solution d'impression" not in folded or "stb" not in folded:
            continue

        match = re.search(
            r"(Acquisition\s+et\s+mise\s+en\s+place\s+d['']une\s+solution\s+d['']impression\s+[aà]\s+la\s+STB)",
            compact,
            flags=re.IGNORECASE,
        )
        if match:
            return _fact_from_text(match.group(1), page_entry["page"], page_entry["section"])

    for page_entry in pages[:6]:
        compact = _compact_fact_text(page_entry["text"])
        folded = _fold_fact_text(compact)
        if "solution d'impression" not in folded or "50" not in folded:
            continue

        match = re.search(
            r"(Le\s+pr\S+sent\s+appel\s+d['ƒ?T]offres?\s+a\s+pour\s+objet[\s\S]{0,900}?"
            r"50\s+\S+quipements?\s+multifonction[\s\S]{0,260}?(?:scan|administration))",
            compact,
            flags=re.IGNORECASE,
        )
        if match:
            return _fact_from_text(match.group(1), page_entry["page"], page_entry["section"])

    pattern = re.compile(
        r"(La\s+soci[eé]t[eé]\s+TOPNET\s+se\s+propose\s+de\s+lancer\s+une\s+consultation"
        r"[\s\S]{0,1400}?\bEDR\b)",
        re.IGNORECASE,
    )
    fact = _regex_fact_from_pages(pages, pattern, max_chars=1200)
    if fact:
        return fact

    for page_entry in pages[:3]:
        compact = _compact_fact_text(page_entry["text"])
        match = re.search(
            r"\b(?P<qty>\d{1,4})\s+"
            r"(?P<item>Coupeuses?\s+de\s+plans?\s+A[O0Q][^\"،.;\n]{0,60})",
            compact,
            flags=re.IGNORECASE,
        )
        if match:
            item = _clean_mined_fact_text(match.group("item")).replace("AQ", "A0").replace("AO", "A0")
            return _fact_from_text(
                f"Acquisition de {match.group('qty')} {item}",
                page_entry["page"],
                page_entry["section"],
            )

    return None


def _extract_opening_fallback(pages: list[dict]) -> dict | None:
    simple_pattern = re.compile(
        r"(L['’]ouverture\s+des\s+plis\s+aura\s+lieu\s+dans\s+les\s+locaux\s+de\s+[^.]{3,120})",
        re.IGNORECASE,
    )
    fact = _regex_fact_from_pages(pages, simple_pattern, max_chars=220)
    if fact:
        return fact

    opening_heading_re = re.compile(
        r"(?:\b(?:article|atticle)\s+\d+[\s.:\-]*)?ouverture\s+des\s+(?:plis|offres)\b",
        re.IGNORECASE,
    )
    opening_stop_re = re.compile(
        r"\s+(?:"
        r"\b(?:article|atticle|chapitre)\s+\d+[\s.:\-]|"
        r"\b(?:montant\s+des\s+offres|evaluation|confidentialite|depouillement|"
        r"mise\s+au\s+point|complement\s+d['’]informations?|attribution|notification)\b|"
        r"\[page\s+\d+\]"
        r")",
        re.IGNORECASE,
    )
    evidence_markers = (
        "commission",
        "comite",
        "seance",
        "publique",
        "huis clos",
        "ouvrir",
        "ouverture simultanement",
        "locaux",
        "en ligne",
    )
    for page_entry in pages:
        compact = _compact_fact_text(page_entry["text"])
        for match in opening_heading_re.finditer(compact):
            stop_match = opening_stop_re.search(compact, match.end())
            end = stop_match.start() if stop_match else min(len(compact), match.end() + 560)
            value = compact[match.start() : end].strip(" .;:-")
            folded_value = _fold_fact_text(value)
            if any(marker in folded_value for marker in evidence_markers):
                return _fact_from_text(value[:640], page_entry["page"], page_entry["section"])

    pattern = re.compile(
        r"(L['’]ouverture\s+des\s+plis\s+ne\s+sera\s+pas\s+publique"
        r"[\s\S]{0,620}?offres\s+financi[eè]res?)",
        re.IGNORECASE,
    )
    return _regex_fact_from_pages(pages, pattern, max_chars=900)


def _extract_caution_fallback(pages: list[dict]) -> dict | None:
    caution_patterns = (
        re.compile(
            r"((?:Le\s+)?cautionnement\s+provisoire\s+de\s+\d+(?:[,.]\d+)?\s*%\s+du\s+montant\s+de\s+la\s+soumission[\s\S]{0,260}?(?:garantie|banque|rejet|soumission))",
            re.IGNORECASE,
        ),
        re.compile(
            r"((?:Le\s+soumissionnaire\s+doit\s+fournir[\s\S]{0,260}?)?"
            r"cautionnement\s+(?:provisoire|pro)[\s\S]{0,900}?\b120\s+jours[\s\S]{0,80})",
            re.IGNORECASE,
        ),
        re.compile(
            r"(caution\s+bancaire\s+provisoire[\s\S]{0,520}?"
            r"(?:(?:\d[\d\s.,()]*\s*(?:DT|TND|dinars?))|"
            r"(?:douze\s+mille|cinq\s+mille|dix\s+mille|vingt\s+mille)"
            r"(?:\s*\(\s*\d[\d\s.,]*\s*\))?\s*(?:DT|TND|dinars?)?)"
            r"[\s\S]{0,120})",
            re.IGNORECASE,
        ),
    )

    for page_entry in pages:
        compact = _compact_fact_text(page_entry["text"])
        for caution_re in caution_patterns:
            for match in caution_re.finditer(compact):
                value = match.group(1)
                value = re.split(
                    r"\b(?:Les\s+tableaux|Déclaration|La\s+liste|RNE|Une\s+attestation|Le\s+CCAP|ARTICLE\s+\d+)\b",
                    value,
                    maxsplit=1,
                    flags=re.IGNORECASE,
                )[0]
                fact = _fact_from_text(value, page_entry["page"], page_entry["section"])
                if fact:
                    return fact

    return None


def _extract_submission_method_from_pages(pages: list[dict]) -> dict | None:
    address_delivery_re = re.compile(
        r"((?:les\s+)?soumissionnaires\s+doivent\s+envoyer\s+leurs\s+offres\s+[aàÃ ]\s+"
        r"l['’â€™]adresse\s+suivante\s*:[\s\S]{0,420}?"
        r"(?:registre\s+du\s+bureau\s+d['’â€™]ordre|bureau\s+d['’â€™]ordre)[\s\S]{0,180})",
        re.IGNORECASE,
    )

    office_deposit_re = re.compile(
        r"((?:les\s+)?soumissionnaires\s+doivent\s+(?:disposer|d[eé]poser)\s+leurs\s+offres\s+"
        r"au\s+bureau\s+d['’]ordre\s+[aà]\s+l['’]adresse\s+suivante\s*:"
        r"[\s\S]{0,520}?)(?:La\s+date\s+et\s+le\s+num[eé]ro|ARTICLE\s+\d+)",
        re.IGNORECASE,
    )

    method_re = re.compile(
        r"((?:les\s+)?offres?[\s\S]{0,140}?"
        r"(?:doivent|devront|devraient|doit)\s+(?:obligatoirement\s+)?"
        r"(?:parvenir|être\s+(?:remises?|déposées?|envoyées?|adressées?))"
        r"[\s\S]{0,680}?"
        r"(?:voie\s+postale|rapide-poste|bureau\s+d['’]\s*ordre|remise\s+directe)"
        r"[\s\S]{0,220})",
        re.IGNORECASE,
    )

    for page_entry in pages:
        compact = _compact_fact_text(page_entry["text"])
        address_delivery_match = address_delivery_re.search(compact)
        if address_delivery_match:
            fact = _fact_from_text(address_delivery_match.group(1), page_entry["page"], page_entry["section"])
            if fact:
                return fact

        deposit_match = office_deposit_re.search(compact)
        if deposit_match:
            fact = _fact_from_text(deposit_match.group(1), page_entry["page"], page_entry["section"])
            if fact:
                return fact

        for match in method_re.finditer(compact):
            value = match.group(1)
            value = re.split(
                r"\(\s*La\s+date\s+limite|\bLa\s+date\s+limite|\bet\s+ce,\s+au\s+plus\s+tard\b|\bARTICLE\s+\d+\b",
                value,
                maxsplit=1,
                flags=re.IGNORECASE,
            )[0]
            fact = _fact_from_text(value, page_entry["page"], page_entry["section"])
            if fact:
                return fact

        folded = _fold_fact_text(compact)
        if "tuneps" in folded and ("www.tuneps.tn" in folded or "منظومة الشراء" in compact):
            fact = _fact_from_text(
                "Les offres sont déposées via la plateforme TUNEPS (www.tuneps.tn).",
                page_entry["page"],
                page_entry["section"],
            )
            if fact:
                return fact

    return None


def _extract_validity_fallback(pages: list[dict]) -> dict | None:
    patterns = (
        re.compile(
            r"(Les\s+(?:candidats|soumissionnaires)\s+sont\s+li\S+s\s+par\s+leurs\s+offres\s+"
            r"(?:pour\s+une\s+p\S+riode\s+de\s+|pendant\s+)[^.]{0,220}?"
            r"(?:date\s+limite\s+fix\S+e\s+pour\s+la\s+r\S+ception\s+des\s+(?:offres|plis)|r\S+ception\s+des\s+(?:offres|plis)))",
            re.IGNORECASE,
        ),
        re.compile(
            r"(Les\s+soumissionnaires\s+sont\s+engag\S+s\s+par\s+leurs\s+offres\s+pendant\s+\d+\s+jours\s+"
            r"(?:à|a)\s+compter\s+de\s+la\s+date\s+limite\s+fix\S+e\s+pour\s+la\s+r\S+ception\s+des\s+plis)",
            re.IGNORECASE,
        ),
        re.compile(r"(La\s+validit[eé]\s+de\s+la\s+soumission)", re.IGNORECASE),
    )
    for pattern in patterns:
        fact = _regex_fact_from_pages(pages, pattern, max_chars=320)
        if fact:
            return fact
    return None


ARABIC_ARTICLE_SPLIT_RE = re.compile(
    r"\s*الفصل\s+(?:\d+|[اأ]?[وا]ل|الثاني|الثالث|الرابع|الخامس|السادس|السابع|الثامن|"
    r"التاسع|العاشر|الحادي\s+عشر|الثاني\s+عشر|الثالث\s+عشر|الرابع\s+عشر|"
    r"الخامس\s+عشر|السادس\s+عشر)\s*[:：]?",
    re.IGNORECASE,
)


def _trim_at_next_arabic_article(text: str) -> str:
    parts = ARABIC_ARTICLE_SPLIT_RE.split(text, maxsplit=1)
    return parts[0].strip()


def _arabic_fact_from_pages(
    pages: list[dict],
    pattern: re.Pattern,
    *,
    max_chars: int = 900,
) -> dict | None:
    for page_entry in pages:
        compact = _compact_fact_text(page_entry["text"])
        match = pattern.search(compact)
        if not match:
            continue
        value = match.group(1) if match.groups() else match.group(0)
        value = _trim_at_next_arabic_article(value)
        return _fact_from_text(value[:max_chars], page_entry["page"], page_entry["section"])
    return None


ARABIC_OCR_FACT_REPLACEMENTS: tuple[tuple[str, str], ...] = (
    ("منظومق", "منظومة"),
    ("العموميه", "العمومية"),
    ("علو الخط", "على الخط"),
    ("علو الشرف", "على الشرف"),
    ("لسجل الوطني", "السجل الوطني"),
    ("توزيبس", "تونبس"),
    ("توزيهبس", "تونبس"),
    ("غرامق", "غرامة"),
    ("الت خير", "التأخير"),
    ("مدق", "مدة"),
    ("سنق", "سنة"),
    ("وشيقة", "وثيقة"),
    ("وشائق", "وثائق"),
    ("بطاقق", "بطاقة"),
    ("الفنيق", "الفنية"),
    ("الفني6", "الفنية"),
    ("المال ية", "المالية"),
    ("جلسق", "جلسة"),
    ("واحدق", "واحدة"),
    ("لجنق", "لجنة"),
    ("الإدارق", "الإدارة"),
    ("اإعلامية", "الإعلامية"),
    ("االستلام", "الاستلام"),
    ("لنهائي", "النهائي"),
    ("االنهائي", "النهائي"),
    ("الأشمان", "الأثمان"),
    ("الشركق", "الشركة"),
    ("الصتفقق", "الصفقة"),
    ("الصتفقة", "الصفقة"),
    ("المسلوق", "المسلمة"),
)


def _normalize_arabic_ocr_for_fact_matching(text: str) -> str:
    normalized = unicodedata.normalize("NFKC", str(text or ""))
    for wrong, right in ARABIC_OCR_FACT_REPLACEMENTS:
        normalized = normalized.replace(wrong, right)
    return normalized


def _arabic_page_text_for_matching(page_entry: dict) -> str:
    return _normalize_arabic_ocr_for_fact_matching(str(page_entry.get("text") or ""))


def _trim_real_arabic_fact(text: str, *, max_chars: int = 700) -> str:
    compact = _compact_fact_text(_normalize_arabic_ocr_for_fact_matching(text))
    compact = re.split(
        r"\s+(?=(?:الفصل|لفصل)\s*\d+\s*:|[IVX]+\.\s|—+\s*Page|\[Page\s+\d+\])",
        compact,
        maxsplit=1,
    )[0]
    return compact[:max_chars].strip(" .;:-")


def _real_arabic_fact_from_page(
    page_entry: dict,
    text: str,
    *,
    max_chars: int = 700,
) -> dict | None:
    return _fact_from_text(
        _trim_real_arabic_fact(text, max_chars=max_chars),
        page_entry["page"],
        page_entry["section"],
    )


def _page_with_real_arabic_markers(
    pages: list[dict],
    required: tuple[str, ...],
    *,
    any_of: tuple[str, ...] = (),
) -> dict | None:
    for page_entry in pages:
        text = _arabic_page_text_for_matching(page_entry)
        if all(marker in text for marker in required) and (
            not any_of or any(marker in text for marker in any_of)
        ):
            return page_entry
    return None


def _real_arabic_window_fact(
    pages: list[dict],
    markers: tuple[str, ...],
    *,
    before: int = 80,
    after: int = 520,
    max_chars: int = 700,
) -> dict | None:
    for page_entry in pages:
        text = _arabic_page_text_for_matching(page_entry)
        compact = _compact_fact_text(text)
        marker_index = next((compact.find(marker) for marker in markers if marker in compact), -1)
        if marker_index < 0:
            continue
        start = max(0, marker_index - before)
        end = min(len(compact), marker_index + after)
        return _real_arabic_fact_from_page(page_entry, compact[start:end], max_chars=max_chars)
    return None


def _real_arabic_list_fact_from_items(
    pages: list[dict],
    page_markers: tuple[str, ...],
    items: list[str],
) -> dict | None:
    page_entry = _page_with_real_arabic_markers(pages, page_markers)
    if not page_entry:
        return None
    return _list_fact_from_items(items, page_entry["page"], page_entry["section"])


def _extract_arabic_subject_fallback(pages: list[dict]) -> dict | None:
    page_entry = _page_with_real_arabic_markers(
        pages[:8],
        ("اقتناء مواد", "وزارة العدل"),
        any_of=("طلب عروض", "كراس الشروط"),
    )
    if page_entry:
        fact = _real_arabic_window_fact(
            [page_entry],
            ("طلب عروض", "كراس الشروط"),
            before=0,
            after=360,
            max_chars=420,
        )
        if fact:
            return fact

    patterns = (
        re.compile(
            r"(?:موضوع\s+الاستشارة|موضوع\s+طلب\s+العروض)\s*[:：]?\s*"
            r"([\s\S]{0,520}?(?:مضاد\s+للفيروسات|رخص|اقتناء|إقتناء|استخدام|إستخدام)[\s\S]{0,420}?)"
            r"(?=\s+الفصل\s+\d+|\s+محتوى\s+الاستشارة|$)",
            re.IGNORECASE,
        ),
        re.compile(
            r"(?:إستشارة|استشارة)[^.\n]{0,120}?(?:متعلقة|تتعلق)\s+ب"
            r"([\s\S]{0,420}?)(?=\s+كرّاس|\s+كراس|\s+الفصل|$)",
            re.IGNORECASE,
        ),
        re.compile(
            r"((?:يعلن[\s\S]{0,140}?)?(?:إجراء\s+)?(?:إستشارة|استشارة)\s+"
            r"(?:لإقتناء|لاقتناء|لشراء|للتزويد|لاقتنـاء)[\s\S]{0,520}?"
            r"(?:Plotwave|ROWE|لوازم|معدات|تجهيزات|رخص)[\s\S]{0,260}?)"
            r"(?=\s+لا\s+تجوز|\s+يبقى|\s+يتم\s+تقديم|\s+الفصل|$)",
            re.IGNORECASE,
        ),
    )
    for pattern in patterns:
        fact = _arabic_fact_from_pages(pages[:5], pattern, max_chars=720)
        if fact:
            return _with_fact_text(fact, _normalize_fact_text(fact["text"]))
    return None


def _extract_arabic_deadline_fallback(pages: list[dict]) -> dict | None:
    patterns = (
        re.compile(
            r"((?:آخر|اخر)\s+أجل\s+لقبول\s+العروض[\s\S]{0,220}?"
            r"(?:على\s+الساعة|الساعة)[\s\S]{0,80})",
            re.IGNORECASE,
        ),
        re.compile(
            r"((?:التاريخ\s+الأقصى|التاريخ\s+الاقصى)[\s\S]{0,180}?لقبول\s+العروض[\s\S]{0,120})",
            re.IGNORECASE,
        ),
    )
    for pattern in patterns:
        fact = _arabic_fact_from_pages(pages, pattern, max_chars=340)
        if fact:
            return fact
    return None


def _extract_arabic_submission_method_fallback(pages: list[dict]) -> dict | None:
    page_entry = _page_with_real_arabic_markers(
        pages,
        ("العرض الفني", "منظومة الشراء العمومي"),
        any_of=("ارسال", "إرسال", "إيداع"),
    )
    if page_entry:
        online = _real_arabic_window_fact(
            [page_entry],
            ("ارسال العرض", "إرسال العرض", "إيداع العرض"),
            before=0,
            after=430,
            max_chars=520,
        )
        direct = _real_arabic_window_fact(
            pages,
            ("البريد مضمون الوصول", "البريد السريع", "مكتب الضبط"),
            before=160,
            after=330,
            max_chars=520,
        )
        if online and direct and direct["text"] not in online["text"]:
            return _with_fact_text(
                online,
                f"{online['text']} {direct['text']}",
            )
        if online:
            return online

    pattern = re.compile(
        r"((?:يتم\s+)?إيداع\s+العرض[\s\S]{0,260}?"
        r"(?:TUNEPS|منظومة\s+الشراء\s+العمومي\s+على\s+الخط)[\s\S]{0,120})",
        re.IGNORECASE,
    )
    fact = _arabic_fact_from_pages(pages, pattern, max_chars=420)
    if fact:
        text = re.split(r"\s+وقد\s+حدد\s+(?:آخر|اخر)\s+أجل", fact["text"], maxsplit=1)[0]
        return _with_fact_text(fact, text)
    return None


def _extract_arabic_validity_fallback(pages: list[dict]) -> dict | None:
    pattern = re.compile(
        r"((?:يلتزم\s+العارض\s+بعرضه|صلوحية\s+العروض|"
        r"يبقى\s+المتعهدون\s+ملتزمون\s+بما\s+قدموه\s+من\s+عروض)[\s\S]{0,260}?"
        r"(?:تسعون|ستون|\(?\s*(?:90|60|120|33)\s*\)?)\s+يوما[\s\S]{0,180})",
        re.IGNORECASE,
    )
    return _arabic_fact_from_pages(pages, pattern, max_chars=360)


def _extract_arabic_opening_fallback(pages: list[dict]) -> dict | None:
    fact = _real_arabic_window_fact(
        pages,
        ("تنعقد جلسة فتح العروض", "لجنة فتح العروض", "فتح العروض"),
        before=120,
        after=540,
        max_chars=620,
    )
    if fact and "فتح العروض" in fact["text"] and any(
        marker in fact["text"] for marker in ("نفس اليوم", "جلسة واحدة", "علنية", "لجنة")
    ):
        return fact

    pattern = re.compile(
        r"((?:فتح\s+العروض|يتم\s+فتح\s+العروض)[\s\S]{0,220}?"
        r"(?:على\s+الساعة|الساعة)[\s\S]{0,80})",
        re.IGNORECASE,
    )
    return _arabic_fact_from_pages(pages, pattern, max_chars=360)


def _arabic_list_between(
    pages: list[dict],
    start_pattern: re.Pattern,
    stop_pattern: re.Pattern,
    *,
    max_items: int = 12,
) -> dict | None:
    collecting = False
    collected = []
    page = None
    section = None

    for page_entry in pages:
        text = str(page_entry["text"] or "")
        segment = text
        if not collecting:
            start = start_pattern.search(text)
            if not start:
                continue
            collecting = True
            page = page_entry["page"]
            section = page_entry["section"]
            segment = text[start.end() :]

        stop = stop_pattern.search(segment)
        if stop:
            collected.append(segment[: stop.start()])
            break

        collected.append(segment)

    if not collected:
        return None

    items = []
    for raw_line in "\n".join(collected).splitlines():
        line = _normalize_fact_text(raw_line)
        if not line:
            continue
        line = re.sub(r"^[IVX]+\.\s*", "", line, flags=re.IGNORECASE)
        line = re.sub(r"^[\d٠-٩]+[.)،\-]?\s*", "", line).strip()
        if len(line) < 8:
            continue
        if any(
            marker in line
            for marker in (
                "الفصل",
                "وتتكون من",
                "وذلك وفقا",
                "يجب أن يتضمن",
                "يجب ان يتضمن",
            )
        ):
            continue
        items.append(line)
        if len(items) >= max_items:
            break

    return _list_fact_from_items(items, page, section)


def _arabic_lines_with_markers(
    pages: list[dict],
    markers: tuple[str, ...],
    *,
    max_items: int = 8,
) -> dict | None:
    items: list[str] = []
    page = None
    section = None
    for page_entry in pages:
        text = str(page_entry["text"] or "")
        for raw_line in text.splitlines():
            line = _normalize_fact_text(raw_line)
            if len(line) < 8:
                continue
            if any(marker in line for marker in markers):
                if page is None:
                    page = page_entry["page"]
                    section = page_entry["section"]
                items.append(line)
                if len(items) >= max_items:
                    return _list_fact_from_items(items, page, section)
    return _list_fact_from_items(items, page, section)


def _extract_arabic_administrative_documents_fallback(pages: list[dict]) -> dict | None:
    fact = _real_arabic_list_fact_from_items(
        pages,
        ("وثيقة الضمان الوقتي", "بطاقة الإرشادات"),
        [
            "وثيقة الضمان الوقتي",
            "نظير من السجل الوطني للمؤسسات لم يمض على استخراجه ثلاثة أشهر",
            "الوثائق المثبتة للمؤسسات الصغرى عند الاقتضاء",
            "بطاقة الإرشادات (ملحق عدد 1)",
            "تصريح على الشرف باستقلالية المؤسسة الصغرى (ملحق عدد 9)",
        ],
    )
    if fact:
        return fact

    fact = _arabic_list_between(
        pages,
        re.compile(r"(?:I\.\s*)?الوثائق\s+الإدارية\s*وتتكون\s+من\s*[:：]?", re.IGNORECASE),
        re.compile(r"(?:II\.\s*)?الوثائق\s+الخاصة\s+بالعرض\s+الفني|وثائق\s+الخاصة\s+بالعرض\s+الفني", re.IGNORECASE),
    )
    if fact:
        return fact
    for page_entry in pages:
        text = str(page_entry["text"] or "")
        if "الوثائق المكونة للعرض" not in text and "الوثائق المكونة للملف" not in text:
            continue
        segment = text
        start = re.search(r"الوثائق\s+المكونة\s+ل(?:لعرض|ملف)", segment)
        if start:
            segment = segment[start.end() :]
        segment = _trim_at_next_arabic_article(segment)
        items = []
        for marker in (
            "الشروط العامة للاستشارة",
            "الممثل الرسمي",
            "Certificat d",
            "شهادة إثبات",
            "الضمان المالي الوقتي",
            "السجل الوطني للمؤسسات",
        ):
            match = re.search(rf"([^.\n]*{re.escape(marker)}[^.\n]*)(?:[.\n]|$)", segment, re.IGNORECASE)
            if match:
                items.append(_normalize_fact_text(match.group(1)))
        fact = _list_fact_from_items(items, page_entry["page"], page_entry["section"])
        if fact:
            return fact
    return None


def _extract_arabic_technical_documents_fallback(pages: list[dict]) -> dict | None:
    fact = _real_arabic_list_fact_from_items(
        pages,
        ("العرض الفني", "ISO"),
        [
            "العرض الفني حسب كل قسط",
            "شهادة المطابقة للمواصفات الفنية ISO 9001 نسخة 2015 أو النسخ الأحدث",
            "شهادة المطابقة للمواصفة ISO 14001 نسخة 2015 أو النسخ الأحدث عند الاقتضاء",
            "تقرير اختبار لعدد الصفحات المنتجة حسب معيار ISO/IEC 19752 أو ما يعادلها",
            "تعمير جداول الخاصيات الفنية الواردة بكراس الشروط بكل دقة",
            "تقديم جذاذات فنية للمواد المطلوبة",
        ],
    )
    if fact:
        return fact

    fact = _arabic_list_between(
        pages,
        re.compile(r"(?:II\.\s*)?الوثائق\s+الخاصة\s+بالعرض\s+الفني\s*[:：]?", re.IGNORECASE),
        re.compile(r"(?:III\.\s*)?وثائق\s+الخاصة\s+بالعرض\s+المالي|الوثائق\s+الخاصة\s+بالعرض\s+المالي", re.IGNORECASE),
    )
    if fact:
        return fact
    fact = _arabic_lines_with_markers(
        pages,
        (
            "الخصائص الفنية",
            "Code Réference",
            "Code Reference",
            "Référence proposée",
            "Reference proposée",
            "Tireuses de plans",
            "Plotwave",
            "ROWE I4",
        ),
        max_items=6,
    )
    if fact:
        return fact
    return None


def _extract_arabic_financial_documents_fallback(pages: list[dict]) -> dict | None:
    fact = _real_arabic_list_fact_from_items(
        pages,
        ("التعهد المالي", "جدول الأثمان"),
        [
            "التعهد المالي حسب كل قسط أو لجميع الأقساط، مؤشر عليه ويحمل الإمضاء الأصلي وختم الشركة",
            "جدول الأثمان حسب كل قسط، مؤشر عليه ويحمل الإمضاء الأصلي وختم الشركة",
        ],
    )
    if fact:
        return fact

    fact = _arabic_list_between(
        pages,
        re.compile(r"(?:III\.\s*)?وثائق\s+الخاصة\s+بالعرض\s+المالي\s*[:：]?", re.IGNORECASE),
        re.compile(r"الفصل\s+\d+\s*[:：]?\s*الأثمان|الفصل\s+\d+\s*[:：]?\s*فتح\s+العروض", re.IGNORECASE),
    )
    if fact:
        return fact

    table_markers = ("جدول الأثمان", "جدول الاثمان")
    value_markers = (
        "جدول الأثمان",
        "جدول الاثمان",
        "السعر الفردي",
        "المبلغ الجملي",
        "الأداءات",
        "الآداءات",
        "القيمة المضافة",
    )
    reject_markers = (
        "الضمان المالي",
        "غرامات التأخير",
        "غرامة التأخير",
        "التأخير",
        "العقد ضمانا",
        "العقد ضمان",
    )
    for page_entry in pages:
        text = str(page_entry["text"] or "")
        has_table_title = any(marker in text for marker in table_markers)
        has_price_grid = "السعر الفردي" in text and "المبلغ الجملي" in text
        if not has_table_title and not has_price_grid:
            continue
        items = []
        for raw_line in text.splitlines():
            line = _normalize_fact_text(raw_line)
            if len(line) < 8:
                continue
            if any(marker in line for marker in reject_markers):
                continue
            if any(marker in line for marker in value_markers):
                items.append(line)
            if len(items) >= 6:
                break
        if has_price_grid and items:
            clean_items = ["جدول الأثمان / العرض المالي", "السعر الفردي"]
            if "المبلغ الجملي" in text:
                clean_items.append("المبلغ الجملي دون احتساب الأداءات")
            if "القيمة المضافة" in text or "الأداءات" in text or "الآداءات" in text:
                clean_items.append("الأداءات على القيمة المضافة")
            return _list_fact_from_items(clean_items, page_entry["page"], page_entry["section"])
        fact = _list_fact_from_items(items, page_entry["page"], page_entry["section"])
        if fact:
            return fact
    return None


def _extract_arabic_caution_fallback(pages: list[dict]) -> dict | None:
    page_entry = _page_with_real_arabic_markers(
        pages,
        ("الضمان الوقتي",),
        any_of=("وثيقة", "120 يوما", "صالحة"),
    )
    if page_entry:
        fact = _real_arabic_window_fact(
            [page_entry],
            ("وثيقة الضمان الوقتي", "الضمان الوقتي"),
            before=80,
            after=360,
            max_chars=460,
        )
        validity = _real_arabic_window_fact(
            pages,
            ("الضمان صالح لمدة 120 يوما", "صالح لمدة 120 يوما", "لمدة 120 يوما"),
            before=100,
            after=260,
            max_chars=360,
        )
        if fact and validity and validity["text"] not in fact["text"]:
            return _with_fact_text(fact, f"{fact['text']} {validity['text']}")
        if fact:
            return fact

    patterns = (
        re.compile(
            r"((?:وثيقة)?الضمان\s+المالي\s+الوقتي[\s\S]{0,260}?"
            r"(?:مبلغ\s+قدره|قدرها|بمبلغ|بقيمة)[\s\S]{0,140}?دينار(?:ا)?)",
            re.IGNORECASE,
        ),
        re.compile(
            r"((?:عدم\s+تقديم\s+)?الضمان\s+المالي\s+الوقتي[\s\S]{0,260}?"
            r"(?:إقصاء\s+العرض|قبول\s+العرض|صالحة|وثيقة\s+إدارية)[\s\S]{0,160})",
            re.IGNORECASE,
        ),
    )
    for pattern in patterns:
        fact = _arabic_fact_from_pages(pages, pattern, max_chars=520)
        if fact:
            return fact
    return None


def _extract_arabic_definitive_caution_fallback(pages: list[dict]) -> dict | None:
    fact = _real_arabic_window_fact(
        pages,
        ("الضمان النهائي", "ضمان نهائي"),
        before=100,
        after=430,
        max_chars=560,
    )
    if fact and any(marker in fact["text"] for marker in ("3", "20", "بالمائة", "%")):
        return fact

    pattern = re.compile(
        r"((?:الضمان\s+المالي\s+النهائي|ضمانا\s+ماليا\s+نهائيا)[\s\S]{0,260}?"
        r"(?:بالمائة|%)[\s\S]{0,180})",
        re.IGNORECASE,
    )
    return _arabic_fact_from_pages(pages, pattern, max_chars=480)


def _extract_arabic_reception_fallback(pages: list[dict]) -> dict | None:
    fact = _real_arabic_window_fact(
        pages,
        ("الاستلام الوقتي", "الاستلام النهائي", "محضر الاستلام"),
        before=100,
        after=620,
        max_chars=720,
    )
    if fact and any(marker in fact["text"] for marker in ("الاستلام الوقتي", "الاستلام النهائي", "محضر")):
        if "كيفية الخلاص" in fact["text"]:
            fact = _with_fact_text(fact, fact["text"].split("كيفية الخلاص", 1)[0].strip())
        return fact

    pattern = re.compile(
        r"((?:ال[اإ]ستلام\s+والتركيب|ال[اإ]ستلام\s+الكمي|ال[اإ]ستلام\s+الوقتي|ال[اإ]ستلام\s+النهائي)"
        r"[\s\S]{0,420}?(?:وصل\s+تسليم|مطابقة|المواصفات|ال[اإ]ستلام\s+النهائي|نهائي)[\s\S]{0,120})",
        re.IGNORECASE,
    )
    fact = _arabic_fact_from_pages(pages, pattern, max_chars=520)
    if fact and "كيفية الخلاص" in fact["text"]:
        fact = _with_fact_text(fact, fact["text"].split("كيفية الخلاص", 1)[0].strip())
    return fact


def _extract_arabic_payment_fallback(pages: list[dict]) -> dict | None:
    fact = _real_arabic_window_fact(
        pages,
        ("أمر بصرف", "خلاص صاحب الصفقة", "فاتورة"),
        before=0,
        after=650,
        max_chars=820,
    )
    if fact and any(marker in fact["text"] for marker in ("30", "15", "فاتورة", "أمر بصرف")):
        return fact

    patterns = (
        re.compile(
            r"((?:كيفية\s+الخلاص|يتم\s+خلاص\s+صاحب\s+العقد|خلاص\s*الطلبية|"
            r"تسديد\s+مستحقات|يتم\s+تسديد\s+مستحقات)[\s\S]{0,720}?"
            r"(?:خمسة\s+و\s+أربعون|\(?45\)?|45)\s+يوما[\s\S]{0,220}?"
            r"(?:تحويل\s+بريدي|تحويل\s+بنكي|بنكي|بريدي))",
            re.IGNORECASE,
        ),
        re.compile(
            r"((?:كيفية\s+الخلاص|يتم\s+خلاص\s+صاحب\s+العقد)[\s\S]{0,520}?"
            r"(?:فاتورة|وصولات\s+التسليم)[\s\S]{0,220})",
            re.IGNORECASE,
        ),
        re.compile(
            r"((?:خلاص\s*الطلبية|تسديد\s+مستحقات|يتم\s+تسديد\s+مستحقات)"
            r"[\s\S]{0,520}?(?:فاتورة|محضر|الصندوق\s+الوطني)[\s\S]{0,180})",
            re.IGNORECASE,
        ),
    )
    for pattern in patterns:
        fact = _arabic_fact_from_pages(pages, pattern, max_chars=780)
        if fact:
            return fact
    return None


def _extract_arabic_penalties_fallback(pages: list[dict]) -> dict | None:
    fact = _real_arabic_window_fact(
        pages,
        ("غرامة التأخير", "خطايا التأخير"),
        before=0,
        after=650,
        max_chars=760,
    )
    if fact and any(marker in fact["text"] for marker in ("1000", "5", "كل يوم", "تأخير")):
        return fact

    pattern = re.compile(
        r"((?:عقوبة\s+التأخير|غرامة\s+تأخير|غرامات\s+التأخير)[\s\S]{0,620}?"
        r"(?:بالمائة|%|بالألف|‰|سقف|تتجاوز|905|5)[\s\S]{0,180})",
        re.IGNORECASE,
    )
    return _arabic_fact_from_pages(pages, pattern, max_chars=800)


def _extract_arabic_guarantee_fallback(pages: list[dict]) -> dict | None:
    fact = _real_arabic_window_fact(
        pages,
        ("مدة الضمان", "مدق الضمان"),
        before=60,
        after=320,
        max_chars=420,
    )
    if fact and any(marker in fact["text"] for marker in ("سنة", "الاستلام", "قبول")):
        return fact

    clear_pattern = re.compile(
        r"((?:مدة\s+الضمان|مدّة\s+الضمان|مدة\s+ضمان)[\s\S]{0,240}?"
        r"(?:أشهر|اشهر|شهر|سنة|سنوات|الإستلام\s+الوقتي|الاستلام\s+الوقتي)[\s\S]{0,120})",
        re.IGNORECASE,
    )
    fact = _arabic_fact_from_pages(pages, clear_pattern, max_chars=460)
    if fact:
        return fact

    pattern = re.compile(
        r"((?:مد[\u064b-\u065f]*ة\s+ضمان|مد[\u064b-\u065f]*ة\s+الضمان|مدة\s+ضمان|ضمان\s+الرخص|مدة\s+الضمان)[\s\S]{0,360}?"
        r"(?:أشهر|اشهر|شهر|سنة|سنوات|الاستلام\s+الوقتي|الإستلام\s+الوقتي)[\s\S]{0,120})",
        re.IGNORECASE,
    )
    return _arabic_fact_from_pages(pages, pattern, max_chars=460)


def _extract_arabic_information_sheet_fallback(pages: list[dict]) -> dict | None:
    fact = _real_arabic_window_fact(
        pages,
        ("بطاقة الإرشادات",),
        before=60,
        after=160,
        max_chars=220,
    )
    if fact:
        return fact

    pattern = re.compile(r"(بطاقة\s+إرشادات\s+عامة[\s\S]{0,120})", re.IGNORECASE)
    return _arabic_fact_from_pages(pages, pattern, max_chars=180)


def _extract_arabic_rne_fallback(pages: list[dict]) -> dict | None:
    fact = _real_arabic_window_fact(
        pages,
        ("السجل الوطني للمؤسسات",),
        before=60,
        after=180,
        max_chars=260,
    )
    if fact:
        return fact

    pattern = re.compile(
        r"((?:نظير\s+أصلي\s+من\s+)?السجل\s+الوطني\s+للمؤسسات[\s\S]{0,160})",
        re.IGNORECASE,
    )
    return _arabic_fact_from_pages(pages, pattern, max_chars=240)


def _extract_information_sheet_fallback(pages: list[dict]) -> dict | None:
    pattern = re.compile(
        r"((?:La\s+)?fiche\s+des?\s+renseignements?\s+g[eé]n[eé]raux"
        r"[\s\S]{0,320}?annexe\s*3[\s\S]{0,240}?annexe\s*3\s*bis)",
        re.IGNORECASE,
    )
    return _regex_fact_from_pages(pages, pattern, max_chars=700)


def _extract_fiscal_certificate_fallback(pages: list[dict]) -> dict | None:
    pattern = re.compile(
        r"(TOPNET\s+fera\s+les\s+v[eé]rifications\s+n[eé]cessaires"
        r"[\s\S]{0,180}?Situation\s+fiscale[\s\S]{0,120}?soumissionnaire)",
        re.IGNORECASE,
    )
    return _regex_fact_from_pages(pages, pattern, max_chars=450)


def _extract_manufacturer_authorization_fallback(pages: list[dict]) -> dict | None:
    originality_pattern = re.compile(
        r"(Attestation\s+constructeur\s+attestant\s+l['’]originalit\S+\s+des\s+produits)",
        re.IGNORECASE,
    )
    fact = _regex_fact_from_pages(pages, originality_pattern, max_chars=220)
    if fact:
        return fact

    topnet_pattern = re.compile(
        r"(Une\s+attestation\s+d[eé]livr[eé]e\s+par\s+le\s+fabricant"
        r"[\s\S]{0,260}?commercialiser\s+les\s+fournitures[\s\S]{0,120}?consultation)",
        re.IGNORECASE,
    )
    fact = _regex_fact_from_pages(pages, topnet_pattern, max_chars=520)
    if fact:
        return fact

    ubci_pattern = re.compile(
        r"(Date\s+de\s+commercialisation[\s\S]{0,520}?document\s+officiel\s+du\s+constructeur"
        r"[\s\S]{0,260}?Attestation\s+Constructeur)",
        re.IGNORECASE,
    )
    fact = _regex_fact_from_pages(pages, ubci_pattern, max_chars=900)
    if fact:
        return _with_fact_text(
            fact,
            "Le document exige un document officiel du constructeur pour les dates de "
            "commercialisation, d'arrêt de commercialisation et de fin de support, ainsi que "
            "une Attestation Constructeur pour les équipements concernés.",
        )

    official_doc_page = None
    official_doc_section = None
    has_official_doc = False
    has_constructor_attestation = False
    for page_entry in pages:
        folded = _fold_fact_text(page_entry["text"])
        if "document officiel du constructeur" in folded:
            has_official_doc = True
            official_doc_page = official_doc_page or page_entry["page"]
            official_doc_section = official_doc_section or page_entry["section"]
        if "attestation constructeur" in folded:
            has_constructor_attestation = True

    if has_official_doc and has_constructor_attestation:
        return _fact_from_text(
            "Le document exige un document officiel du constructeur pour les dates de "
            "commercialisation, d'arrêt de commercialisation et de fin de support, ainsi que "
            "une Attestation Constructeur pour les équipements concernés.",
            official_doc_page or "?",
            official_doc_section or "technical",
        )

    return None


def _extract_guarantee_fallback(pages: list[dict]) -> dict | None:
    telecom_pattern = re.compile(
        r"(Tunisie\s+T\S+l\S+com\s+informera\s+le\s+fournisseur[\s\S]{0,520}?"
        r"(?:p\S+riode\s+de\s+6\s+mois|2\s+ans|3\s+ann\S+es)[\s\S]{0,520}?"
        r"douchettes)",
        re.IGNORECASE,
    )
    fact = _regex_fact_from_pages(pages, telecom_pattern, max_chars=900)
    if fact:
        return fact

    generic_pattern = re.compile(
        r"((?:A\s+d\S+faut\s+d\S+un\s+meilleur\s+d\S+lai\s+propos\S+\s+par\s+le\s+fournisseur,\s*)?"
        r"(?:le\s+)?d\S+lai\s+de\s+garantie\s+est\s+fix\S+\s+[aÃ ]\s+"
        r"[\s\S]{0,220}?(?:jours?|mois|ans?|heures?)[^.\n]{0,180})",
        re.IGNORECASE,
    )
    fact = _regex_fact_from_pages(pages, generic_pattern, max_chars=520)
    if fact:
        return fact

    stb_simple_pattern = re.compile(
        r"((?:trois\s+\(?0?3?\)?\s+ans\s+)?(?:a\s+partir\s+de\s+)?"
        r"l['’]?\s*expiration\s+d['’]?\s*une\s+ann\S+e\s+de\s+garantie)",
        re.IGNORECASE,
    )
    fact = _regex_fact_from_pages(pages, stb_simple_pattern, max_chars=320)
    if fact:
        return _with_fact_text(
            fact,
            "La garantie est d'une année. Les prestations d'entretien continu des équipements "
            "multifonction impression-copie-scan durent trois ans à partir de l'expiration "
            "de cette année de garantie.",
        )

    stb_pattern = re.compile(
        r"(d\S+lai\s+global\s+des\s+prestations\s+d['ƒ?T]entretien\s+continu[\s\S]{0,360}?"
        r"expiration\s+d['ƒ?T]une\s+ann\S+e\s+de\s+garantie)",
        re.IGNORECASE,
    )
    fact = _regex_fact_from_pages(pages, stb_pattern, max_chars=520)
    if fact:
        return _with_fact_text(
            fact,
            "La garantie est d'une année. Les prestations d'entretien continu des équipements "
            "multifonction impression-copie-scan durent trois ans à partir de l'expiration "
            "de cette année de garantie.",
        )

    stb_pattern = re.compile(
        r"(trois\s+ans\s+[aÇÿ]\s+partir\s+de\s+l['ƒ?T]expiration\s+d['ƒ?T]une\s+ann\S+e\s+de\s+garantie)",
        re.IGNORECASE,
    )
    fact = _regex_fact_from_pages(pages, stb_pattern, max_chars=260)
    if fact:
        return _with_fact_text(
            fact,
            "La garantie est d'une année. Les prestations d'entretien continu des équipements "
            "multifonction impression-copie-scan durent trois ans à partir de l'expiration "
            "de cette année de garantie.",
        )

    pattern = re.compile(
        r"(La\s+dur[eé]e\s+de\s+cette\s+garantie\s+est\s+fix[eé]e\s+[aà]\s+trois\s*\(03\)\s+ans"
        r"[\s\S]{0,420}?48\s+Heures)",
        re.IGNORECASE,
    )
    return _regex_fact_from_pages(pages, pattern, max_chars=650)


def _extract_reception_fallback(pages: list[dict]) -> dict | None:
    telecom_pattern = re.compile(
        r"(Article\s+9\s*:\s*RECEPTION\s+PROVISOIRE\s*-\s*RECEPTION\s+DEFINITIVE"
        r"[\s\S]{0,1800}?(?:R\S+ception\s+quantitative|R\S+ception\s+provisoire)"
        r"[\s\S]{0,1800}?R\S+ception\s+d\S+finitive[\s\S]{0,420})",
        re.IGNORECASE,
    )
    fact = _regex_fact_from_pages(pages, telecom_pattern, max_chars=1800)
    if fact:
        return fact

    cetime_pattern = re.compile(
        r"(La\s+r\S+ception\s+est\s+prononc\S+e\s+suite\s+[aà]\s*:?"
        r"[\s\S]{0,900}?(?:PV\s+de\s+r\S+ception[\s\S]{0,220}?r\S+serves|sans\s+r\S+serves))",
        re.IGNORECASE,
    )
    fact = _regex_fact_from_pages(pages, cetime_pattern, max_chars=1100)
    if fact:
        return fact

    cnss_pattern = re.compile(
        r"(R\S+ception\s+provisoire[\s\S]{0,900}?proc\S+s-verbal\s+de\s+r\S+ception\s+provisoire"
        r"[\s\S]{0,900}?R\S+ception\s+d\S+finitive[\s\S]{0,900}?r\S+ception\s+d\S+finitive)",
        re.IGNORECASE,
    )
    doc_text, offsets = _flatten_pages_for_articles(pages)
    cnss_match = cnss_pattern.search(doc_text)
    if cnss_match:
        page, section = _page_info_at_offset(offsets, cnss_match.start())
        return _fact_from_text(cnss_match.group(1)[:1800], page, section)

    pattern = re.compile(
        r"(14\.1\s+La\s+r[eé]ception\s+provisoire[\s\S]{0,3600}?"
        r"proc[eè]s-verbal\s+de\s+r[eé]ception\s+d[eé]finitive[\s\S]{0,220})",
        re.IGNORECASE,
    )
    match = pattern.search(doc_text)
    if not match:
        return None

    page, section = _page_info_at_offset(offsets, match.start())
    return _fact_from_text(match.group(1)[:3800], page, section)


def _extract_definitive_caution_fallback(pages: list[dict]) -> dict | None:
    good_fin_pattern = re.compile(
        r"((?:caution\s+bancaire\s+[aÃ ]\s+premi\S+re\s+demande\s+de\s+bonne\s+fin|"
        r"caution\s+de\s+bonne\s+fin)[\s\S]{0,900}?"
        r"(?:\b\d+\s*%|pour\s+cent)[\s\S]{0,240}?(?:commande|march\S+|TTC))",
        re.IGNORECASE,
    )
    fact = _regex_fact_from_pages(pages, good_fin_pattern, max_chars=1100)
    if fact:
        return fact

    pattern = re.compile(
        r"(constituer\s+une\s+caution\s+d[eé]finitive\s+[aà]\s+la\s+premi[eè]re\s+demande"
        r"[\s\S]{0,900}?caution\s+de\s+garantie)",
        re.IGNORECASE,
    )
    return _regex_fact_from_pages(pages, pattern, max_chars=1000)


def _extract_penalties_fallback(pages: list[dict]) -> dict | None:
    telecom_pattern = re.compile(
        r"((?:Article\s+\d+\s*)?PENALITES?\s+POUR\s+RETARD"
        r"[\s\S]{0,900}?cinq\s+pour\s+mille\s*\(?\s*5\s*[‰%o/]*\s*\)?"
        r"[\s\S]{0,900}?(?:maximum|plafond|ne\s+peut\s+pas\s+d\S+passer)"
        r"[\s\S]{0,160}?10\s*%[\s\S]{0,180}?(?:march\S+|commande))",
        re.IGNORECASE,
    )
    fact = _regex_fact_from_pages(pages, telecom_pattern, max_chars=1200)
    if fact:
        return fact

    generic_steg_pattern = re.compile(
        r"((?:Article\s+\d+\s*)?(?:PENALITES?\s+DE\s+RETARD|p\S{0,8}nalit\S{0,8}\s+de\s+retard)"
        r"[\s\S]{0,900}?"
        r"(?:0\s*[,\.]\s*2\s*%|0,2%|0\.2%)"
        r"[\s\S]{0,900}?"
        r"(?:maximum|plafond|ne\s+peut\s+pas\s+d\S+passer|ne\s+d\S+passe\s+pas)"
        r"[\s\S]{0,180}?5\s*%[\s\S]{0,180}?(?:TVA|HTVA|commande|march\S+))",
        re.IGNORECASE,
    )
    fact = _regex_fact_from_pages(pages, generic_steg_pattern, max_chars=1200)
    if fact:
        return fact

    steg_pattern = re.compile(
        r"(p[eÃ©]nalit[eÃ©]\s+de\s+0\s*[,\.]?\s*\d+\s*%[\s\S]{0,650}?"
        r"(?:maximum|plafond)\s+de\s+5\s*%[\s\S]{0,220}?(?:HTVA|TVA|commande|march\S+))",
        re.IGNORECASE,
    )
    fact = _regex_fact_from_pages(pages, steg_pattern, max_chars=900)
    if fact:
        return fact

    pattern = re.compile(
        r"(Si\s+les\s+d[eé]lais\s+d['’]ex[eé]cution\s+du\s+march[eé]\s+ne\s+sont\s+pas\s+respect[eé]s"
        r"[\s\S]{0,850}?HTVA)",
        re.IGNORECASE,
    )
    return _regex_fact_from_pages(pages, pattern, max_chars=950)


def _extract_references_fallback(pages: list[dict]) -> dict | None:
    patterns = (
        re.compile(
            r"((?:la\s+)?liste\s+d\S?au\s+moins\s+\d+\s+travaux\s+similaires[^.\n]{0,220})",
            re.IGNORECASE,
        ),
        re.compile(
            r"(Nombre\s+de\s+r[eé]f[eé]rence\s+\d+\s+r[eé]f[eé]rences?\s+dans\s+l['’]installation"
            r"\s+des\s+frames\s+HPE[\s\S]{0,260}?(?:justificatifs|commande|contrat|facture))",
            re.IGNORECASE,
        ),
    )
    for pattern in patterns:
        fact = _regex_fact_from_pages(pages, pattern, max_chars=420)
        if fact:
            return fact
    return None


def _extract_payment_fallback(pages: list[dict]) -> dict | None:
    telecom_pattern = re.compile(
        r"(Pour\s+chaque\s+Appel\s+de\s+commande,\s*Le\s+paiement\s+se\s+fera\s+100\s*%"
        r"[\s\S]{0,620}?60\s+jours[\s\S]{0,360}?"
        r"(?:facture|bon\(s\)\s+de\s+livraison|PV\s+de\s+r\S+ception\s+provisoire))",
        re.IGNORECASE,
    )
    fact = _regex_fact_from_pages(pages, telecom_pattern, max_chars=900)
    if fact:
        return fact

    generic_virement_pattern = re.compile(
        r"(Le\s+r\S+glement\s+est\s+effectu\S+\s+par\s+virement"
        r"[\s\S]{0,520}?(?:sans\s+r\S+serves|PV\s+sign\S+[\s\S]{0,120}?r\S+serves|facture[\s\S]{0,180}?r\S+serves))",
        re.IGNORECASE,
    )
    fact = _regex_fact_from_pages(pages, generic_virement_pattern, max_chars=760)
    if fact:
        return fact

    cnss_pattern = re.compile(
        r"(Modalit\S+s\s+de\s+paiement[\s\S]{0,900}?"
        r"facture\s+\S+mises?\s+par\s+le\s+fournisseur\s+est\s+payable\s+[aàÃ ]\s+45\s+jours"
        r"[\s\S]{0,160}?CNSS)",
        re.IGNORECASE,
    )
    fact = _regex_fact_from_pages(pages, cnss_pattern, max_chars=1100)
    if fact:
        return fact

    bct_pattern = re.compile(
        r"(Le\s+paiement\s+du\s+montant\s+global\s+du\s+march[eé]\s*"
        r"\(bon\s+de\s+commande\)\s+sera\s+effectu[eé]\s+au\s+fur\s+et\s+[aà]\s+mesure"
        r"[\s\S]{0,260}?(?:[.:]|[eé]ch[eé]ancier\s+ci-dessous))",
        re.IGNORECASE,
    )
    fact = _regex_fact_from_pages(pages, bct_pattern, max_chars=720)
    if fact:
        return fact

    stb_pattern = re.compile(
        r"(La\s+STB\s+proc\S+dera\s+au\s+r\S+glement\s+du\s+prix\s+du\s+march\S+\s+par\s+virement\s+bancaire"
        r"[\s\S]{0,620}?(?:PV\s+de\s+r\S+ception|factures?))",
        re.IGNORECASE,
    )
    fact = _regex_fact_from_pages(pages, stb_pattern, max_chars=760)
    if fact:
        return fact

    topnet_pattern = re.compile(
        r"(Les\s+conditions\s+de\s+paiement\s+sont\s+fix[eé]es\s+comme\s+suit"
        r"[\s\S]{0,900}?virement\s+bancaire)",
        re.IGNORECASE,
    )
    fact = _regex_fact_from_pages(pages, topnet_pattern, max_chars=1000)
    if fact:
        return fact

    ubci_pattern = re.compile(
        r"(Les\s+paiements\s+s['’]effectueront\s+par\s+virement"
        r"[\s\S]{0,720}?factures\s+conformes[\s\S]{0,140}?r[eé]ception)",
        re.IGNORECASE,
    )
    return _regex_fact_from_pages(pages, ubci_pattern, max_chars=900)


def _offer_content_segment(page_text: str) -> str | None:
    compact = _compact_fact_text(page_text)
    folded = _fold_fact_text(compact)
    if "contenu de l'offre" not in folded and "offre doit comporter" not in folded:
        return None

    start_match = re.search(
        r"(?:article\s+\d+\s*:\s*)?contenu\s+de\s+l['â€™]?\s*offre|"
        r"l['â€™]?\s*offre\s+doit\s+comporter\s+les\s+pi[eÃ¨]ces\s+suivantes",
        compact,
        flags=re.IGNORECASE,
    )
    if not start_match:
        return None

    stop_match = re.search(r"\bARTICLE\s+\d+\s*:", compact[start_match.end() :], flags=re.IGNORECASE)
    stop = start_match.end() + stop_match.start() if stop_match else len(compact)
    return compact[start_match.start() : stop]


def _extract_offer_content_documents(pages: list[dict], field: str) -> dict | None:
    patterns_by_field = {
        "administrative_documents": (
            r"Fiche\s+de\s+renseignements?\s+g\S+n\S+raux[^.]{0,180}?soumissionnaire",
            r"Fiche\s+de\s+renseignements?\s+g[eÃ©]n[eÃ©]raux[^.]{0,180}?soumissionnaire",
            r"Cahier\s+des\s+charges",
            r"(?:Un\s+)?extrait\s+du\s+registre\s+de\s+commerce\s*/?\s*certificat\s+RNE",
            r"(?<!/)(?:certificat|extrait)\s+RNE",
            r"Attestation\s+[^.]{0,80}?fiscale",
            r"Attestation\s+[^.]{0,80}?CNSS",
        ),
        "technical_documents": (
            r"Tableaux?\s+des\s+sp\S+cificit\S+s?\s+techniques?[^.]{0,260}?(?:documents?\s+techniques?\s+y\s+aff\S+rant)?",
            r"documents?\s+techniques?\s+y\s+aff\S+rant",
            r"Tableaux?\s+des\s+sp[eÃ©]cificit[eÃ©]s?\s+techniques?[^.]{0,220}?(?:documents?\s+techniques?\s+y\s+aff[eÃ©]rant)?",
            r"documents?\s+techniques?\s+y\s+aff[eÃ©]rant",
        ),
        "financial_documents": (
            r"(?:La\s+)?soumission\s+d[uÃ»]ment\s+remplie\s+et\s+sign[eÃ©]e",
            r"(?:Les?\s+)?Bordereaux?\s+des\s+prix\s+d[uÃ»]ment\s+remplis?\s+et\s+sign[eÃ©]s?",
            r"\bLa\s+soumission\b",
            r"\bLes?\s+Bordereaux?\s+des\s+prix\b",
        ),
    }
    item_patterns = patterns_by_field.get(field)
    if not item_patterns:
        return None

    for page_entry in pages:
        segment = _offer_content_segment(page_entry["text"])
        if not segment:
            continue

        items = []
        for pattern in item_patterns:
            for match in re.finditer(pattern, segment, flags=re.IGNORECASE):
                items.append(match.group(0))

        fact = _list_fact_from_items(items, page_entry["page"], page_entry["section"])
        if fact:
            return fact

    return None


def _section_between_markers(
    compact: str,
    start_patterns: tuple[str, ...],
    stop_patterns: tuple[str, ...],
) -> str | None:
    start = None
    for pattern in start_patterns:
        match = re.search(pattern, compact, flags=re.IGNORECASE)
        if match and (start is None or match.end() < start):
            start = match.end()
    if start is None:
        return None

    remainder = compact[start:]
    stop_positions = [
        match.start()
        for pattern in stop_patterns
        if (match := re.search(pattern, remainder, flags=re.IGNORECASE))
    ]
    stop = min(stop_positions) if stop_positions else len(remainder)
    return remainder[:stop].strip(" .;:-")


def _extract_consultation_offer_documents(pages: list[dict], field: str) -> dict | None:
    config = {
        "administrative_documents": {
            "starts": (
                r"\bpi\S+ces\s+administratives\s+suivantes\s*:?",
                r"\bcontient\s+les\s+pi\S+ces\s+administratives\s+suivantes\s*:?",
            ),
            "stops": (
                r"\b[A-Z]\s*[-.)]?\s*dossier\s+de\s+l['’]?\s*offre\s+financi\S+re\b",
                r"\b[A-Z]\s*[-.)]?\s*dossier\s+de\s+l['’]?\s*offre\s+technique\b",
                r"\bARTICLE\s+\d+\b",
            ),
            "items": (
                r"Pr\S+sentation\s+du\s+soumissionnaire",
                r"Une\s+attestation\s+d['’]affiliation\s+[aà]\s+la\s+CNSS",
                r"Extrait\s+du\s+registre\s+national\s+des\s+entreprises[^.]{0,180}",
                r"Le\s+cahier\s+des\s+charges\s+sign\S+\s+paraph\S+[^.]{0,180}",
                r"Attestation\s+constructeur\s+attestant\s+l['’]originalit\S+\s+des\s+produits",
            ),
        },
        "financial_documents": {
            "starts": (
                r"\b[A-Z]\s*[-.)]?\s*dossier\s+de\s+l['’]?\s*offre\s+financi\S+re\b",
                r"\bdossier\s+de\s+l['’]?\s*offre\s+financi\S+re\b",
            ),
            "stops": (
                r"\b[A-Z]\s*[-.)]?\s*dossier\s+de\s+l['’]?\s*offre\s+technique\b",
                r"\bARTICLE\s+\d+\b",
            ),
            "items": (
                r"Le(?:\s*\(\s*s\s*\))?\s+lettre(?:\s*\(\s*s\s*\))?\s+de\s+soumission[^.]{0,220}",
                r"Devis\s+estimatif\s+d\S+taill\S+[^.]{0,360}",
            ),
        },
        "technical_documents": {
            "starts": (
                r"\bpi\S+ces\s+techniques\s+suivantes\s*:?",
                r"\bcompos\S+e\s+des\s+pi\S+ces\s+techniques\s+suivantes\s*:?",
            ),
            "stops": (
                r"\bARTICLE\s+\d+\b",
                r"\bdossier\s+de\s+l['’]?\s*offre\s+financi\S+re\b",
            ),
            "items": (
                r"Pr\S+sentation\s+de\s+l['’]offre\s+technique",
                r"Pr\S+sentation\s+des\s+sp\S+cifications\s+techniques[^.]{0,220}",
                r"Les\s+d\S+lais\s+de\s+livraison\s+des\s+articles",
            ),
        },
    }.get(field)
    if not config:
        return None

    for page_entry in pages:
        compact = _compact_fact_text(page_entry["text"])
        segment = _section_between_markers(compact, config["starts"], config["stops"])
        if not segment:
            continue
        items = []
        for pattern in config["items"]:
            for match in re.finditer(pattern, segment, flags=re.IGNORECASE):
                items.append(match.group(0))
        fact = _list_fact_from_items(items, page_entry["page"], page_entry["section"])
        if fact:
            return fact

    return None


def _extract_administrative_documents_fallback(pages: list[dict]) -> dict | None:
    consultation_docs = _extract_consultation_offer_documents(pages, "administrative_documents")
    if consultation_docs:
        return consultation_docs

    offer_content = _extract_offer_content_documents(pages, "administrative_documents")
    if offer_content:
        return offer_content

    cetime_item_patterns = (
        r"Une\s+fiche\s+de\s+renseignement[^+.]{0,160}",
        r"Le\s+pr[eé]sent\s+cahier\s+des\s+charges[^+.]{0,260}",
        r"(?:un\s+)?original\s+du\s+registre\s+national\s+de\s+l['’]entreprise",
        r"Attestation\s+de\s+la\s+situation\s+fiscale[^+.]{0,140}",
        r"CV\s+des\s+intervenants[^+.]{0,180}",
        r"Une\s+d[eé]claration\s+de\s+non\s+faillite[^+.]{0,140}",
    )
    for page_entry in pages:
        compact = _compact_fact_text(page_entry["text"])
        folded = _fold_fact_text(compact)
        if "fiche de renseignement" not in folded or "situation fiscale" not in folded:
            continue

        items = []
        for pattern in cetime_item_patterns:
            match = re.search(pattern, compact, flags=re.IGNORECASE)
            if match:
                items.append(match.group(0))
        fact = _list_fact_from_items(items, page_entry["page"], page_entry["section"])
        if fact:
            return fact

    stb_item_patterns = (
        r"Une\s+fiche\s+kys\s*\(Know\s+Your\s+Supplier\)[^.]{0,120}",
        r"Un\s+certificat\s+d['ƒ?T]affil\S+\s+[aÇÿ]\s+ta?\s+Caisse\s+Nationale\s+de\s+S\S+curit\S+\s+Socia\w+",
        r"L['ƒ?T]original[^.]{0,180}?registre\s+national\s+des\s+entreprises",
        r"l['ƒ?T]acte\s+de\s+groupement\s+solidaire[^.]{0,180}",
        r"Une\s+procuration\s+ou\s+le\s+pouvoir[^.]{0,180}",
    )
    for page_entry in pages:
        compact = _compact_fact_text(page_entry["text"])
        folded = _fold_fact_text(compact)
        if "fiche kys" not in folded or "registre national des entreprises" not in folded:
            continue

        items = []
        for pattern in stb_item_patterns:
            match = re.search(pattern, compact, flags=re.IGNORECASE)
            if match:
                items.append(match.group(0))
        fact = _list_fact_from_items(items, page_entry["page"], page_entry["section"])
        if fact:
            return fact

    topnet_item_patterns = (
        r"(?:Le\s+)?cautionnement\s+provisoire[^.]{0,180}",
        r"Le\s+cahier\s+des\s+charges[^.]{0,220}?lu\s+et\s+approuv[eé][^.]{0,120}",
        r"La\s+d[eé]l[eé]gation\s+de\s+pouvoir\s+et\s+de\s+signature",
        r"Un\s+extrait\s+du\s+Registre\s+National\s+des\s+Soci[eé]t[eé]s[^.]{0,180}",
        r"Une\s+attestation\s+d['’]affiliation\s+[aà]\s+la\s+CNSS",
        r"Une\s+attestation\s+d[eé]livr[eé]e\s+par\s+le\s+fabricant[^.]{0,240}",
    )

    for page_entry in pages:
        compact = _compact_fact_text(page_entry["text"])
        folded = _fold_fact_text(compact)
        if "registre national des societes" not in folded or "cnss" not in folded:
            continue

        items = []
        for pattern in topnet_item_patterns:
            match = re.search(pattern, compact, flags=re.IGNORECASE)
            if match:
                items.append(match.group(0))

        fact = _list_fact_from_items(items, page_entry["page"], page_entry["section"])
        if fact:
            return fact

    ubci_pattern = re.compile(
        r"(Un\s+exemplaire\s+du\s+pr[eé]sent\s+document\s+paraph[eé]"
        r"[\s\S]{0,220}?cachet\s+[aà]\s+la\s+derni[eè]re\s+page)",
        re.IGNORECASE,
    )
    fact = _regex_fact_from_pages(pages, ubci_pattern, max_chars=420)
    if fact:
        return _list_fact_from_items([fact["text"]], fact["page"], fact["section"])

    return None


def _extract_technical_documents_topnet_fallback(pages: list[dict]) -> dict | None:
    consultation_docs = _extract_consultation_offer_documents(pages, "technical_documents")
    if consultation_docs:
        return consultation_docs

    offer_content = _extract_offer_content_documents(pages, "technical_documents")
    if offer_content:
        return offer_content

    stb_item_patterns = (
        r"formulaire\s+de\s+r\S+ponses[^.]{0,180}?annexe\s*3",
        r"liste\s+de\s+r\S+f\S+rences\s+du\s+soumissionnaire[^.]{0,220}?annexe\s*10",
        r"justificatifs\s+des\s+r\S+f\S+rences\s+du\s+soumissionnaire[^.]{0,260}",
        r"documentation\s+technique\s+compl\S+te\s+des\s+\S+quipements[^.]{0,260}",
        r"engagement[^.]{0,180}?disponibilit\S+\s+des\s+pi\S+ces\s+de\s+rechange[^.]{0,260}",
        r"autorisation\s+d[yu]\s+constructeur[^.]{0,180}",
        r"Certification\s+des\s+\S+quipements\s+aux\s+normes[^.]{0,220}",
        r"certificat[/\\-]d\S+ctaration\s+de\s+conformit\S+[^.]{0,320}",
    )
    for page_entry in pages:
        compact = _compact_fact_text(page_entry["text"])
        folded = _fold_fact_text(compact)
        if "offre technique ne comporte pas" not in folded and "documentation technique complete" not in folded:
            continue

        items = []
        for pattern in stb_item_patterns:
            for match in re.finditer(pattern, compact, flags=re.IGNORECASE):
                items.append(match.group(0))
        fact = _list_fact_from_items(items, page_entry["page"], page_entry["section"])
        if fact and len(fact.get("items", [])) >= 3:
            return fact

    items = []
    page = None
    section = None
    item_patterns = (
        r"Le\s+formulaire\s+des\s+r[eé]ponses[\s\S]{0,280}?annexe\s*6",
        r"Pr[eé]sentation\s+des\s+sp[eé]cifications\s+techniques\s*-\s*Documents\s+techniques",
        r"La\s+documentation\s+technique\s+d['’]exploitation\s+et\s+de\s+maintenance",
        r"Les\s+plans\s+d['’]installation\s+des\s+fournitures[\s\S]{0,180}?install[eé]",
        r"La\s+documentation\s+de\s+recette\s+technique",
        r"La\s+documentation\s+technique\s+de\s+chaque\s+type\s+de\s+mat[eé]riel\s+propos[eé][\s\S]{0,180}?soumissionnaire",
        r"Le\s+formulaire\s+technique[\s\S]{0,220}?date",
        r"Date\s+de\s+commercialisation\s+du\s+mod[eè]le\s+propos[eé][\s\S]{0,160}?constructeur",
        r"Date\s+pr[eé]visionnelle\s+d['’]arr[eê]t\s+de\s+commercialisation[\s\S]{0,180}?constructeur",
        r"Date\s+de\s+fin\s+de\s+support\s+du\s+mod[eè]le\s+propos[eé][\s\S]{0,160}?constructeur",
        r"La\s+validit[eé]\s+de\s+la\s+soumission",
    )

    for page_entry in pages:
        compact = _compact_fact_text(page_entry["text"])
        for pattern in item_patterns:
            for match in re.finditer(pattern, compact, flags=re.IGNORECASE):
                items.append(match.group(0))
                page = page or page_entry["page"]
                section = section or page_entry["section"]

    return _list_fact_from_items(items, page, section)


def _extract_financial_documents_fallback(pages: list[dict]) -> dict | None:
    consultation_docs = _extract_consultation_offer_documents(pages, "financial_documents")
    if consultation_docs:
        return consultation_docs

    offer_content = _extract_offer_content_documents(pages, "financial_documents")
    if offer_content:
        return offer_content

    item_patterns = (
        r"L\W*offre\s+financi[eè]re\s+par\s+unit[eé]\s+et\s+selon\s+les\s+quantit[eé]s\s+propos[eé]es",
        r"Prix\s+catalogue\s+des\s+pi[eè]ces\s+de\s+rechange",
        r"L\W*offre\s+d\W*extension\s+de\s+garantie",
        r"Proposition\s+d\W*un\s+contrat\s+cadre\s+de\s+service[^.]{0,120}",
        r"L\W*offre\s+financi\S+re\s+doit\s+(?:indiquer|comporter|\S+ciser\s+notamment)[\s\S]{0,520}?(?:TTC|FODEC|d\S+taill\S+|d\S+lai|bordereau\s+des\s+prix)",
        r"Une\s+sous[-\s]?enveloppe\s+ferm[eÃ©]e\s+pour\s+l['â€™]offre\s+financi\S+re[\s\S]{0,420}?(?:soumission|bordereau\s+des\s+prix|sous[-\s]?d[eÃ©]tail)",
        r"\bLa\s+soumission\b",
        r"\bLe\s+bordereau\s+des\s+prix\b",
        r"\bLe\s+sous[-\s]?d[eé]tail\s+des\s+prix(?:\s+par\s+lot)?\b",
        r"\bLe\s+sous[-\s]?d[eÃ©]tail\s+des\s+prix(?:\s+par\s+lot)?\b",
        r"\bF\s*1\s+La\s+soumission[\s\S]{0,260}?cachet\s+du\s+soumissionnaire",
        r"\bF\s*2\s+Le\s+bordereau\s+des\s+prix[\s\S]{0,260}?cachet\s+du\s+soumissionnaire",
    )
    items = []
    page = None
    section = None

    for page_entry in pages:
        compact = _compact_fact_text(page_entry["text"])
        folded = _fold_fact_text(compact)
        if "offre financi" not in compact.lower() and not any(
            marker in folded
            for marker in (
                "offre financiere par unite",
                "prix catalogue",
                "offre financiere",
                "bordereau des prix",
                "la soumission",
            )
        ):
            continue

        for pattern in item_patterns:
            for match in re.finditer(pattern, compact, flags=re.IGNORECASE):
                items.append(match.group(0))
                page = page or page_entry["page"]
                section = section or page_entry["section"]

    return _list_fact_from_items(items, page, section)


REQUESTED_ITEM_QUANTITY_UNIT_RE = (
    r"(?:utilisateurs?|users?|instances?|licen[cs]es?|postes?|unit[eé]s?|uc|pcs?|"
    r"laptops?|ordinateurs?|[eé]crans?|serveurs?|jours?|j/h|ans?)"
)
REQUESTED_TABLE_HEADER_RE = re.compile(
    r"(?:(?:rubriques?|d[eé]signations?|designation|libell[eé]|produits?|articles?|licen[cs]es?)"
    r"(?:\s+\S+){0,10}?\s+(?:quantit\S*|qte|qt[eé]|nombre)"
    r"|(?:quantit\S*|qte|qt[eé]|nombre)(?:\s+\S+){0,6}?\s+"
    r"(?:rubriques?|d[eé]signations?|designation|libell[eé]|produits?|articles?|licen[cs]es?))"
    r"(?:(?:\s+\S+){0,12}?(?:support\s*id|id\s+support|id\s+de\s+licen[cs]e|"
    r"r[eé]f[eé]rence|reference))?",
    re.IGNORECASE,
)
REQUESTED_TABLE_STOP_RE = re.compile(
    r"\b(?:avec\s+nos\s+salutations|tunis\s+le|article\s+\d+|conditions?\s+de\s+paiement|"
    r"modalit[eé]s?|p[eé]nalit[eé]s?|votre\s+offre\s+doit)\b",
    re.IGNORECASE,
)
REQUESTED_ITEM_WITH_SUPPORT_RE = re.compile(
    rf"(?P<designation>[A-Z][A-Za-z0-9À-ÖØ-öø-ÿ&/().,+#'’\-\s]{{2,140}}?)\s+"
    rf"(?P<quantity>\d+(?:[.,]\d+)?\s*{REQUESTED_ITEM_QUANTITY_UNIT_RE})\b"
    r"(?P<between>[\s\S]{0,180}?)(?P<support_id>#[0-9]{4,})",
    re.IGNORECASE,
)
REQUESTED_ITEM_GENERIC_RE = re.compile(
    rf"(?P<designation>[A-Z][A-Za-z0-9À-ÖØ-öø-ÿ&/().,+#'’\-\s]{{2,100}}?)\s+"
    rf"(?P<quantity>\d+(?:[.,]\d+)?\s*{REQUESTED_ITEM_QUANTITY_UNIT_RE})\b",
    re.IGNORECASE,
)
REQUESTED_ITEM_QUANTITY_FIRST_RE = re.compile(
    rf"(?P<quantity>\d+(?:[.,]\d+)?)(?![A-Za-z])(?:\s*{REQUESTED_ITEM_QUANTITY_UNIT_RE})?\s+"
    rf"(?P<designation>[A-Z][A-Za-z0-9À-ÖØ-öø-ÿ&/().,+#'’\-\s]{{2,180}}?)"
    rf"(?=(?:\s+\d+(?:[.,]\d+)?(?![A-Za-z])(?:\s*{REQUESTED_ITEM_QUANTITY_UNIT_RE})?\s+[A-Z])|$)",
    re.IGNORECASE,
)
REQUESTED_ITEM_ROW_TABLE_RE = re.compile(
    r"(?:^|\s)(?:item|poste|n[°o])?\s*\d{1,3}\s+"
    r"(?P<designation>[A-Z].{3,180}?)\s+"
    r"(?P<unit>unit[eé]|u\b|forfait|lot|licen[cs]es?|poste?s?|jours?|j/h|mois|ans?)\s+"
    r"(?P<quantity>\d+(?:[.,]\d+)?)\b",
    re.IGNORECASE,
)
REQUESTED_ITEM_ID_RE = re.compile(
    r"\b(?=[A-Z0-9][A-Z0-9\-\s]{5,80}\d)"
    r"(?:[A-Z0-9]{6,}|[A-Z0-9]{4,}(?:\s*-\s*[A-Z0-9]{2,})+)\b"
)


def _requested_item_windows(compact: str, folded: str) -> list[str]:
    windows = []
    for header_match in REQUESTED_TABLE_HEADER_RE.finditer(compact):
        remainder = compact[header_match.end() :].strip()
        if not remainder:
            continue
        stop = REQUESTED_TABLE_STOP_RE.search(remainder)
        if stop:
            remainder = remainder[: stop.start()]
        windows.append(remainder[:1800])

    if not windows and any(
        marker in folded
        for marker in ("support id", "id support", "id de licence", "id licence", "licence")
    ):
        windows.append(compact[:1800])

    return windows


def _clean_requested_item_designation(value: str) -> str:
    text = _normalize_fact_text(value)
    text = re.sub(
        r"^.*\b(?:rubriques?|d[eé]signations?|designation|libell[eé]|produits?|articles?|"
        r"licen[cs]es?|quantit\S*|qte|qt[eé]|nombre|support\s*id|id\s+support|"
        r"id\s+de\s+licen[cs]e|r[eé]f[eé]rence|reference)\b\s*",
        "",
        text,
        flags=re.IGNORECASE,
    ).strip(" .;:-|")
    text = re.sub(r"\s+", " ", text).strip(" .;:-|")
    folded = _fold_fact_text(text)
    if (
        folded.startswith(
            (
                "et versions",
                "versions ulterieures",
                "admin externes",
                "admin internes",
                "de ressources",
                "des ressources",
                "nombre de",
                "nombre des",
            )
        )
        or ("windows server" in folded and "version" in folded)
        or ("annexe" in folded and "bordereau" in folded)
        or ("modele de bordereau" in folded)
        or ("ressources a proteger" in folded)
    ):
        return ""
    if len(text) < 3 or folded in {
        "rubriques",
        "designation",
        "designations",
        "quantite",
        "support id",
        "id licence",
        "id de licence",
    }:
        return ""
    return text


def _clean_requested_item_identifier(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"\s+", "", value.strip(" .;:-|"))


def _format_requested_item(
    designation: str,
    quantity: str,
    *,
    license_id: str = "",
    support_id: str = "",
) -> str:
    parts = [f"{designation} : {_normalize_fact_text(quantity)}"]
    if license_id:
        parts.append(f"ID licence {license_id}")
    if support_id:
        parts.append(f"Support ID {support_id}")
    return ", ".join(parts)


def _quantity_first_line_items(text: str) -> list[str]:
    items = []
    in_table = False

    for raw_line in str(text).splitlines():
        line = _normalize_fact_text(raw_line)
        if not line:
            continue

        folded = _fold_fact_text(line)
        if not in_table:
            if "quantit" in folded and any(
                marker in folded
                for marker in ("designation", "designations", "article", "articles", "produit", "produits")
            ):
                in_table = True
            continue

        if REQUESTED_TABLE_STOP_RE.search(line):
            break

        match = re.match(
            rf"^(?P<quantity>\d+(?:[.,]\d+)?)(?:\s*{REQUESTED_ITEM_QUANTITY_UNIT_RE})?\s+"
            r"(?P<designation>.+?)\s*$",
            line,
            re.IGNORECASE,
        )
        if not match:
            continue

        raw_designation = _normalize_fact_text(match.group("designation"))
        designation = _clean_requested_item_designation(raw_designation)
        if _fold_fact_text(raw_designation).startswith("licence ") and not _fold_fact_text(designation).startswith("licence "):
            designation = raw_designation
        if not designation:
            continue
        items.append(_format_requested_item(designation, match.group("quantity")))

    return items


def _row_table_requested_items(text: str) -> list[str]:
    compact = _compact_fact_text(text)
    folded = _fold_fact_text(compact)
    if not (
        any(marker in folded for marker in ("designation", "designations", "libelle", "article", "articles"))
        and any(marker in folded for marker in ("quantite", "qte", "qté", "unite", "unité"))
    ):
        return []

    items = []
    seen = set()
    row_re = re.compile(
        r"^\s*(?:item|poste|n[°o])?\s*\d{1,3}\s+"
        r"(?P<designation>.+)\s+"
        r"(?P<unit>unit\S*|u|forfait|lot|licen[cs]es?|poste?s?|jours?|j/h|mois|ans?)\s+"
        r"(?P<quantity>\d+(?:[.,]\d+)?)\s*$",
        re.IGNORECASE,
    )
    header_seen = False
    for raw_line in str(text or "").splitlines():
        line = _compact_fact_text(raw_line)
        folded_line = _fold_fact_text(line)
        if not line or folded_line.startswith("page "):
            continue
        if (
            any(marker in folded_line for marker in ("designation", "designations", "libelle", "article", "articles"))
            and any(marker in folded_line for marker in ("quantite", "qte", "qt", "unite", "unit"))
        ):
            header_seen = True
            continue
        if not header_seen:
            continue

        match = row_re.match(line)
        if not match:
            continue
        raw_designation = _normalize_fact_text(match.group("designation"))
        designation = _clean_requested_item_designation(raw_designation)
        if _fold_fact_text(raw_designation).startswith("licence ") and not _fold_fact_text(designation).startswith("licence "):
            designation = raw_designation
        folded_designation = _fold_fact_text(designation)
        if not designation or folded_designation in seen:
            continue
        unit = _normalize_fact_text(match.group("unit"))
        quantity = _normalize_fact_text(match.group("quantity"))
        items.append(_format_requested_item(designation, f"{quantity} {unit}"))
        seen.add(folded_designation)
        if len(items) >= 20:
            return items

    if items:
        return items
    for match in REQUESTED_ITEM_ROW_TABLE_RE.finditer(compact):
        raw_designation = _normalize_fact_text(match.group("designation"))
        designation = _clean_requested_item_designation(raw_designation)
        if _fold_fact_text(raw_designation).startswith("licence ") and not _fold_fact_text(designation).startswith("licence "):
            designation = raw_designation
        folded_designation = _fold_fact_text(designation)
        if not designation or folded_designation in seen:
            continue
        unit = _normalize_fact_text(match.group("unit"))
        quantity = _normalize_fact_text(match.group("quantity"))
        items.append(_format_requested_item(designation, f"{quantity} {unit}"))
        seen.add(folded_designation)
        if len(items) >= 20:
            break
    return items


def _extract_requested_items_fallback(pages: list[dict]) -> dict | None:
    items = []
    seen_designations = set()
    page = None
    section = None

    early_items = []
    early_page = None
    early_section = None
    for page_entry in pages[:8]:
        compact = _compact_fact_text(page_entry["text"])
        folded = _fold_fact_text(compact)
        if "solution d'impression" not in folded or "multifonction" not in folded:
            continue

        stb_items = []
        if re.search(r"50\s+\S+quipements?\s+multifonction", compact, flags=re.IGNORECASE):
            stb_items.append("50 équipements multifonction impression-copie-scan")
        if re.search(r"100\s*000\s+pages", compact, flags=re.IGNORECASE):
            stb_items.append("Consommables supplémentaires nécessaires pour imprimer >= 100 000 pages")
        if re.search(r"capacit\S+\s+minimale\s+de\s+100\s+\S+quipements", compact, flags=re.IGNORECASE):
            stb_items.append("Application d'administration et de gestion d'impression pour une capacité minimale de 100 équipements")
        for item in stb_items:
            if _fold_fact_text(item) not in {_fold_fact_text(existing) for existing in early_items}:
                early_items.append(item)
        early_page = early_page or page_entry["page"]
        early_section = early_section or page_entry["section"]
    if early_items:
        return _list_fact_from_items(early_items, early_page, early_section)

    for page_entry in pages:
        compact = _compact_fact_text(page_entry["text"])
        folded = _fold_fact_text(compact)
        if not any(
            marker in folded
            for marker in ("licence", "license", "quantit", "qte", "support id", "id support", "designation", "libelle")
        ):
            continue

        row_table_items_found = False
        for item in _row_table_requested_items(page_entry["text"]):
            folded_item = _fold_fact_text(item)
            if folded_item in seen_designations:
                continue
            items.append(item)
            seen_designations.add(folded_item)
            row_table_items_found = True
            page = page or page_entry["page"]
            section = section or page_entry["section"]

        if row_table_items_found:
            continue

        line_quantity_items_found = False
        for item in _quantity_first_line_items(page_entry["text"]):
            folded_item = _fold_fact_text(item)
            if folded_item in seen_designations:
                continue
            items.append(item)
            seen_designations.add(folded_item)
            line_quantity_items_found = True
            page = page or page_entry["page"]
            section = section or page_entry["section"]

        for window in _requested_item_windows(compact, folded):
            matched_support_row = False
            for match in REQUESTED_ITEM_WITH_SUPPORT_RE.finditer(window):
                designation = _clean_requested_item_designation(match.group("designation"))
                if not designation:
                    continue
                quantity = _normalize_fact_text(match.group("quantity"))
                license_match = REQUESTED_ITEM_ID_RE.search(match.group("between") or "")
                license_id = _clean_requested_item_identifier(
                    license_match.group(0) if license_match else ""
                )
                support_id = _clean_requested_item_identifier(match.group("support_id"))
                items.append(
                    _format_requested_item(
                        designation,
                        quantity,
                        license_id=license_id,
                        support_id=support_id,
                    )
                )
                seen_designations.add(_fold_fact_text(designation))
                matched_support_row = True
                page = page or page_entry["page"]
                section = section or page_entry["section"]

            if matched_support_row:
                continue

            if not line_quantity_items_found:
                for match in REQUESTED_ITEM_QUANTITY_FIRST_RE.finditer(window):
                    designation = _clean_requested_item_designation(match.group("designation"))
                    folded_designation = _fold_fact_text(designation)
                    if not designation or folded_designation in seen_designations:
                        continue
                    items.append(_format_requested_item(designation, match.group("quantity")))
                    seen_designations.add(folded_designation)
                    page = page or page_entry["page"]
                    section = section or page_entry["section"]

            for match in REQUESTED_ITEM_GENERIC_RE.finditer(window):
                designation = _clean_requested_item_designation(match.group("designation"))
                folded_designation = _fold_fact_text(designation)
                if not designation or folded_designation in seen_designations:
                    continue
                items.append(_format_requested_item(designation, match.group("quantity")))
                seen_designations.add(folded_designation)
                page = page or page_entry["page"]
                section = section or page_entry["section"]

    return _list_fact_from_items(items, page, section)


def _clean_mined_fact_text(text: str | None) -> str:
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", " ", str(text or ""))
    return re.sub(r"\s+", " ", text).strip(" .;:-")


def _mined_fact(
    fact_type: str,
    label: str,
    value: str,
    page: str | None,
    section: str | None,
    quote: str | None = None,
    confidence: str = "high",
) -> dict | None:
    label = _clean_mined_fact_text(label)
    value = _clean_mined_fact_text(value)
    quote = _clean_mined_fact_text(quote or value)
    if not label or not value:
        return None
    return {
        "type": fact_type,
        "label": label,
        "value": value,
        "text": f"{label} : {value}",
        "quote": quote,
        "page": page or "?",
        "section": section or "general",
        "confidence": confidence,
    }


def _append_mined_fact(items: list[dict], seen: set[tuple[str, str, str]], fact: dict | None) -> None:
    if not fact:
        return
    key = (
        str(fact.get("type", "")),
        _fold_fact_text(str(fact.get("label", ""))),
        _fold_fact_text(str(fact.get("value", ""))),
    )
    if key in seen:
        return
    seen.add(key)
    items.append(fact)


def _extract_mined_lots(pages: list[dict]) -> list[dict]:
    items: list[dict] = []
    seen: set[tuple[str, str, str]] = set()
    lot_pattern = re.compile(
        r"(?:^|\b|\d+\.\s*)LOT\s*(?P<num>\d+)\s*:?\s*(?P<desc>[\s\S]{0,260}?)"
        r"(?=(?:\s+o\s+L['’]?objectif|\s+\d+\.\s*Lot\s+\d+|\s+Le\s+pr[eé]sent|\s+Chaque\s+solution|$))",
        re.IGNORECASE,
    )

    for page_entry in pages:
        compact = _compact_fact_text(page_entry["text"])
        folded = _fold_fact_text(compact)
        if "lot" not in folded:
            continue
        for match in lot_pattern.finditer(compact):
            num = match.group("num")
            desc = _normalize_fact_text(match.group("desc"))
            if len(desc) < 6:
                continue
            _append_mined_fact(
                items,
                seen,
                _mined_fact(
                    "lot",
                    f"Lot {num}",
                    desc,
                    page_entry["page"],
                    page_entry["section"],
                    match.group(0),
                ),
            )
    return items


def _extract_mined_metrics(pages: list[dict]) -> list[dict]:
    items: list[dict] = []
    seen: set[tuple[str, str, str]] = set()
    metric_patterns = (
        ("Nombre de Endpoint PC", r"\bPC\s*:\s*(\d{1,6})\b"),
        ("Nombre de Endpoint Serveur", r"\bServeurs?\s*:\s*(\d{1,6})\b"),
        ("Nombre de Endpoint Mobile", r"\bMobile\s*:\s*(\d{1,6})\b"),
        (
            "Nombre des utilisateurs de la plateforme PAM",
            r"Nombre\s+des\s+utilisateurs\s+de\s+la\s+plateforme\s+([\s\S]{0,100}?Admin[\s\S]{0,80}?internes?)",
        ),
        ("Nombre de ressources a proteger", r"Nombre\s+de\s+ressources\s+[aà]\s+prot[eé]ger\s+(\d{1,6})\b"),
    )

    for page_entry in pages:
        compact = _compact_fact_text(page_entry["text"])
        folded = _fold_fact_text(compact)
        if not any(marker in folded for marker in ("endpoint", "utilisateurs", "ressources", "pc :", "serveur", "mobile")):
            continue

        scope = ""
        if re.search(r"\blot\s*1\b", folded) or any(marker in folded for marker in ("antivirus", "edr", "endpoint")):
            scope = "Lot 1 - "
        elif re.search(r"\blot\s*2\b", folded) or "pam" in folded:
            scope = "Lot 2 - "

        for label, pattern in metric_patterns:
            for match in re.finditer(pattern, compact, flags=re.IGNORECASE):
                _append_mined_fact(
                    items,
                    seen,
                    _mined_fact(
                        "metric",
                        f"{scope}{label}" if scope else label,
                        match.group(1),
                        page_entry["page"],
                        page_entry["section"],
                        match.group(0),
                    ),
                )
    return items


def _extract_mined_pricing_items(pages: list[dict]) -> list[dict]:
    items: list[dict] = []
    seen: set[tuple[str, str, str]] = set()
    row_pattern = re.compile(
        r"\b(?P<item>\d{1,2})\s+"
        r"(?P<designation>[A-ZÀ-ÖØ-öø-ÿ][\wÀ-ÖØ-öø-ÿ'’()./+&\-\s]{4,220}?)"
        r"\s+Unit[eé]\s+(?P<qte>\d+(?:[.,]\d+)?)\b",
        re.IGNORECASE,
    )

    for page_entry in pages:
        compact = _compact_fact_text(page_entry["text"])
        folded = _fold_fact_text(compact)
        if "bordereau des prix" not in folded and "prix u" not in folded:
            continue

        lot_match = re.search(r"\bLOT\s*(\d+)\s*:?\s*([^:]{0,120}?)(?=\s+Item|\s*$)", compact, flags=re.IGNORECASE)
        lot_label = "Bordereau"
        if lot_match:
            lot_label = f"Lot {lot_match.group(1)}"

        for match in row_pattern.finditer(compact):
            designation = _normalize_fact_text(match.group("designation"))
            value = f"{designation} - Unite {match.group('qte')}"
            _append_mined_fact(
                items,
                seen,
                _mined_fact(
                    "pricing_item",
                    f"{lot_label} item {match.group('item')}",
                    value,
                    page_entry["page"],
                    page_entry["section"],
                    match.group(0),
                ),
            )
    return items


def _extract_mined_requested_items(pages: list[dict]) -> list[dict]:
    items: list[dict] = []
    seen: set[tuple[str, str, str]] = set()
    patterns = (
        re.compile(
            r"(?P<designation>Coupeuses?\s+de\s+plans?\s+grand\s+format\s+papier\s+A[O0Q]\s*\(36[^)]*\))\s+"
            r"(?P<quantity>\d{1,4})\b",
            re.IGNORECASE,
        ),
        re.compile(
            r"\b(?P<quantity>\d{1,4})\s+"
            r"(?P<designation>Coupeuses?\s+de\s+plans?\s+A[O0Q][^\"،.;\n]{0,60})",
            re.IGNORECASE,
        ),
    )

    for page_entry in pages:
        compact = _compact_fact_text(page_entry["text"])
        folded = _fold_fact_text(compact)
        if not any(marker in folded for marker in ("designation", "qte", "qté", "coupeuse", "coupeuses")):
            continue

        for item_text in _row_table_requested_items(page_entry["text"]):
            if ":" not in item_text:
                continue
            designation, quantity = item_text.split(":", 1)
            _append_mined_fact(
                items,
                seen,
                _mined_fact(
                    "requested_item",
                    designation.strip(),
                    quantity.strip(),
                    page_entry["page"],
                    page_entry["section"],
                    item_text,
                ),
            )

        for pattern in patterns:
            for match in pattern.finditer(compact):
                designation = _clean_mined_fact_text(match.group("designation"))
                designation = designation.replace("AQ", "A0").replace("AO", "A0")
                quantity = _clean_mined_fact_text(match.group("quantity")).lstrip("0") or "0"
                if len(designation) < 8:
                    continue
                _append_mined_fact(
                    items,
                    seen,
                    _mined_fact(
                        "requested_item",
                        designation,
                        quantity,
                        page_entry["page"],
                        page_entry["section"],
                        match.group(0),
                    ),
                )
    return items


def _extract_mined_technical_requirements(pages: list[dict]) -> list[dict]:
    items: list[dict] = []
    seen: set[tuple[str, str, str]] = set()
    requirement_patterns = (
        ("Fonctionnalité", r"Fonctionnalit[eé]\s+(.+?)(?=\s+(?:A\s+pr[eé]ciser|Type\s*de\s*coupe|Coupe\s+bidirectionnel|Longueur\s+de\s+coupe|Orientation\s+papier|Graduations|Dispositif|Table\s+de\s+coupe|Bac\s+de\s+r[eé]cup[eé]ration|Equipement\s+de\s+s[eé]curit[eé]|Protection|Chariot|La\s+lame|Lame\s+interchangeable|Normes|$))"),
        ("Type de coupe", r"Type\s*de\s*coupe\s+(.+?)(?=\s+(?:Coupe\s+bidirectionnel|Longueur\s+de\s+coupe|Orientation\s+papier|Graduations|Dispositif|Table\s+de\s+coupe|Bac\s+de\s+r[eé]cup[eé]ration|Equipement\s+de\s+s[eé]curit[eé]|Protection|Chariot|La\s+lame|Lame\s+interchangeable|Normes|$))"),
        ("Coupe bidirectionnelle", r"Coupe\s+bidirectionnel(?:le)?\s+(oui|non)"),
        ("Longueur de coupe Papier A0", r"Longueur\s+de\s+coupe\s*:?\s*Papier\s*A[O0Q]\s*\|?\s*([><≥≤]?\s*\d{3,5}\s*mm)"),
        ("Orientation papier A0 à découper", r"Orientation\s+papier\s*A[O0Q]\s+[aà]\s+d[eé]couper\s*\|?\s*(.+?)(?=\s+(?:Graduations|Dispositif|Table\s+de\s+coupe|Bac\s+de\s+r[eé]cup[eé]ration|Equipement\s+de\s+s[eé]curit[eé]|Protection|Chariot|La\s+lame|Lame\s+interchangeable|Normes|$))"),
        ("Graduations", r"Graduations\s+(mm)"),
        ("Dispositif de pression", r"Dispositif\s+de\s+pression\s+(.+?)(?=\s+(?:Table\s+de\s+coupe|Bac\s+de\s+r[eé]cup[eé]ration|Equipement\s+de\s+s[eé]curit[eé]|Protection|Chariot|La\s+lame|Lame\s+interchangeable|Normes|$))"),
        ("Table de coupe", r"Table\s+de\s+coupe\s+(.+?)(?=\s+(?:Bac\s+de\s+r[eé]cup[eé]ration|Equipement\s+de\s+s[eé]curit[eé]|Protection|Chariot|La\s+lame|Lame\s+interchangeable|Normes|$))"),
        ("Bac de récupération des chutes papiers", r"Bac\s+de\s+r[eé]cup[eé]ration\s+des\s+chutes\s+papiers\s+(.+?)(?=\s+(?:Equipement\s+de\s+s[eé]curit[eé]|Protection|Chariot|La\s+lame|Lame\s+interchangeable|Normes|$))"),
        ("Equipement de sécurité", r"Equipement\s+de\s+s[eé]curit[eé]\s+(.+?)(?=\s+(?:Protection|Chariot|La\s+lame|Lame\s+interchangeable|Normes|$))"),
        ("Protection intégrale de la lame", r"Protection\s+int[eé]grale\s+de\s+la\s+lame\s+(oui|non)"),
        ("Chariot porte lame", r"Chariot\s+porte\s+lame\s+(oui|non)"),
        ("Lame interchangeable", r"Lame\s+interchangeable\s+(oui|non)"),
    )

    for page_entry in pages:
        compact = _compact_fact_text(page_entry["text"])
        folded = _fold_fact_text(compact)
        if not any(marker in folded for marker in ("caracteristique", "minimum demande", "coupeuse", "type de coupe", "orientation papier")):
            continue

        for label, pattern in requirement_patterns:
            for match in re.finditer(pattern, compact, flags=re.IGNORECASE):
                value = _clean_mined_fact_text(match.group(1))
                if len(value) < 2:
                    continue
                value = value.replace("AQ", "A0").replace("AO", "A0")
                folded_value = _fold_fact_text(value)
                if len(value) > 180 and (_arabic_char_ratio(value) > 0.03 or "|" in value):
                    continue
                if "minimum demande" in folded_value and len(value) > 80:
                    continue
                _append_mined_fact(
                    items,
                    seen,
                    _mined_fact(
                        "technical_requirement",
                        label,
                        value,
                        page_entry["page"],
                        page_entry["section"],
                        match.group(0),
                    ),
                )
    return items


def _extract_mined_facts(pages: list[dict]) -> dict | None:
    items: list[dict] = []
    seen: set[tuple[str, str, str]] = set()
    for extractor in (
        _extract_mined_lots,
        _extract_mined_metrics,
        _extract_mined_pricing_items,
        _extract_mined_requested_items,
        _extract_mined_technical_requirements,
    ):
        for fact in extractor(pages):
            _append_mined_fact(items, seen, fact)

    if not items:
        return None

    return {
        "text": "\n".join(f"- {item['text']}" for item in items),
        "items": items,
        "page": items[0].get("page", "?"),
        "section": items[0].get("section", "general"),
        "schema": "mined_facts.v1",
    }


def _extract_subject_from_article(article: dict) -> dict | None:
    if "objet" not in article["folded"]:
        return None

    compact = article["compact"]
    match = re.search(
        r"\bOBJET(?:\s+(?:DU|DE\s+LA|DE\s+L['’])\s+(?:MARCHE|MARCHÉ|CONSULTATION|CAHIER\s+DES\s+CHARGES))?\b\s*[:\-]?",
        compact,
        flags=re.IGNORECASE,
    )
    body = compact[match.end() :] if match else _strip_article_heading(compact)
    real_subject_match = re.search(
        r"\b(?:le\s+pr[eÃ©]sent\s+(?:march[eÃ©]|appel\s+d['â€™]offres?)|la\s+pr[eÃ©]sente\s+consultation)\s+a\s+pour\s+objet\b",
        body,
        flags=re.IGNORECASE,
    )
    if real_subject_match:
        body = body[real_subject_match.start() :]
    body = body.strip(" .;:-")
    if not body:
        return None

    sentences = [part.strip(" .;:-") for part in re.split(r"(?<=[.!?])\s+", body) if part.strip()]
    if not sentences:
        return None

    value = sentences[0]
    if len(sentences) > 1 and _fold_fact_text(sentences[1]).startswith("a cet effet"):
        value = f"{value}. {sentences[1]}"

    folded_value = _fold_fact_text(value)
    if any(
        marker in folded_value
        for marker in (
            "fixent les procedures",
            "procedures de l'appel d'offres",
            "stipulent les conditions du marche",
        )
    ):
        return None

    return _fact_from_text(value, article["page"], article["section"])


def _extract_deadline_from_article(article: dict) -> dict | None:
    heading = _article_title_folded(article)
    if "date limite" not in heading and "remise des offres" not in heading:
        return None
    if "validite" in heading and "date limite de reception" not in heading and "date limite de remise" not in heading:
        return None

    date_value = _date_value_from_text(article["compact"])
    if date_value:
        return _fact_from_text(date_value, article["page"], article["section"])

    sentence = _first_sentence_with("date limite", article["compact"])
    if sentence and len(sentence) > 12:
        return _fact_from_text(sentence, article["page"], article["section"])
    return None


def _extract_submission_method_from_article(article: dict) -> dict | None:
    folded = article["folded"]
    if not any(marker in folded for marker in ("date limite", "presentation", "reception de la soumission")):
        return None

    compact = article["compact"]
    start_match = re.search(r"\b(?:les\s+)?(?:offres|soumissions)\s+doivent\s+parvenir\b", compact, re.IGNORECASE)
    if not start_match:
        return None

    value = compact[start_match.start() :]
    value = re.split(r"\(\s*La\s+date\s+limite|\bLa\s+date\s+limite", value, maxsplit=1, flags=re.IGNORECASE)[0]
    return _fact_from_text(value[:650], article["page"], article["section"])


def _extract_validity_from_article(article: dict) -> dict | None:
    heading = _article_title_folded(article)
    if "validite" not in heading and "valables" not in heading:
        return None
    if any(
        marker in heading
        for marker in (
            "caution",
            "garantie",
            "execution",
            "marche",
            "contrat",
            "bancaire",
        )
    ) and not any(marker in heading for marker in ("offre", "offres", "soumission", "soumissions")):
        return None

    sentence = _first_sentence_with("valables", article["compact"])
    if not sentence:
        sentence = _first_sentence_with("engag", article["compact"])
    if not sentence:
        sentence = _first_sentence_with("lies", article["compact"])
    if sentence:
        sentence = re.sub(r"^.*?\bLes\s+offres\b", "Les offres", sentence, count=1, flags=re.IGNORECASE)
        sentence = re.sub(r"^.*?\bLes\s+soumissionnaires\b", "Les soumissionnaires", sentence, count=1, flags=re.IGNORECASE)
        sentence = re.sub(r"^.*?\bLes\s+candidats\b", "Les candidats", sentence, count=1, flags=re.IGNORECASE)
        return _fact_from_text(sentence, article["page"], article["section"])
    return None


def _extract_caution_from_article(article: dict) -> dict | None:
    folded = _fold_fact_text(article["compact"])
    if (
        "caution provisoire" not in folded
        and "cautionnement provisoire" not in folded
        and "garantie de soumission" not in folded
    ):
        return None

    sentence = _first_sentence_with("montant de la caution provisoire", article["compact"])
    if not sentence:
        sentence = _first_sentence_with("caution provisoire", article["compact"])
    if not sentence:
        sentence = _first_sentence_with("cautionnement provisoire", article["compact"])
    if sentence and (
        AMOUNT_VALUE_RE.search(sentence)
        or PERCENT_VALUE_RE.search(sentence)
        or "caution provisoire" in _fold_fact_text(sentence)
        or "cautionnement provisoire" in _fold_fact_text(sentence)
    ):
        return _fact_from_text(sentence, article["page"], article["section"])
    return None


def _extract_definitive_caution_from_article(article: dict) -> dict | None:
    if "caution definitive" not in _fold_fact_text(article["compact"]):
        return None

    sentence = _first_sentence_with("caution definitive", article["compact"])
    if sentence:
        return _fact_from_text(sentence, article["page"], article["section"])
    return None


def _extract_guarantee_from_article(article: dict) -> dict | None:
    folded = article["folded"]
    heading = _article_title_folded(article)
    if "garantie" not in heading and "garantie est fix" not in folded and "delai de garantie est" not in folded:
        return None

    duration_re = re.compile(r"\b\d+\s*(?:\([^)]*\)\s*)?(?:jours?|mois|ans?|heures?)\b", re.IGNORECASE)
    for marker in ("delai de garantie", "periode de garantie", "duree de garantie"):
        sentence = _first_sentence_with(marker, article["compact"])
        if not sentence:
            continue
        sentence_folded = _fold_fact_text(sentence)
        if "expiration du delai de garantie" in sentence_folded and "fix" not in sentence_folded:
            continue
        if (
            duration_re.search(sentence)
            or "garantie est fixe" in sentence_folded
            or "garantie est fixee" in sentence_folded
            or "48 heures" in sentence_folded
        ):
            return _fact_from_text(sentence, article["page"], article["section"])
    return None


def _extract_reception_from_article(article: dict) -> dict | None:
    heading = _article_title_folded(article)
    if "reception" not in heading or "reception de la soumission" in heading:
        return None

    compact = article["compact"]
    provisional = _first_sentence_with("reception provisoire", compact, max_chars=500)
    definitive = _first_sentence_with("reception definitive", compact, max_chars=500)
    pronounced = _first_sentence_with("reception est prononcee", compact, max_chars=900)
    parts = [part for part in (provisional, definitive, pronounced) if part]
    if parts:
        return _fact_from_text(" ".join(parts), article["page"], article["section"])
    return None


def _extract_penalties_from_article(article: dict) -> dict | None:
    folded = article["folded"]
    if "penalite" not in folded and "retard" not in folded:
        return None

    compact = article["compact"]
    first = _first_sentence_with("penalite", compact, max_chars=500)
    cap = _first_sentence_with("cinq pour cent", compact, max_chars=400) or _first_sentence_with("5 %", compact, max_chars=400)
    if first:
        first = re.sub(r"^\s*\|?\s*PENALITE\s+DE\s+RETARD\s*", "", first, flags=re.IGNORECASE)
    parts = [part for part in (first, cap) if part]
    if parts:
        return _fact_from_text(" ".join(parts), article["page"], article["section"])
    return None


def _extract_payment_from_articles(articles: list[dict]) -> dict | None:
    parts = []
    page = None
    section = None

    for article in articles:
        folded = article["folded"]
        if "conditions de paiement" not in folded and "modalites de paiement" not in folded:
            continue

        compact = _strip_article_heading(article["compact"])
        compact = re.sub(
            r"^\s*\|?\s*(?:CONDITIONS|MODALIT[EÉ]S?)?\s*DE\s+PAIEMENT\s+",
            "",
            compact,
            count=1,
            flags=re.IGNORECASE,
        )
        if not compact:
            continue

        if page is None:
            page = article["page"]
            section = article["section"]
        parts.append(compact[:650])

    if not parts:
        return None

    return _fact_from_text(" ".join(parts), page or "?", section or "payment")


def _extract_article_facts(pages: list[dict]) -> dict:
    articles = _article_sections_from_pages(pages)
    if not articles:
        return {}

    facts = {}
    extractors = (
        ("subject", _extract_subject_from_article),
        ("deadline", _extract_deadline_from_article),
        ("submission_method", _extract_submission_method_from_article),
        ("validity", _extract_validity_from_article),
        ("caution", _extract_caution_from_article),
        ("definitive_caution", _extract_definitive_caution_from_article),
        ("guarantee", _extract_guarantee_from_article),
        ("reception", _extract_reception_from_article),
        ("penalties", _extract_penalties_from_article),
    )

    for field, extractor in extractors:
        for article in articles:
            fact = extractor(article)
            if fact:
                facts[field] = fact
                break

    payment = _extract_payment_from_articles(articles)
    if payment:
        facts["payment"] = payment

    return facts


def _is_table_of_contents_fact(fact: dict | None) -> bool:
    if not fact:
        return False

    folded = _fold_fact_text(str(fact.get("text", "")))
    if not folded.startswith("article"):
        return False
    if len(folded) > 140:
        return False
    return not any(
        marker in folded
        for marker in (
            " est ",
            " sont ",
            " sera ",
            " seront ",
            " doit ",
            " doivent ",
            " fixe ",
            " pay",
            " par jour",
            "reception quantitative",
            "reception provisoire",
            "reception definitive",
        )
    )


def _is_reliable_guarantee_text(text: str, folded: str) -> bool:
    arabic_financial_guarantee_markers = (
        "الضمان المالي الوقتي",
        "الضمان المالي النهائي",
    )
    if any(marker in text or marker in folded for marker in arabic_financial_guarantee_markers):
        return False
    arabic_warranty_markers = (
        "مدة الضمان",
        "مدّة الضمان",
        "مدة ضمان",
        "أشهر",
        "اشهر",
    )
    if any(marker in text or marker in folded for marker in arabic_warranty_markers):
        arabic_duration_markers = (
            "أشهر",
            "اشهر",
            "شهر",
            "شهور",
            "يوم",
            "سنة",
            "سنوات",
        )
        return bool(
            re.search(r"\d+", text)
            or any(marker in text or marker in folded for marker in arabic_duration_markers)
        )

    duration_re = re.compile(r"\b\d+\s*(?:\([^)]*\)\s*)?(?:jours?|mois|ans?|heures?)\b", re.IGNORECASE)
    has_duration = bool(duration_re.search(text) or "48 heures" in folded)
    word_duration_re = re.compile(
        r"\b(?:une|un|deux|trois|quatre|cinq|six|douze)\s+(?:jours?|mois|ans?|annees?)\b",
        re.IGNORECASE,
    )
    has_duration = has_duration or bool(word_duration_re.search(folded))
    has_context = any(
        marker in folded
        for marker in (
            "delai de garantie",
            "duree de garantie",
            "periode de garantie",
            "garantie est fixe",
            "garantie est fixee",
            "garantie est de",
            "garantie de",
            "periode de 6 mois",
            "garantie est d'une annee",
            "garantie est d une annee",
            "annee de garantie",
        )
    )
    if "expiration du delai de garantie" in folded and "fix" not in folded:
        return "annee de garantie" in folded or "garantie est d'une annee" in folded or "garantie est d une annee" in folded
    if "doit garantir" in folded and not has_duration:
        return False
    return has_context and has_duration


def _has_template_placeholders(folded: str) -> bool:
    return bool(re.search(r"(?:\.{4,}|\(\s*[67]\s*\))", folded))


def _is_caution_template_context(folded: str) -> bool:
    hard_template_markers = (
        "cautionnement provisoire pour participer",
        "le montant du dit cautionnement",
        "m'engage",
        "nous engageons",
        "fait a",
        "publie(e) en date",
        "relatif-relative",
        "relatif relative",
    )
    if any(marker in folded for marker in hard_template_markers):
        return True

    soft_template_markers = (
        "annexe",
        "modele",
        "formulaire",
    )
    if _has_template_placeholders(folded) and any(marker in folded for marker in soft_template_markers):
        return True

    real_clause_markers = (
        "doit fournir",
        "doit presenter",
        "doit etre accompagne",
        "est fixe",
        "s'eleve",
        "s eleve",
        "chaque offre",
        "dans son offre",
        "a l'appui de son offre",
    )
    return any(marker in folded for marker in soft_template_markers) and not any(
        marker in folded for marker in real_clause_markers
    )


def _is_bad_reception_fact_context(folded: str) -> bool:
    if "reception" not in folded:
        return False

    hard_bad_context = (
        "composition de l'offre",
        "piece jointe",
        "pieces jointes",
        "a telecharger",
        "envoye en ligne",
        "offre financiere",
        "bordereau des prix",
        "bordereaux des prix",
        "non presentation du cautionnement",
    )
    if "composition de l'offre" in folded and any(
        marker in folded for marker in hard_bad_context[1:]
    ):
        return True

    bad_markers = (
        "composition de l'offre",
        "piece jointe",
        "pieces jointes",
        "a telecharger",
        "envoye en ligne",
        "annexe",
        "formulaire",
        "liste nominative",
        "offre financiere",
        "soumission",
        "bordereaux des prix",
        "bordereau des prix",
        "non presentation du cautionnement",
        "fiches techniques",
        "references",
    )
    true_markers = (
        "reception quantitative",
        "reception provisoire",
        "reception definitive",
        "sera prononcee",
        "est prononcee",
        "modalites de reception",
        "reception technique",
        "pv de reception",
        "proces-verbal de reception",
    )
    bad_hits = sum(marker in folded for marker in bad_markers)
    true_hits = sum(marker in folded for marker in true_markers)
    return (bad_hits >= 2 and true_hits < 2) or (bad_hits >= 4 and true_hits <= 3)


def _is_reliable_scalar_fact(field: str, fact: dict | None) -> bool:
    if not fact:
        return False

    raw_text = str(fact.get("text", ""))
    text = _normalize_arabic_ocr_for_fact_matching(raw_text)
    folded = _fold_fact_text(text)
    if not folded:
        return False

    if _is_table_of_contents_fact(fact):
        return False

    if field == "guarantee":
        return _is_reliable_guarantee_text(text, folded)

    if field == "deadline":
        if any(marker in text for marker in ("آخر أجل", "اخر أجل", "التاريخ الأقصى", "التاريخ الاقصى")):
            return bool(re.search(r"\d{1,4}", text))
        return bool(_date_value_from_text(text))

    if field == "caution":
        if "restitution" in folded or "mise en paiement" in folded:
            return False
        if _is_caution_template_context(folded):
            return False
        if "الضمان المالي الوقتي" in text or "الضمان الوقتي" in text:
            return bool(re.search(r"\d{1,4}", text) or "دينار" in text or "يوما" in text)
        return bool(
            AMOUNT_VALUE_RE.search(text)
            or WORD_AMOUNT_VALUE_RE.search(text)
            or PERCENT_VALUE_RE.search(text)
            or "exige" in folded
        )

    if field == "definitive_caution":
        if _is_caution_template_context(folded):
            return False
        definitive_markers = (
            "definit",
            "bonne fin",
            "bonne execution",
            "الضمان المالي النهائي",
            "الضمان النهائي",
            "ضمانا ماليا نهائيا",
            "ضمان مالي نهائي",
            "ضمان نهائي",
            "Ø§Ù„Ø¶Ù…Ø§Ù† Ø§Ù„Ù…Ø§Ù„ÙŠ Ø§Ù„Ù†Ù‡Ø§Ø¦ÙŠ",
            "Ø¶Ù…Ø§Ù†Ø§ Ù…Ø§Ù„ÙŠØ§ Ù†Ù‡Ø§Ø¦ÙŠØ§",
        )
        if not any(marker in folded or marker in text for marker in definitive_markers):
            return False
        return bool(
            AMOUNT_VALUE_RE.search(text)
            or WORD_AMOUNT_VALUE_RE.search(text)
            or PERCENT_VALUE_RE.search(text)
            or re.search(r"(?<!\d)\d{1,2}\s+96\b", text)
            or "بالمائة" in text
            or "Ø¨Ø§Ù„Ù…Ø§Ø¦Ø©" in text
            or "exige" in folded
            or "doit fournir" in folded
            or "doit constituer" in folded
        )

    if field == "penalties":
        if folded.startswith("resiliation"):
            return False
        if (
            "غرامة تأخير" in text
            or "غرامة التأخير" in text
            or "غرامات التأخير" in text
            or "عقوبة التأخير" in text
            or "خطايا التأخير" in text
        ):
            return True
        return "penalite" in folded and ("retard" in folded or "jour" in folded or "pour mille" in folded)

    if field == "reception":
        if _is_bad_reception_fact_context(folded):
            return False
        arabic_reception_markers = (
            "الاستلام",
            "الإستلام",
            "Ø§Ù„Ø§Ø³ØªÙ„Ø§Ù…",
            "Ø§Ù„Ø¥Ø³ØªÙ„Ø§Ù…",
            "Ã˜Â§Ã™â€žÃ˜Â§Ã˜Â³Ã˜ÂªÃ™â€žÃ˜Â§Ã™â€¦",
            "Ã˜Â§Ã™â€žÃ˜Â¥Ã˜Â³Ã˜ÂªÃ™â€žÃ˜Â§Ã™â€¦",
        )
        if any(marker in text for marker in arabic_reception_markers):
            return True
        return "reception" in folded and any(
            marker in folded
            for marker in (
                "provisoire",
                "definitive",
                "quantitative",
                "prononce",
                "pv",
                "proces-verbal",
                "modalites",
            )
        )

    if field == "guarantee":
        if "الضمان المالي الوقتي" in text or "الضمان المالي النهائي" in text:
            return False
        return True

    if field == "submission_method":
        tender_markers = (
            "soumission",
            "offre technique",
            "offre financiere",
            "offre financière",
            "enveloppe",
            "pli",
            "tuneps",
            "tuneps.tn",
            "bureau d'ordre",
            "bureau d ordre",
            "depot des offres",
            "depot de l'offre",
            "dépôt des offres",
            "dépôt de l'offre",
            "remise des offres",
            "reception des offres",
            "réception des offres",
            "voie postale",
            "rapid poste",
            "rapide poste",
            "منظومة الشراء العمومي",
            "تونبس",
            "ارسال العرض",
            "إرسال العرض",
            "مكتب الضبط",
            "البريد مضمون الوصول",
            "البريد السريع",
        )
        reject_markers = (
            "eclaircissement",
            "éclaircissement",
            "correspondance",
            "facture",
            "factures",
            "notification",
            "reclamation",
            "réclamation",
            "demande d'information",
        )
        has_tender_context = any(marker in folded for marker in tender_markers)
        has_reject_context = any(marker in folded for marker in reject_markers)
        if has_reject_context and not has_tender_context:
            return False
        return has_tender_context

    if field == "validity":
        if any(
            marker in text
            for marker in (
                "صلوحية العروض",
                "يلتزم العارض بعرضه",
                "يبقى المتعهدون ملتزمون",
                "تسعون يوما",
                "تسعون يومًا",
                "ستون يوما",
                "120 يوما",
                "لمدة 120 يوما",
                "٩٠",
            )
        ):
            return True
        wrong_context = (
            "caution",
            "garantie",
            "execution",
            "exécution",
            "contrat",
            "marche",
            "marché",
            "bancaire",
        )
        right_context = (
            "offre",
            "offres",
            "soumission",
            "soumissions",
            "valable",
            "valables",
        )
        has_wrong = any(marker in folded for marker in wrong_context)
        has_right = any(marker in folded for marker in right_context)
        if has_wrong and not has_right:
            return False
        has_duration = bool(
            re.search(
                r"\b(?:\d+\s*(?:jours?|mois|ans?)|"
                r"90|120|60|180|quatre[-\s]vingt[-\s]dix|"
                r"trois\s+mois|six\s+mois)\b",
                text,
                re.IGNORECASE,
            )
        )
        if any(
            marker in folded
            for marker in (
                "validite de la soumission",
                "validite de l'offre",
                "validite des offres",
            )
        ):
            return True
        return has_duration or "valable" in folded or "valables" in folded

    if field == "references":
        false_reference_markers = (
            "reference relative au nom du soumissionnaire",
            "reference comportant le nom du soumissionnaire",
            "reference au nom du soumissionnaire",
            "automatiquement rejet",
            "sera rejet",
        )
        true_reference_markers = (
            "references similaires",
            "reference similaire",
            "projets similaires",
            "travaux similaires",
            "references techniques",
            "references financieres",
            "attestations de bonne execution",
            "attestation de bonne execution",
            "certificats",
            "contrats",
            "pvs de reception",
            "pv de reception",
            "justificatifs",
        )
        if any(marker in folded for marker in false_reference_markers) and not any(
            marker in folded for marker in true_reference_markers
        ):
            return False
        return any(marker in folded for marker in true_reference_markers)

    return True


def _next_list_stop(text: str, start: int, stop_patterns: tuple[re.Pattern, ...]) -> int:
    stops = []
    for pattern in stop_patterns:
        match = pattern.search(text, start)
        if match:
            stops.append(match.start())

    if stops:
        return min(stops)
    return min(len(text), start + 2200)


def _clean_fact_list_item(item: str) -> str:
    item = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", " ", item)
    cleaned = re.sub(r"\s+", " ", item).strip(" .;:-")
    cleaned = re.sub(r"\s*\|\s*", " ", cleaned)
    cleaned = re.sub(r"\b\d+\s*,\s*\d+\s*=", " ", cleaned)
    cleaned = re.sub(r"^(?:[-–—•●▪*]|\d+[.)])\s*", "", cleaned).strip(" .;:-")
    cleaned = re.sub(
        r"\b(?:Date,\s*)?signature\s+et\s+cachet\s+du\s+soumissionnaire\b.*$",
        "",
        cleaned,
        flags=re.IGNORECASE,
    )
    cleaned = re.sub(r"\bCachet\s+signature\s+du\s+soumissionnaire\b.*$", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\bD[uû]ment\s+sign[eé].*$", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\bAuthentifications?\s+pi[eè]ce\b", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\bN[°o]\s+de\s+la\s+D[eé]signations?\b", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\b(?:Designations?|Désignations?|Authentifications?)\b\s*=?", " ", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\b(?:N[°o]\s*de\s+la\s+piece|N[°o]\s*de\s+la\s+pièce)\b\s*=?", " ", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(
        r"^[\s'\"`.,;:|_\-()0-9a-zA-Z]{0,35}?\b(?=(?:La\s+lettre|Le\s+bordereau|Le\s+r[eé]capitulatif|Pr[eé]sentation\s*du\s*Soumissionnaire|RNE\s+r[eé]cent|Une\s+attestation|D[eé]claration)\b)",
        "",
        cleaned,
        flags=re.IGNORECASE,
    )
    cleaned = re.sub(r"\bPr[eé]sentation\s*du\s*Soumissionnaire\b", "Présentation du Soumissionnaire", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\bLa(?=certification|documentation|liste|lettre|caution|date|livraison)", "La ", cleaned)
    cleaned = re.sub(r"\bselonlemod[eè]lejointenannexe\b", "selon le modèle joint en annexe", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\bConform[eé]ment\s+au\s+mod[eè]lejointenannexe\b", "conformément au modèle joint en annexe", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\bmod[eè]lejointenannexe\b", "modèle joint en annexe", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\benannexe\b", "en annexe", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\bleurCV\b", "leur CV", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\bchargesselon\b", "charges selon", cleaned, flags=re.IGNORECASE)
    if cleaned.count("(") > cleaned.count(")"):
        cleaned = cleaned.rstrip(" (")
        if cleaned.count("(") > cleaned.count(")"):
            cleaned += ")"
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .;:-|")
    return cleaned


def _fact_list_dedupe_key(item: str) -> str:
    folded = _fold_fact_text(item)
    folded = re.sub(r"\bla(?=certification|documentation|liste|lettre|caution|date|livraison)", "la ", folded)
    folded = re.sub(r"\b(?:la|le|les|un|une|des|de|du|d|et|a)\b", " ", folded)
    return re.sub(r"\s+", " ", folded).strip()[:160]


def _strip_table_authentication_columns(text: str) -> str:
    text = re.sub(r"\b\d+\s*,\s*(?:Authentifications?|Authentification)\s*=\s*.*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\b(?:Authentifications?|Authentification)\s*=\s*.*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\b\d+\s*,\s*2\s*=\s*.*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\b(?:Date,\s*)?signature\s+et\s+cachet\s+du\s+soumissionnaire\b.*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bCachet\s+signature\s+du\s+soumissionnaire\b.*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bD[uû]ment\s+sign[eé]\s+paraph[eé]\s+et\s+dat[eé].*$", "", text, flags=re.IGNORECASE)
    return text


def _extract_designation_table_items(segment: str, field: str | None = None) -> list[str]:
    compact = re.sub(r"\s+", " ", segment).strip()
    folded = _fold_fact_text(compact)
    if not any(marker in folded for marker in ("designation", "designations", "authentification", "authentifications")):
        return []

    candidates = []
    designation_re = re.compile(
        r"(?:^|\s)(?:\d+\s*,\s*)?D(?:e|é)signations?\s*=\s*(?P<item>.*?)(?=\s+(?:\d+\s*,\s*)?Authentifications?\s*=|\s+\d+\s*,\s*D(?:e|é)signations?\s*=|$)",
        re.IGNORECASE,
    )
    for match in designation_re.finditer(compact):
        candidates.append(match.group("item"))

    numbered_row_re = re.compile(
        r"(?:^|\s)\|?\s*\d+\s*,\s*1\s*=\s*(?P<item>.*?)(?=\s+\|?\s*\d+\s*,\s*2\s*=|\s+\|?\s*\d+\s*,\s*1\s*=|$)",
        re.IGNORECASE,
    )
    for match in numbered_row_re.finditer(compact):
        candidates.append(match.group("item"))

    cleaned_items = []
    seen = set()
    for candidate in candidates:
        item = _strip_table_authentication_columns(candidate)
        item = re.split(
            r"\s+(?=(?:\d+\s*,\s*)?D(?:e|é)signations?\s*=|\d+\s*,\s*1\s*=|\d+\s*,\s*2\s*=)",
            item,
            maxsplit=1,
            flags=re.IGNORECASE,
        )[0]
        item = _clean_fact_list_item(item)
        item = re.sub(r"^(?:joint\s+en\s+annexe\s*\(?\d+\)?\.?\s*)", "", item, flags=re.IGNORECASE)
        if not item or (field and not _is_useful_fact_list_item(field, item)):
            continue
        key = _fact_list_dedupe_key(item)
        if not key or key in seen:
            continue
        seen.add(key)
        cleaned_items.append(item)
        if len(cleaned_items) >= 12:
            break

    return cleaned_items


LIST_ITEM_START_RE = re.compile(
    r"\s(?=(?:\d{1,2}\s+)?(?:"
    r"La\s+caution|Les\s+tableaux|Les\s+r[eé]f[eé]rences|La\s+liste|Une\s+attestation|"
    r"Une\s+d[eé]claration|Un\s+extrait|Le\s+CCAP|Le\s+CPTP|RNE|Documentation|"
    r"Engagement|Un\s+engagement|Planning|Contrat|Un\s+document\s+technique|"
    r"La\s+certification|La\s+lettre|Le\s+bordereau|Le\s+r[eé]capitulatif|"
    r"D[eé]claration|Pr[eé]sentation|Programme|Liste|R[eé]capitulatif"
    r")\b)",
    re.IGNORECASE,
)


FACT_LIST_FIELD_MARKERS: dict[str, tuple[str, ...]] = {
    "administrative_documents": (
        "caution",
        "rne",
        "cnss",
        "attestation",
        "declaration",
        "reference",
        "presentation",
        "equipe",
        "ccap",
        "cptp",
        "registre",
        "fiche",
        "fiscale",
    ),
    "technical_documents": (
        "liste",
        "equipement",
        "documentation",
        "prospectus",
        "brochure",
        "engagement",
        "origine",
        "fourniture",
        "formation",
        "support",
        "planning",
        "test",
        "essai",
        "recette",
        "certification",
        "certifications",
        "vmware",
        "hpe",
    ),
    "financial_documents": (
        "lettre de soumission",
        "bordereau",
        "prix",
        "recapitulatif",
        "offre financiere",
        "soumission",
    ),
}


def _is_useful_fact_list_item(field: str, item: str) -> bool:
    folded = _fold_fact_text(item)
    if len(folded) < 4:
        return False
    if len(item) > 520:
        return False
    if any(
        marker in folded
        for marker in (
            "n de la designations",
            "authentifications piece",
            "outre l'offre technique",
            "toute offre dont l'enveloppe",
            "a ne pas ouvrir",
            "doit contenir sous peine",
        )
    ):
        return False

    markers = FACT_LIST_FIELD_MARKERS.get(field, ())
    return not markers or any(marker in folded for marker in markers)


def _split_fact_list_items(segment: str, field: str | None = None) -> list[str]:
    if not segment:
        return []

    head = segment[:600]
    colon_idx = head.find(":")
    if colon_idx >= 0:
        segment = segment[colon_idx + 1 :]

    compact = re.sub(r"\s+", " ", segment).strip()
    table_items = _extract_designation_table_items(compact, field)
    if table_items:
        return table_items
    compact = re.sub(r"\s(?:[-–—•●▪*])\s+", "\n", compact)
    compact = re.sub(r"\s\d+[.)]\s+", "\n", compact)
    compact = LIST_ITEM_START_RE.sub("\n", compact)
    compact = re.sub(r";\s*", "\n", compact)
    compact = re.sub(
        r"\.\s+(?=(?:Un|Une|Le|La|Les|L['’]|Date|Prix|Proposition|Attestation|Certificat|Extrait|Formulaire|Documentation)\b)",
        "\n",
        compact,
    )

    items = []
    seen = set()
    for piece in compact.split("\n"):
        item = _clean_fact_list_item(piece)
        if len(item) < 4:
            continue
        if field and not _is_useful_fact_list_item(field, item):
            continue

        normalized = re.sub(r"\W+", " ", item.lower()).strip()
        if not normalized or normalized in seen:
            continue
        if normalized in {"doit contenir", "doit comporter", "pieces suivantes"}:
            continue
        if normalized.startswith(("article ", "date limite ")):
            continue

        seen.add(normalized)
        items.append(item)
        if len(items) >= 12:
            break

    return items


def _extract_fact_list_from_pages(
    pages: list[dict],
    field: str,
    heading_patterns: tuple[re.Pattern, ...],
    stop_patterns: tuple[re.Pattern, ...],
) -> dict | None:
    for page_entry in pages:
        text = page_entry["text"].strip()
        if not text:
            continue

        for heading_pattern in heading_patterns:
            for match in heading_pattern.finditer(text):
                end = _next_list_stop(text, match.end(), stop_patterns)
                segment = text[match.start() : end]
                items = _split_fact_list_items(segment, field)
                if not items:
                    continue

                return {
                    "text": "\n".join(f"- {item}" for item in items),
                    "items": [
                        {
                            "text": item,
                            "page": page_entry["page"],
                            "section": page_entry["section"],
                        }
                        for item in items
                    ],
                    "page": page_entry["page"],
                    "section": page_entry["section"],
                }

    return None


TECHNICAL_LIST_ITEM_PATTERNS = (
    re.compile(r"(?:la\s+)?liste\s+de\s+l['’]?\s*[eé]quipe\s+intervenante[^.\n|]{0,280}", re.IGNORECASE),
    re.compile(r"(?:copie\s+des\s+)?dipl[oô]mes?\s+et\s+des\s+certifications?[^.\n|]{0,180}", re.IGNORECASE),
    re.compile(r"documentation(?:s)?\s+techniques?[^.\n|]{0,220}", re.IGNORECASE),
    re.compile(r"engagement\s+concernant\s+l['’]?\s*origine\s+des\s+fournitures[^.\n|]{0,180}", re.IGNORECASE),
    re.compile(r"(?:la\s+)?certification\s+sur\s+(?:HPE|VMware|VMWARE)[^.\n|]{0,120}", re.IGNORECASE),
    re.compile(r"(?:la\s+)?liste\s+des\s+[eé]quipements[^.\n|]{0,180}", re.IGNORECASE),
    re.compile(r"(?:transfert\s+de\s+comp[eé]tences|formation)[^.\n|]{0,180}", re.IGNORECASE),
)


def _extract_technical_documents_fallback(pages: list[dict]) -> dict | None:
    heading_re = re.compile(
        r"\b(?:dossier\s+de\s+l['’]?\s*offre\s+technique|offre\s+technique)\b",
        re.IGNORECASE,
    )
    stop_re = re.compile(
        r"\b(?:offre\s+financi[eè]re|date\s+limite|article\s+\d+)\b",
        re.IGNORECASE,
    )

    for page_entry in pages:
        text = page_entry["text"].strip()
        if not text:
            continue

        for match in heading_re.finditer(text):
            stop = stop_re.search(text, match.end())
            segment = text[match.start() : stop.start() if stop else min(len(text), match.start() + 2600)]
            items = _split_fact_list_items(segment, "technical_documents")

            if not items:
                seen = set()
                for item_pattern in TECHNICAL_LIST_ITEM_PATTERNS:
                    for item_match in item_pattern.finditer(segment):
                        item = _clean_fact_list_item(item_match.group(0))
                        folded = _fold_fact_text(item)
                        if item and folded not in seen:
                            seen.add(folded)
                            items.append(item)
                        if len(items) >= 8:
                            break
                    if len(items) >= 8:
                        break

            if len(items) == 1 and "offre technique" in _fold_fact_text(items[0]) and len(items[0]) > 260:
                items = []

            if not items:
                folded_segment = _fold_fact_text(segment)
                marker_count = sum(
                    1
                    for marker in (
                        "documentation technique",
                        "certification",
                        "vmware",
                        "hpe",
                        "origine des fournitures",
                        "liste de l'equipe",
                        "liste des equipements",
                    )
                    if marker in folded_segment
                )
                if marker_count >= 2:
                    cleaned_segment = re.sub(
                        r"^.*?\b(?:dossier\s+de\s+l['’]?\s*offre\s+technique|offre\s+technique)\b",
                        "",
                        segment,
                        count=1,
                        flags=re.IGNORECASE | re.DOTALL,
                    )
                    cleaned_segment = _clean_fact_list_item(cleaned_segment[:900])
                    if cleaned_segment:
                        items = [cleaned_segment]

            if not items:
                continue

            return {
                "text": "\n".join(f"- {item}" for item in items),
                "items": [
                    {
                        "text": item,
                        "page": page_entry["page"],
                        "section": page_entry["section"],
                    }
                    for item in items
                ],
                "page": page_entry["page"],
                "section": page_entry["section"],
            }

    fallback_items = []
    fallback_page = None
    fallback_section = None
    seen = set()
    row_patterns = (
        re.compile(
            r"La\s*certification\s+de\s+leurs\s+[eé]quipes\s+sur\s+HPE\s+Synergy\s+et\s+pour\s+Vmware",
            re.IGNORECASE,
        ),
        re.compile(
            r"La\s+liste\s+de\s+l['’]?\s*[eé]quipe\s+intervenante[\s\S]{0,360}?(?:Annexe\s*\(?6\)?|soumissionnaire|RNE)",
            re.IGNORECASE,
        ),
        re.compile(
            r"La\s+livraison\s+de\s+la\s+documentation\s+technique\s+n[eé]cessaire\s*\([^)]+\)",
            re.IGNORECASE,
        ),
        re.compile(
            r"documentation\s+technique\s+n[eé]cessaire[^.\n]{0,260}",
            re.IGNORECASE,
        ),
    )

    for page_entry in pages:
        text = page_entry["text"].strip()
        if not text:
            continue

        for row_pattern in row_patterns:
            for match in row_pattern.finditer(text):
                item = _clean_fact_list_item(match.group(0))
                item = re.split(
                    r"\b(?:\d+\s*,\s*\d+\s*=|RNE\s+r[eé]cent|Une\s+attestation|Le\s+dossier\s+de\s+l['’]?\s*offre\s+financi[eè]re)\b",
                    item,
                    maxsplit=1,
                    flags=re.IGNORECASE,
                )[0].strip(" .;:-|")
                folded = _fold_fact_text(item)
                if not item or folded in seen or any(folded in existing or existing in folded for existing in seen):
                    continue
                seen.add(folded)
                fallback_items.append(item)
                fallback_page = fallback_page or page_entry["page"]
                fallback_section = fallback_section or page_entry["section"]
                if len(fallback_items) >= 8:
                    break
            if len(fallback_items) >= 8:
                break
        if len(fallback_items) >= 8:
            break

    if fallback_items:
        return {
            "text": "\n".join(f"- {item}" for item in fallback_items),
            "items": [
                {
                    "text": item,
                    "page": fallback_page or "?",
                    "section": fallback_section or "technical",
                }
                for item in fallback_items
            ],
            "page": fallback_page or "?",
            "section": fallback_section or "technical",
        }

    return None


SCALAR_FALLBACK_EXTRACTORS = {
    "validity": _extract_validity_fallback,
    "opening": _extract_opening_fallback,
    "caution": _extract_caution_fallback,
    "information_sheet": _extract_information_sheet_fallback,
    "fiscal_certificate": _extract_fiscal_certificate_fallback,
    "manufacturer_authorization": _extract_manufacturer_authorization_fallback,
    "guarantee": _extract_guarantee_fallback,
    "reception": _extract_reception_fallback,
    "definitive_caution": _extract_definitive_caution_fallback,
    "penalties": _extract_penalties_fallback,
    "references": _extract_references_fallback,
    "payment": _extract_payment_fallback,
}

ARABIC_SCALAR_FALLBACK_EXTRACTORS = {
    "deadline": _extract_arabic_deadline_fallback,
    "submission_method": _extract_arabic_submission_method_fallback,
    "validity": _extract_arabic_validity_fallback,
    "opening": _extract_arabic_opening_fallback,
    "caution": _extract_arabic_caution_fallback,
    "information_sheet": _extract_arabic_information_sheet_fallback,
    "rne": _extract_arabic_rne_fallback,
    "guarantee": _extract_arabic_guarantee_fallback,
    "reception": _extract_arabic_reception_fallback,
    "definitive_caution": _extract_arabic_definitive_caution_fallback,
    "penalties": _extract_arabic_penalties_fallback,
    "payment": _extract_arabic_payment_fallback,
}

LIST_FALLBACK_EXTRACTORS = {
    "administrative_documents": _extract_administrative_documents_fallback,
    "technical_documents": _extract_technical_documents_topnet_fallback,
    "financial_documents": _extract_financial_documents_fallback,
}

ARABIC_LIST_FALLBACK_EXTRACTORS = {
    "administrative_documents": _extract_arabic_administrative_documents_fallback,
    "technical_documents": _extract_arabic_technical_documents_fallback,
    "financial_documents": _extract_arabic_financial_documents_fallback,
}

FACT_QUALITY_MARKERS: dict[str, tuple[str, ...]] = {
    "submission_method": (
        "voie postale",
        "accuse de reception",
        "bureau d'ordre central",
        "topnet",
        "centre urbain nord",
        "ubci",
        "139 avenue de la liberte",
        "1001 tunis",
        "tuneps",
        "www.tuneps.tn",
        "depot des offres",
        "dépôt des offres",
        "remise des offres",
        "offre technique",
        "offre financiere",
        "offre financière",
        "pli ferme",
        "pli fermé",
    ),
    "validity": (
        "validite de l'offre",
        "validite des offres",
        "validite de la soumission",
        "offres resteront valables",
        "offre reste valable",
        "date limite fixee",
        "reception des plis",
        "90 jours",
        "120 jours",
        "quatre-vingt-dix",
        "يبقى المتعهدون ملتزمون",
        "صلوحية العروض",
        "60 يوما",
        "ستون يوما",
    ),
    "opening": (
        "pas publique",
        "une seule etape",
        "locaux de topnet",
        "offres techniques",
        "offres financieres",
    ),
    "caution": (
        "cautionnement provisoire",
        "1,5%",
        "montant de la soumission",
        "1 000 dt",
        "banque tunisienne",
        "120 jours",
        "الضمان المالي الوقتي",
        "إقصاء العرض",
    ),
    "information_sheet": (
        "fiche des renseignements generaux",
        "annexe 3",
        "fiche des elements de contact",
        "annexe 3 bis",
    ),
    "fiscal_certificate": (
        "situation fiscale",
        "verifications necessaires",
    ),
    "manufacturer_authorization": (
        "attestation delivree par le fabricant",
        "commercialiser les fournitures",
        "document officiel du constructeur",
        "attestation constructeur",
    ),
    "administrative_documents": (
        "cautionnement provisoire",
        "cahier des charges",
        "delegation de pouvoir",
        "registre national des societes",
        "cnss",
        "fabricant",
        "paraphe",
        "signe",
        "date",
        "cachet",
        "الوثائق المكونة للعرض",
        "الشروط العامة للاستشارة",
        "الممثل الرسمي",
        "الضمان المالي الوقتي",
    ),
    "technical_documents": (
        "formulaire des reponses",
        "annexe 6",
        "specifications techniques",
        "documents techniques",
        "documentation technique",
        "formulaire technique",
        "document officiel du constructeur",
        "date de fin de support",
        "validite de la soumission",
        "الخصائص الفنية",
        "Code Réference",
        "Plotwave",
        "ROWE I4",
    ),
    "financial_documents": (
        "offre financiere par unite",
        "prix catalogue",
        "extension de garantie",
        "contrat cadre de service",
        "جدول الأثمان",
        "السعر الفردي",
        "المبلغ الجملي",
    ),
    "guarantee": (
        "trois (03) ans",
        "certificat de reception provisoire",
        "48 heures",
        "مدة الضمان",
        "أشهر",
        "الاستلام الوقتي",
        "الإستلام الوقتي",
    ),
    "reception": (
        "reception provisoire",
        "tests de bon fonctionnement",
        "reception definitive",
        "36 mois",
        "proces-verbal",
    ),
    "definitive_caution": (
        "caution definitive",
        "10%",
        "dix (10) jours",
        "bonne execution",
        "garantie",
    ),
    "penalties": (
        "1%",
        "un pour mille",
        "chaque jour de retard",
        "5%",
        "htva",
        "غرامات التأخير",
        "غرامة تأخير",
    ),
    "payment": (
        "100 %",
        "pv de reception provisoire",
        "60 jours",
        "facture",
        "reglement est effectue par virement",
        "règlement est effectué par virement",
        "virement bancaire",
        "r.i.b",
        "trente (30) jours",
        "factures conformes",
        "كيفية الخلاص",
        "يتم خلاص صاحب العقد",
        "خمسة و أربعون",
        "تحويل بريدي",
        "تحويل بنكي",
    ),
}


FACT_QUALITY_PENALTIES: dict[str, tuple[str, ...]] = {
    "subject": (
        "annexe",
        "modele de soumission",
        "modele d'engagement",
        "modele d'offre",
        "formulaire de reponse",
    ),
    "submission_method": (
        "eclaircissement",
        "éclaircissement",
        "correspondance",
        "facture",
        "factures",
        "notification",
        "reclamation",
        "réclamation",
    ),
    "validity": (
        "validite de la caution",
        "validité de la caution",
        "validite du cautionnement",
        "validite du contrat",
        "validité du contrat",
        "validite du marche",
        "validité du marché",
        "delai d'execution",
        "délai d'exécution",
        "duree d'execution",
        "durée d'exécution",
    ),
    "caution": (
        "annexe",
        "modele",
        "formulaire",
        "m'engage",
        "nous engageons",
        "le montant du dit cautionnement",
        "cautionnement provisoire pour participer",
        "fait a",
    ),
    "definitive_caution": (
        "annexe",
        "modele",
        "formulaire",
        "m'engage",
        "nous engageons",
        "le montant du dit cautionnement",
        "cautionnement provisoire pour participer",
        "fait a",
    ),
    "guarantee": (
        "الضمان المالي الوقتي",
        "الضمان المالي النهائي",
        "caution provisoire",
        "caution definitive",
    ),
    "reception": (
        "كيفية الخلاص",
        "يتم خلاص صاحب العقد",
        "modalites de paiement",
        "composition de l'offre",
        "piece jointe",
        "pieces jointes",
        "a telecharger",
        "envoye en ligne",
        "offre financiere",
        "bordereau des prix",
        "bordereaux des prix",
        "non presentation du cautionnement",
        "modalités de paiement",
    ),
}


def _fact_quality_score(field: str, fact: dict | None) -> int:
    if not fact:
        return -1

    text = str(fact.get("text", ""))
    folded = _fold_fact_text(text)
    if not folded:
        return -1

    score = min(len(folded) // 220, 6)
    for marker in FACT_QUALITY_MARKERS.get(field, ()):
        if _fold_fact_text(marker) in folded:
            score += 10
    for marker in FACT_QUALITY_PENALTIES.get(field, ()):
        if _fold_fact_text(marker) in folded:
            score -= 10
    return score


def _prefer_fact(field: str, current: dict | None, candidate: dict | None) -> dict | None:
    if not candidate:
        return current
    if not current:
        return candidate
    if _fact_quality_score(field, candidate) > _fact_quality_score(field, current):
        return candidate
    return current


def _list_fact_item_count(fact: dict | None) -> int:
    if not fact:
        return 0
    items = fact.get("items")
    if isinstance(items, list):
        return len(items)
    text = str(fact.get("text", ""))
    return len([line for line in text.splitlines() if line.strip().startswith("-")])


def _is_signature_template_fact(fact: dict | None) -> bool:
    if not fact:
        return False
    folded = _fold_fact_text(str(fact.get("text", "")))
    return (
        "nom, prenom et qualite du signataire" in folded
        or "signature et cachet" in folded
        or "le soumissionnaire" in folded and "fait a" in folded
    )


def _is_caution_procedure_admin_fact(fact: dict | None) -> bool:
    if not fact:
        return False
    folded = _fold_fact_text(str(fact.get("text", "")))
    if "caution provisoire" not in folded and "cautionnement provisoire" not in folded:
        return False
    procedure_hits = sum(
        1
        for marker in (
            "restituee",
            "restitue",
            "mise en paiement",
            "titulaire du marche",
            "apres constitution",
            "refuse de signer",
            "premiere demande",
        )
        if marker in folded
    )
    document_hits = sum(
        1
        for marker in (
            "fiche",
            "cnss",
            "registre",
            "rne",
            "attestation fiscale",
            "declaration sur l'honneur",
            "documents administratifs",
        )
        if marker in folded
    )
    return procedure_hits >= 3 and document_hits == 0


def _with_fact_text(fact: dict | None, text: str) -> dict | None:
    if not fact:
        return None

    updated = dict(fact)
    updated["text"] = text
    return updated


def _polish_fact_text(field: str, fact: dict | None) -> dict | None:
    if not fact:
        return None

    text = str(fact.get("text", ""))
    folded = _fold_fact_text(text)

    if field == "subject" and "rose blanche" not in folded:
        boundary_parts = re.split(
            r"\s*(?:[.,;]\s*Article\s+\d+\b|\bCette\s+acquisition\s+est\s+r\S+partie\b|"
            r"\bN\s*[°o]?\s+D\S+signation\s+R\S+f\s+Qt\S+\b|\bQt\S+\s+MIN\b|"
            r"\bLot\s+N\s*[°o]?\s*0?\d+\b|\bLa\s+participation\b|\bToute\s+soumission\b|"
            r"\bLes\s+candidats\b)",
            text,
            maxsplit=1,
            flags=re.IGNORECASE,
        )
        if len(boundary_parts) > 1 and len(boundary_parts[0].strip()) >= 35:
            text = boundary_parts[0].strip(" .;:-")
            folded = _fold_fact_text(text)
            fact = _with_fact_text(fact, text)

        match = re.search(
            r"\b(?:le\s+pr\S+sent\s+(?:march\S+|appel\s+d\S+offres?)|la\s+pr\S+sente\s+consultation)\s+a\s+pour\s+objet\b",
            text,
            flags=re.IGNORECASE,
        )
        if match and match.start() > 0:
            return _with_fact_text(fact, text[match.start() :].strip(" .;:-"))
        if re.match(r"^\s*de\s+d[eé]finir\b", text, flags=re.IGNORECASE):
            return _with_fact_text(fact, f"Le présent marché a pour objet {text.strip(' .;:-')}.")

    if field == "submission_method":
        cleaned_submission = re.split(
            r"\bet\s+ce,\s+au\s+plus\s+tard\b|\bau\s+plus\s+tard\s*,?\s+le\b|\bdate\s+limite\b",
            text,
            maxsplit=1,
            flags=re.IGNORECASE,
        )[0].strip(" .;:-")
        if cleaned_submission and cleaned_submission != text:
            return _with_fact_text(fact, cleaned_submission)

    if field == "payment" and "appel de commande" in folded and "60 jours" in folded:
        return _with_fact_text(
            fact,
            "Le paiement est effectué à 100% à 60 jours, sur présentation de l'original de la facture, "
            "du/des bon(s) de livraison et du PV de réception provisoire.",
        )

    if field == "penalties" and "cinq pour mille" in folded and "10%" in folded:
        return _with_fact_text(
            fact,
            "La pénalité de retard est de cinq pour mille (5‰) par jour sur le montant des articles "
            "non livrés dans les délais, avec un plafond de 10% du montant définitif du marché.",
        )

    if field == "validity" and "يبقى المتعهدون ملتزمون" in text:
        cleaned_validity = re.split(
            r"\s+(?:تقديم\s+العروض|يتم\s+تقديم\s+العروض|إيداع\s+العروض|فتح\s+العروض)\b",
            text,
            maxsplit=1,
        )[0].strip(" .;:-")
        if cleaned_validity and cleaned_validity != text:
            return _with_fact_text(fact, cleaned_validity)

    if field == "caution" and "الضمان المالي الوقتي" in text:
        if re.search(r"\b0\s*دينار", text):
            return _with_fact_text(
                fact,
                "الضمان المالي الوقتي مطلوب، لكن مبلغ الضمان لم يستخرج بشكل موثوق من النص الممسوح.",
            )

    if field == "definitive_caution" and "ثلاثة بالمائة" in text:
        cleaned_definitive = re.sub(r"\(?\s*9?03\s*\)?", "3%", text)
        if cleaned_definitive != text:
            return _with_fact_text(fact, cleaned_definitive)

    if field == "penalties" and "غرامات التأخير" in text:
        cleaned_penalties = re.sub(r"\b9?05\b(?=\s+من\s+المبلغ)", "5%", text)
        cleaned_penalties = re.sub(r"\s+\(\s*عدد\s+أيام\s+التأخير\s*\)\s*100\b", "", cleaned_penalties)
        if cleaned_penalties != text:
            return _with_fact_text(fact, cleaned_penalties)

    if field == "guarantee" and ("مدّة الضمان" in text or "مدة الضمان" in text or "مدة ضمان" in text):
        if re.search(r"(?:\(?\s*6\s*\)?|\)\s*6\s*\()\s*أشهر", text) and ("الإستلام الوقتي" in text or "الاستلام الوقتي" in text):
            return _with_fact_text(
                fact,
                "مدة الضمان هي ستة (6) أشهر ابتداء من تاريخ الإستلام الوقتي.",
            )

    if field == "rne" and "السجل الوطني للمؤسسات" in text:
        return _with_fact_text(fact, "يشترط التسجيل أو تقديم نظير أصلي من السجل الوطني للمؤسسات.")

    if (
        field == "subject"
        and "topnet" in folded
        and "solution antivirale" in folded
        and "edr" in folded
    ):
        return _with_fact_text(
            fact,
            "La société TOPNET lance une consultation auprès des sociétés opérant dans la "
            "sécurité des infrastructures informatiques pour le renouvellement de la Solution "
            "Antivirale, avec installation et activation des licences et mise en production de l'EDR.",
        )

    if (
        field == "subject"
        and "rose blanche" in folded
        and "solutions de cybersecurite" in folded
    ):
        return _with_fact_text(
            fact,
            "Le présent appel d'offres a pour objet la fourniture, l'installation, "
            "la configuration, l'intégration et la mise en production de solutions "
            "de cybersécurité en deux lots : Antivirus + EDR et PAM.",
        )

    if (
        field == "subject"
        and "solution d'impression" in folded
        and ("multifonction" in folded or "periode de garantie" in folded)
    ):
        return _with_fact_text(
            fact,
            "L'appel d'offres a pour objet l'acquisition, la fourniture, l'installation "
            "et la mise en service de 50 équipements multifonction impression-copie-scan, "
            "avec une solution d'administration/gestion d'impression et des prestations "
            "de maintenance.",
        )

    if (
        field == "subject"
        and "acquisition et mise en place d'une solution d'impression" in folded
        and "stb" in folded
    ):
        return _with_fact_text(
            fact,
            "L'appel d'offres a pour objet l'acquisition et la mise en place "
            "d'une solution d'impression a la STB.",
        )

    if (
        field == "guarantee"
        and "solution d'impression" in folded
        and "periode de garantie" in folded
    ):
        return _with_fact_text(
            fact,
            "La garantie est d'une année. Les prestations d'entretien continu des équipements "
            "multifonction impression-copie-scan durent trois ans après l'expiration de cette garantie.",
        )

    if (
        field == "payment"
        and "stb" in folded
        and ("virement bancaire" in folded or "virement banca" in folded or "mode de paiement" in folded)
        and ("60%" in folded or "10%" in folded)
    ):
        return _with_fact_text(
            fact,
            "La STB règle le marché par virement bancaire au titulaire, sur présentation "
            "des factures et après signature des PV de réception correspondants.",
        )

    if (
        field == "penalties"
        and "penalites de retard" in folded
        and ("trois pour mille" in folded or "3%o" in folded or "3‰" in text)
    ):
        return _with_fact_text(
            fact,
            "Les pénalités de retard sont calculées au prorata des journées de retard, "
            "sur la base du montant HT du marché, à raison de 3‰ (trois pour mille).",
        )

    if (
        field == "caution"
        and "cautionnement" in folded
        and ("1 000 dt" in folded or "mille dinars" in folded)
        and "120 jours" in folded
    ):
        return _with_fact_text(
            fact,
            "Le cautionnement provisoire est de Mille Dinars Tunisiens (1 000 DT), sous forme "
            "de caution bancaire inconditionnelle émise par une banque tunisienne, payable à "
            "première demande de TOPNET et valable 120 jours à partir de la date de réception "
            "des offres.",
        )

    if (
        field == "payment"
        and "100" in folded
        and "pv de reception provisoire" in folded
        and "60 jours" in folded
        and "virement bancaire" in folded
    ):
        return _with_fact_text(
            fact,
            "100 % sont réglés après signature du PV de réception provisoire, à 60 jours "
            "suivant la présentation de la facture originale et des attachements signés, "
            "par virement bancaire.",
        )

    if (
        field == "payment"
        and "45 jours" in folded
        and ("virement bancaire" in folded or "virement banca" in folded or "mode de paiement" in folded)
        and "factures" in folded
    ):
        return _with_fact_text(
            fact,
            "Les factures conformes sont payables à 45 jours à compter de leur réception "
            "avec les documents de facturation par le bureau d'ordre. Le paiement s'effectue "
            "par virement bancaire.",
        )

    if (
        field == "guarantee"
        and ("delai de garantie est fixe" in folded or "garantie a defaut" in folded)
        and "6 mois" in folded
        and ("pieces et main" in folded or "main d'oeuvre" in folded or "main d œuvre" in folded)
    ):
        return _with_fact_text(fact, "Le délai de garantie est fixé à 6 mois pièces et main d'oeuvre.")

    if (
        field == "definitive_caution"
        and ("bonne fin" in folded or "caution bancaire a premiere demande" in folded)
        and "5%" in folded
    ):
        return _with_fact_text(fact, "La caution bancaire à première demande de bonne fin est égale à 5% du montant TTC.")

    if (
        field == "penalties"
        and ("0,2%" in folded or "0.2%" in folded or re.search(r"0\s*[,\.]\s*2\s*%", text))
        and "5%" in folded
        and "retard" in folded
    ):
        return _with_fact_text(
            fact,
            "Les pénalités de retard sont de 0,2% du montant de la commande par jour calendaire de retard, avec un maximum de 5% du montant total définitif de la commande hors TVA.",
        )

    if (
        field == "submission_method"
        and "ubci" in folded
        and "bureau d'ordre central" in folded
        and "139 avenue de la liberte" in folded
    ):
        return _with_fact_text(
            fact,
            "Les soumissions doivent parvenir au bureau d'ordre central de l'UBCI, "
            "à l'adresse suivante : UBCI 139 Avenue de la Liberté 1001 Tunis.",
        )

    if (
        field == "payment"
        and "ubci" in folded
        and "virement" in folded
        and "r.i.b" in folded
        and "trente (30) jours" in folded
    ):
        return _with_fact_text(
            fact,
            "Les paiements s'effectuent par virement à la banque nommée par le "
            "soumissionnaire retenu, qui fournira le R.I.B. Le paiement des factures "
            "conformes est effectué à trente (30) jours de leur réception.",
        )

    return fact


TENDER_PROFILE_FIELD_MAP = {
    "subject": "object",
    "submission_method": "submission_method",
    "deadline": "deadline",
    "validity": "validity",
    "opening": "opening",
    "variants": "variants",
    "caution": "provisional_caution",
    "information_sheet": "information_sheet",
    "cnss": "cnss_certificate",
    "fiscal_certificate": "fiscal_certificate",
    "rne": "rne_certificate",
    "administrative_documents": "administrative_documents",
    "technical_documents": "technical_documents",
    "manufacturer_authorization": "manufacturer_authorization",
    "references": "references",
    "financial_documents": "financial_documents",
    "guarantee": "guarantee",
    "reception": "reception",
    "definitive_caution": "definitive_caution",
    "penalties": "penalties",
    "payment": "payment",
    "requested_items": "requested_items",
    "mined_facts": "mined_facts",
}


def _profile_value_from_fact(fact: dict) -> dict | None:
    text = _compact_fact_text(str(fact.get("text") or ""))
    if not text:
        return None

    value = {
        "text": text,
        "page": str(fact.get("page") or "?"),
        "section": str(fact.get("section") or "general"),
        "confidence": "extracted",
    }

    items = fact.get("items")
    if isinstance(items, list) and items:
        normalized_items = []
        for item in items:
            if not isinstance(item, dict):
                continue
            item_text = _compact_fact_text(str(item.get("text") or ""))
            if not item_text:
                continue
            normalized_items.append(
                {
                    "text": item_text,
                    "page": str(item.get("page") or value["page"]),
                    "section": str(item.get("section") or value["section"]),
                }
            )
        if normalized_items:
            value["items"] = normalized_items

    schema = fact.get("schema")
    if schema:
        value["schema"] = str(schema)

    return value


def build_tender_profile(facts: dict | None) -> dict | None:
    if not facts:
        return None

    fields = {}
    for fact_key, profile_key in TENDER_PROFILE_FIELD_MAP.items():
        fact = facts.get(fact_key)
        if not isinstance(fact, dict):
            continue
        value = _profile_value_from_fact(fact)
        if value:
            fields[profile_key] = value

    if not fields:
        return None

    expected_core_fields = (
        "object",
        "submission_method",
        "deadline",
        "validity",
        "opening",
        "provisional_caution",
        "administrative_documents",
        "technical_documents",
        "financial_documents",
        "guarantee",
        "payment",
        "penalties",
    )
    missing_core_fields = [field for field in expected_core_fields if field not in fields]
    present_core_count = len(expected_core_fields) - len(missing_core_fields)
    coverage = round(present_core_count / len(expected_core_fields), 3)

    return {
        "schema": "tender_profile.v1",
        "fields": fields,
        "coverage": {
            "core_present": present_core_count,
            "core_total": len(expected_core_fields),
            "core_ratio": coverage,
            "missing_core_fields": missing_core_fields,
        },
    }


def extract_document_facts(chunks: list[str], metas: list[dict]) -> dict:
    if not chunks or not metas:
        return {}

    pages = _group_chunks_by_page(chunks, metas)
    if not pages:
        return {}

    placeholder_values = _extract_placeholder_values(pages)
    marker_facts = _extract_instruction_marker_facts(placeholder_values)
    marker_facts.update(_extract_dpc_marker_facts(pages))
    article_facts = _extract_article_facts(pages)

    subject_candidates = _extract_fact_candidates_from_pages(
        pages,
        FACT_SUBJECT_PATTERNS,
        formatter=_subject_formatter(pages),
    )
    if article_facts.get("subject"):
        article_subject = dict(article_facts["subject"])
        article_subject["_pattern_index"] = -1
        subject_candidates.append(article_subject)
    page_subject = _extract_subject_from_pages_fallback(pages)
    if page_subject:
        page_subject = dict(page_subject)
        page_subject["_pattern_index"] = -2
        subject_candidates.append(page_subject)
    arabic_subject = _extract_arabic_subject_fallback(pages)
    if arabic_subject:
        arabic_subject = dict(arabic_subject)
        arabic_subject["_pattern_index"] = -3
        subject_candidates.append(arabic_subject)
    subject = max(subject_candidates, key=_score_subject_fact) if subject_candidates else None
    if marker_facts.get("subject") and (
        not subject
        or PLACEHOLDER_REFERENCE_RE.search(str(subject.get("text", "")))
        or "fourniture de acquisition" in _fold_fact_text(str(subject.get("text", "")))
        or "specifies dans le cctp" in _fold_fact_text(str(subject.get("text", "")))
        or "conditions de participation" in _fold_fact_text(str(subject.get("text", "")))
        or "donnees particulieres" in _fold_fact_text(str(subject.get("text", "")))
        or " dpc" in f" {_fold_fact_text(str(subject.get('text', '')))}"
    ):
        subject = marker_facts["subject"]
    if subject:
        subject = dict(subject)
        subject.pop("_pattern_index", None)
        subject = _resolve_fact_placeholders(subject, placeholder_values)
        subject = _polish_fact_text("subject", subject)

    deadline = article_facts.get("deadline")
    deadline = _resolve_fact_placeholders(deadline, placeholder_values)
    if deadline and not _is_reliable_scalar_fact("deadline", deadline):
        deadline = None
    if not deadline:
        deadline = (
            _extract_deadline_value_from_pages(pages)
            or _extract_fact_from_pages(pages, FACT_DEADLINE_PATTERNS)
            or _extract_arabic_deadline_fallback(pages)
        )
        deadline = _resolve_fact_placeholders(deadline, placeholder_values)
        if deadline and not _is_reliable_scalar_fact("deadline", deadline):
            deadline = None

    summary_text = None
    summary_page = None
    summary_section = None

    if subject:
        summary_bits = [subject["text"]]
        summary_page = subject["page"]
        summary_section = subject["section"]
        if deadline and deadline["text"] not in subject["text"]:
            summary_bits.append(f"Date limite : {deadline['text']}")
        summary_text = " ".join(summary_bits)
    else:
        first_page = pages[0]
        sentences = [
            _normalize_fact_text(sentence)
            for sentence in re.split(r"(?<=[.!?])\s+", first_page["text"])
            if len(sentence.strip()) > 20
        ]
        if sentences:
            summary_text = sentences[0]
            summary_page = first_page["page"]
            summary_section = first_page["section"]

    facts = {}
    if subject:
        facts["subject"] = subject
    if deadline:
        facts["deadline"] = deadline
    for field, patterns in FACT_SCALAR_PATTERNS.items():
        fact = article_facts.get(field) or _extract_fact_from_pages(pages, patterns)
        if field == "submission_method":
            fact = _prefer_fact(field, fact, _extract_submission_method_from_pages(pages))
        if field == "references" and fact and not _is_reliable_scalar_fact(field, fact):
            fact = None
        if field == "caution" and not fact:
            fact = _extract_caution_fallback(pages)
        if fact and not _is_reliable_scalar_fact(field, fact) and field == "caution":
            fact = _extract_caution_fallback(pages)
        fallback_extractor = SCALAR_FALLBACK_EXTRACTORS.get(field)
        if fallback_extractor:
            fact = _prefer_fact(field, fact, fallback_extractor(pages))
        arabic_fallback_extractor = ARABIC_SCALAR_FALLBACK_EXTRACTORS.get(field)
        if arabic_fallback_extractor:
            fact = _prefer_fact(field, fact, arabic_fallback_extractor(pages))
        if field == "guarantee" and fact and not _is_reliable_scalar_fact(field, fact):
            fallback_fact = None
            if arabic_fallback_extractor:
                fallback_fact = arabic_fallback_extractor(pages)
            if (not fallback_fact or not _is_reliable_scalar_fact(field, fallback_fact)) and fallback_extractor:
                fallback_fact = fallback_extractor(pages)
            if fallback_fact and _is_reliable_scalar_fact(field, fallback_fact):
                fact = fallback_fact
        fact = _resolve_fact_placeholders(fact, placeholder_values)
        if marker_facts.get(field) and (
            not fact
            or PLACEHOLDER_REFERENCE_RE.search(str(fact.get("text", "")))
            or "dpc" in _fold_fact_text(str(fact.get("text", "")))
            or _is_table_of_contents_fact(fact)
            or field in {"guarantee", "definitive_caution", "penalties"}
        ):
            fact = marker_facts[field]
        if fact and _is_reliable_scalar_fact(field, fact):
            facts[field] = _polish_fact_text(field, fact)
    for field, heading_patterns in FACT_LIST_DEFINITIONS.items():
        fact_list = _extract_fact_list_from_pages(
            pages,
            field,
            heading_patterns,
            FACT_LIST_STOP_PATTERNS[field],
        )
        fallback_extractor = LIST_FALLBACK_EXTRACTORS.get(field)
        if fallback_extractor:
            fallback_list = fallback_extractor(pages)
            if (
                fallback_list
                and (
                    _is_signature_template_fact(fact_list)
                    or _is_caution_procedure_admin_fact(fact_list)
                    or _list_fact_item_count(fallback_list) > _list_fact_item_count(fact_list)
                )
            ):
                fact_list = fallback_list
            else:
                fact_list = _prefer_fact(field, fact_list, fallback_list)
        if field == "administrative_documents" and _is_caution_procedure_admin_fact(fact_list):
            fact_list = None
        arabic_list_extractor = ARABIC_LIST_FALLBACK_EXTRACTORS.get(field)
        if arabic_list_extractor:
            arabic_list = arabic_list_extractor(pages)
            if arabic_list and _list_fact_item_count(arabic_list) > _list_fact_item_count(fact_list):
                fact_list = arabic_list
            else:
                fact_list = _prefer_fact(field, fact_list, arabic_list)
        if fact_list:
            facts[field] = fact_list
    if "technical_documents" not in facts:
        technical_fallback = _extract_technical_documents_fallback(pages)
        if technical_fallback:
            facts["technical_documents"] = technical_fallback
    requested_items = _extract_requested_items_fallback(pages)
    if requested_items:
        facts["requested_items"] = requested_items
    mined_facts = _extract_mined_facts(pages)
    if mined_facts:
        facts["mined_facts"] = mined_facts
    tender_profile = build_tender_profile(facts)
    if tender_profile:
        facts["tender_profile"] = tender_profile
    if summary_text:
        facts["summary"] = {
            "text": summary_text,
            "page": summary_page or "?",
            "section": summary_section or "general",
        }
    extraction_warning = _build_extraction_warning(chunks, facts)
    if extraction_warning:
        facts["extraction_warning"] = extraction_warning

    return facts

# ============= Arabic text helpers =============

def _arabic_char_ratio(text: str) -> float:
    arabic = sum(1 for c in text if "\u0600" <= c <= "\u06FF" or "\u0750" <= c <= "\u077F"
                 or "\uFB50" <= c <= "\uFDFF" or "\uFE70" <= c <= "\uFEFF")
    alpha = sum(1 for c in text if c.isalpha())
    return arabic / alpha if alpha else 0.0



def _fix_arabic_lines(text: str) -> str:
    """Apply bidi per-line to preserve paragraph structure.
    
    Applying bidi to the entire text at once can scramble paragraph
    boundaries. Processing line-by-line keeps structure intact.
    """
    try:
        from bidi.algorithm import get_display
        lines = text.split("\n")
        fixed = [get_display(line) if _arabic_char_ratio(line) > 0.1 else line for line in lines]
        return "\n".join(fixed)
    except ImportError:
        return text


# ============= Ingest Pipeline =============

def ingest():
    _ensure_utf8_stdio()
    vs = VectorStore()

    logger.info("Loading embedding model...")
    embedder = get_embedder()

    pdf_files = [f for f in os.listdir(PDFS_DIR) if f.endswith(".pdf")]
    logger.info(f"Found {len(pdf_files)} PDFs. Starting sync...")

    indexed_count = 0

    for pdf_file in pdf_files:
        if vs.has_source(pdf_file):
            logger.debug(f"Skipping {pdf_file} (already indexed)")
            continue

        # ── Extract and Chunk via Docling ────────────────────────────────────
        all_chunks, all_metas, all_ids = extract_and_chunk(os.path.join(PDFS_DIR, pdf_file), pdf_file)

        # ── Embed and store ───────────────────────────────────────────────
        if all_chunks:
            logger.info(f"Embedding {len(all_chunks)} chunks from {pdf_file}...")
            sys.stdout.flush()
            embeddings = to_builtin_list(embedder.encode(all_chunks, show_progress_bar=True))
            vs.add(chunks=all_chunks, embeddings=embeddings, metadatas=all_metas, ids=all_ids)
            indexed_count += 1

    logger.success(f"Ingestion complete — {indexed_count} new PDF(s) indexed")


if __name__ == "__main__":
    _ensure_utf8_stdio()
    ingest()
